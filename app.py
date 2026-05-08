from flask import Flask, request, redirect, session, send_file, jsonify
import psycopg2, os, qrcode, json, hashlib, hmac, base64, re, time
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4
from datetime import datetime, timedelta
from io import BytesIO
from cryptography.fernet import Fernet
import pyotp, urllib.request, urllib.parse

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "estudio_carlon_ultra_secure_2026_#$@!")
app.config["PROPAGATE_EXCEPTIONS"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
DB_URL = os.getenv("DB_URL")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
# Numero personal del admin - recibe alertas de seguridad
WHATSAPP_ADMIN    = os.getenv("WHATSAPP_ADMIN", "3855164943")
# Numero del estudio - usado para enviar mensajes a clientes
WHATSAPP_ESTUDIO  = os.getenv("WHATSAPP_ESTUDIO", "3843674464")
# Mantener por compatibilidad - apunta al admin
WHATSAPP_NUMERO   = WHATSAPP_ADMIN
# API keys separadas para cada numero
CALLMEBOT_APIKEY         = os.getenv("CALLMEBOT_APIKEY", "")          # admin (3855164943)
CALLMEBOT_APIKEY_ESTUDIO = os.getenv("CALLMEBOT_APIKEY_ESTUDIO", "2530568")  # estudio (3843674464)
ENCRYPT_KEY = os.getenv("ENCRYPT_KEY", "").encode() if os.getenv("ENCRYPT_KEY") else None

# ── Encriptación de datos sensibles ──────────────────────────────────────────
def get_fernet():
    k = ENCRYPT_KEY
    if not k:
        raw = (app.secret_key + "_enc_salt_carlon_2026").encode()
        k = base64.urlsafe_b64encode(hashlib.sha256(raw).digest())
    return Fernet(k)

def enc(val):
    if not val: return val
    try: return get_fernet().encrypt(str(val).encode()).decode()
    except: return val

def dec(val):
    if not val: return val
    try: return get_fernet().decrypt(val.encode()).decode()
    except: return val

# ── Envío de alertas WhatsApp via CallMeBot ───────────────────────────────────
def enviar_whatsapp(mensaje):
    """Alerta de seguridad al celular personal del admin"""
    try:
        apikey = CALLMEBOT_APIKEY
        if not apikey: return
        num = WHATSAPP_ADMIN.replace("+","").replace(" ","")
        msg_enc = urllib.parse.quote(mensaje)
        url = f"https://api.callmebot.com/whatsapp.php?phone=54{num}&text={msg_enc}&apikey={apikey}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        urllib.request.urlopen(req, timeout=5)
    except: pass

def enviar_whatsapp_estudio(numero_cliente, mensaje):
    """Envia WA a un cliente desde el numero del estudio"""
    try:
        apikey = CALLMEBOT_APIKEY_ESTUDIO
        if not apikey: return False
        num = str(numero_cliente).replace("+","").replace(" ","").replace("-","")
        if not num.startswith("54"): num = "54" + num
        msg_enc = urllib.parse.quote(mensaje)
        url = f"https://api.callmebot.com/whatsapp.php?phone={num}&text={msg_enc}&apikey={apikey}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        urllib.request.urlopen(req, timeout=6)
        return True
    except: return False

# ── Rate limiting & bloqueo de login ─────────────────────────────────────────
LOGIN_INTENTOS = {}   # {ip: {"count": n, "ts": timestamp, "blocked_until": ts}}
MAX_INTENTOS = 5
BLOQUEO_MINS = 30

def get_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()

def verificar_bloqueo_ip(ip):
    d = LOGIN_INTENTOS.get(ip, {})
    bu = d.get("blocked_until", 0)
    if bu and time.time() < bu:
        mins = int((bu - time.time()) / 60) + 1
        return True, mins
    return False, 0

def registrar_intento_fallido(ip):
    now = time.time()
    d = LOGIN_INTENTOS.get(ip, {"count": 0, "ts": now})
    if now - d.get("ts", now) > 3600:
        d = {"count": 0, "ts": now}
    d["count"] += 1
    d["ts"] = now
    if d["count"] >= MAX_INTENTOS:
        d["blocked_until"] = now + BLOQUEO_MINS * 60
        enviar_whatsapp(
            f"⚠️ ALERTA SEGURIDAD - Estudio Carlon\n"
            f"🔴 IP {ip} BLOQUEADA por {MAX_INTENTOS} intentos fallidos\n"
            f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"⏱ Bloqueada por {BLOQUEO_MINS} minutos"
        )
        registrar_evento_seguridad("BLOQUEO_IP", f"IP {ip} bloqueada tras {MAX_INTENTOS} intentos", ip)
    LOGIN_INTENTOS[ip] = d
    return d["count"]

def limpiar_intento(ip):
    if ip in LOGIN_INTENTOS:
        LOGIN_INTENTOS[ip] = {"count": 0, "ts": time.time()}

def desbloquear_ip_admin(ip):
    if ip in LOGIN_INTENTOS:
        LOGIN_INTENTOS[ip] = {"count": 0, "ts": time.time(), "blocked_until": 0}

# ── Bloqueo por país (IPs Argentina) ─────────────────────────────────────────
PAISES_PERMITIDOS = ["AR"]

def verificar_pais(ip):
    try:
        if ip in ("127.0.0.1", "::1", "localhost"): return True
        if ip.startswith("192.168.") or ip.startswith("10."): return True
        url = f"http://ip-api.com/json/{ip}?fields=countryCode"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=3)
        data = json.loads(resp.read())
        cc = data.get("countryCode", "AR")
        if cc not in PAISES_PERMITIDOS:
            registrar_evento_seguridad("ACCESO_PAIS_BLOQUEADO", f"IP {ip} - País {cc}", ip)
            enviar_whatsapp(f"🌎 ACCESO BLOQUEADO\nIP: {ip} - País: {cc}\nFecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            return False
        return True
    except: return True  # Si falla la consulta, permite (no bloquear por error de API)

# ── Registro de eventos de seguridad ─────────────────────────────────────────
def registrar_evento_seguridad(tipo, detalle, ip="", usuario=""):
    try:
        conn = conectar(); c = conn.cursor()
        c.execute("""INSERT INTO seguridad_eventos(tipo,detalle,ip,usuario,fecha)
                     VALUES(%s,%s,%s,%s,%s)""",
                  (tipo, detalle, ip, usuario or session.get("display",""), now_ar()))
        conn.commit(); conn.close()
    except: pass

MEDIOS_PAGO = ["Transferencia -> Natasha Carlon","Transferencia -> Maira Carlon",
               "Efectivo","Cheque","Dolares","Otro"]
# Categorías solo para admin (secretarias NO pueden ver ni cargar estas)
CATEGORIAS_ADMIN = ["Sueldo","Gastos Personales Natasha","Gastos Personales Maira","Tarjetas","Retiro Natasha","Retiro Maira","Otros Admin"]
# Categorías permitidas para secretarias
CATEGORIAS_SEC   = ["Luz","Internet","Agua","Gastos de Oficina","Artículos de Limpieza","Papelería","Otros"]
# Todas las categorías (admin ve todo)
CATEGORIAS_GASTO = CATEGORIAS_ADMIN + CATEGORIAS_SEC
VENCIMIENTOS_IMPOSITIVOS = [
    {"id":"ib_cat_a",   "nombre":"Ingresos Brutos - Cat. A", "dia":18,"tipo":"IIBB Sgo.","detalle":"Categoria A - vence el 18 de cada mes"},
    {"id":"ib_cat_b",   "nombre":"Ingresos Brutos - Cat. B", "dia":15,"tipo":"IIBB Sgo.","detalle":"Categoria B - vence el 15 de cada mes"},
    {"id":"iva_ddjj",   "nombre":"IVA - DJ (F.731)",          "dia":20,"tipo":"AFIP",     "detalle":"Formulario 731 - presentacion mensual"},
    {"id":"f931_1",     "nombre":"Aportes F.931 - 1ra quinc.","dia":9, "tipo":"AFIP",     "detalle":"Sueldos 1ra quincena - vence el 9"},
    {"id":"f931_2",     "nombre":"Aportes F.931 - 2da quinc.","dia":11,"tipo":"AFIP",     "detalle":"Sueldos 2da quincena - vence el 11"},
    {"id":"ganancias",  "nombre":"Ganancias - Anticipo",      "dia":23,"tipo":"AFIP",     "detalle":"Anticipo mensual segun cronograma AFIP"},
    {"id":"monotributo","nombre":"Monotributo - Cuota",       "dia":20,"tipo":"AFIP",     "detalle":"Cuota unificada mensual"},
    {"id":"suss_ddjj",  "nombre":"SUSS - DJ",                 "dia":9, "tipo":"AFIP",     "detalle":"Sistema Unico de Seguridad Social"},
    {"id":"rentas_prov","nombre":"Rentas Provinciales",       "dia":20,"tipo":"Rentas SGO","detalle":"Dir. Gral. de Rentas - Sgo. del Estero"},
]

CSS = """
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600&display=swap');
:root{--bg:#F7F5F0;--card:#fff;--primary:#1A3A2A;--accent:#C8A96E;--danger:#C0392B;--success:#27AE60;--warning:#E67E22;--info:#2475B0;--purple:#7B68EE;--text:#1C1C1C;--muted:#888;--border:#E4DDD0;--r:12px;--shadow:0 2px 18px rgba(0,0,0,0.07)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
nav{background:var(--primary);padding:0 28px;display:flex;align-items:center;justify-content:space-between;height:62px;position:sticky;top:0;z-index:200;box-shadow:0 2px 14px rgba(0,0,0,0.2)}
.brand{font-family:'DM Serif Display',serif;color:var(--accent);font-size:1.18rem;white-space:nowrap}
.nav-links{display:flex;align-items:center;gap:2px;flex-wrap:wrap}
.nav-links a{color:rgba(255,255,255,.72);text-decoration:none;padding:6px 11px;border-radius:7px;font-size:.82rem;font-weight:500;transition:all .18s}
.nav-links a:hover,.nav-links a.act{background:rgba(255,255,255,.11);color:#fff}
.nav-links a.logout{color:rgba(255,120,120,.85)}
.nav-links a.logout:hover{background:rgba(192,57,43,.22);color:#ffaaaa}
.user-pill{background:rgba(255,255,255,.1);border-radius:20px;padding:4px 12px;font-size:.76rem;color:rgba(255,255,255,.82);display:flex;align-items:center;gap:6px;white-space:nowrap}
.rbadge{font-size:.64rem;font-weight:700;padding:2px 7px;border-radius:10px;text-transform:uppercase;letter-spacing:.4px}
.rbadge.admin{background:var(--accent);color:#fff}.rbadge.sec{background:#2475B0;color:#fff}
.wrap{max-width:1180px;margin:0 auto;padding:32px 20px}
.page-title{font-family:'DM Serif Display',serif;font-size:1.9rem;color:var(--primary);margin-bottom:5px}
.page-sub{color:var(--muted);font-size:.84rem;margin-bottom:26px}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px;margin-bottom:24px}
.scard{background:var(--card);border-radius:var(--r);padding:18px 22px;box-shadow:var(--shadow);border-left:4px solid var(--accent);position:relative;overflow:hidden;transition:transform .18s}
.scard:hover{transform:translateY(-2px)}
.scard.g{border-left-color:var(--success)}.scard.r{border-left-color:var(--danger)}.scard.b{border-left-color:var(--info)}.scard.o{border-left-color:var(--warning)}.scard.p{border-left-color:var(--purple)}
.slabel{font-size:.68rem;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:var(--muted);margin-bottom:7px}
.sval{font-family:'DM Serif Display',serif;font-size:1.65rem;color:var(--primary)}
.sicon{position:absolute;right:14px;top:12px;font-size:1.8rem;opacity:.1}
.btn{display:inline-flex;align-items:center;gap:5px;padding:8px 17px;border-radius:8px;font-family:'DM Sans',sans-serif;font-weight:600;font-size:.84rem;cursor:pointer;border:none;text-decoration:none;transition:all .18s;white-space:nowrap}
.btn-p{background:var(--primary);color:#fff}.btn-p:hover{background:#254d38}
.btn-a{background:var(--accent);color:#fff}.btn-a:hover{background:#b8955a}
.btn-g{background:var(--success);color:#fff}.btn-g:hover{background:#1f9149}
.btn-r{background:var(--danger);color:#fff}.btn-r:hover{background:#a93226}
.btn-b{background:var(--info);color:#fff}.btn-b:hover{background:#1a5e8a}
.btn-o{background:transparent;border:1.5px solid var(--border);color:var(--text)}.btn-o:hover{border-color:var(--primary);color:var(--primary)}
.btn-wa{background:#25D366;color:#fff}.btn-wa:hover{background:#1ebe5d}
.btn-arca{background:#0055a5;color:#fff}.btn-arca:hover{background:#004080}
.btn-sm{padding:5px 10px;font-size:.77rem}.btn-xs{padding:3px 8px;font-size:.71rem}
.fcard{background:var(--card);border-radius:var(--r);padding:22px;box-shadow:var(--shadow);margin-bottom:22px}
.fcard h3{font-family:'DM Serif Display',serif;font-size:1.1rem;color:var(--primary);margin-bottom:15px;padding-bottom:10px;border-bottom:1px solid var(--border)}
.fgrid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:15px}
.fg{display:flex;flex-direction:column;gap:4px}
.fg label{font-size:.7rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.4px}
.fg input,.fg select,.fg textarea{padding:9px 11px;border:1.5px solid var(--border);border-radius:8px;font-family:'DM Sans',sans-serif;font-size:.88rem;background:var(--bg);transition:border-color .18s;outline:none}
.fg input:focus,.fg select:focus,.fg textarea:focus{border-color:var(--primary);background:#fff}
.search{display:flex;align-items:center;gap:9px;background:var(--card);border:1.5px solid var(--border);border-radius:10px;padding:8px 14px;margin-bottom:14px;box-shadow:var(--shadow)}
.search input{border:none;background:none;outline:none;font-family:'DM Sans',sans-serif;font-size:.9rem;width:100%}
.dtable{background:var(--card);border-radius:var(--r);box-shadow:var(--shadow);overflow:hidden;margin-bottom:22px}
.dtable table{width:100%;border-collapse:collapse}
.dtable thead tr{background:var(--primary)}
.dtable thead th{color:rgba(255,255,255,.72);font-size:.68rem;font-weight:600;letter-spacing:.8px;text-transform:uppercase;padding:11px 14px;text-align:left}
.dtable tbody tr{border-bottom:1px solid var(--border);transition:background .13s}
.dtable tbody tr:last-child{border-bottom:none}
.dtable tbody tr:hover{background:#f9f7f3}
.dtable td{padding:10px 14px;font-size:.85rem;vertical-align:middle}
.dtable td.nm{font-weight:600;color:var(--primary)}.dtable td.mu{color:var(--muted);font-size:.78rem}
.arow{background:var(--card);border-radius:var(--r);padding:14px 18px;box-shadow:var(--shadow);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:9px;margin-bottom:10px}
.period{font-family:'DM Serif Display',serif;font-size:1rem;color:var(--primary);min-width:78px}
.badge{display:inline-block;padding:3px 9px;border-radius:20px;font-size:.71rem;font-weight:700;letter-spacing:.3px}
.bp{background:#d5f5e3;color:#1a7a42}.bpar{background:#fef3cd;color:#9a6700}.bd{background:#fde8e8;color:#c0392b}
.badm{background:var(--accent);color:#fff}.bsec{background:#dce8ff;color:#1a4a8a}
.bmedio{background:#f0f4ff;color:#2475B0;font-size:.68rem;padding:2px 8px;border-radius:10px;font-weight:600}
.dcard{background:var(--card);border-radius:var(--r);padding:14px 18px;box-shadow:var(--shadow);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:9px;border-left:4px solid var(--danger);margin-bottom:9px}
.dname{font-weight:600;color:var(--primary)}.damt{font-family:'DM Serif Display',serif;font-size:1.15rem;color:var(--danger)}
.logrow{display:flex;gap:10px;padding:8px 0;border-bottom:1px solid var(--border);font-size:.8rem;align-items:flex-start}
.logrow:last-child{border-bottom:none}
.log-time{color:var(--muted);white-space:nowrap;min-width:120px;font-size:.72rem}
.log-user{font-weight:600;color:var(--primary);min-width:72px}
.log-msg{color:var(--text);flex:1}
.log-dot{width:7px;height:7px;border-radius:50%;background:var(--accent);margin-top:5px;flex-shrink:0}
.log-dot.red{background:var(--danger)}.log-dot.orange{background:var(--warning)}.log-dot.green{background:var(--success)}
.mo{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:300;align-items:center;justify-content:center}
.mo.on{display:flex}
.modal{background:var(--card);border-radius:16px;padding:26px;min-width:300px;max-width:500px;width:94%;box-shadow:0 20px 60px rgba(0,0,0,.22);animation:su .22s ease;max-height:90vh;overflow-y:auto}
@keyframes su{from{transform:translateY(26px);opacity:0}to{transform:none;opacity:1}}
.modal h3{font-family:'DM Serif Display',serif;font-size:1.25rem;color:var(--primary);margin-bottom:5px}
.modal .msub{color:var(--muted);font-size:.82rem;margin-bottom:16px}
.mact{display:flex;gap:9px;justify-content:flex-end;margin-top:16px}
.progwrap{background:var(--border);border-radius:6px;height:8px;overflow:hidden}
.progbar{height:100%;background:var(--success);border-radius:6px;transition:width .5s}
.chartrow{display:flex;align-items:center;gap:8px;margin-bottom:6px;font-size:.78rem}
.chartrow .cl{width:150px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-weight:500}
.chartrow .cbg{flex:1;background:var(--border);border-radius:4px;height:7px}
.chartrow .cfill{height:100%;background:var(--accent);border-radius:4px}
.chartrow .cv{width:82px;text-align:right;color:var(--muted);font-size:.74rem}
.ucard{background:var(--card);border-radius:var(--r);padding:14px 18px;box-shadow:var(--shadow);display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:9px;border-left:4px solid var(--info);margin-bottom:9px}
.ucard.adm{border-left-color:var(--accent)}
.flash{padding:10px 15px;border-radius:8px;margin-bottom:14px;font-size:.85rem;font-weight:500}
.fok{background:#d5f5e3;color:#1a7a42}.ferr{background:#fde8e8;color:#c0392b}
.lwrap{min-height:100vh;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,var(--primary) 0%,#0d1f14 100%)}
.lcard{background:var(--card);border-radius:20px;padding:38px 34px;width:360px;box-shadow:0 24px 80px rgba(0,0,0,.3)}
.ltitle{font-family:'DM Serif Display',serif;font-size:1.65rem;color:var(--primary);text-align:center;margin-bottom:4px}
.lsub{color:var(--muted);font-size:.8rem;text-align:center;margin-bottom:24px}
.denied{text-align:center;padding:80px 24px}
.denied .di{font-size:3.5rem;margin-bottom:14px}
.denied h2{font-family:'DM Serif Display',serif;font-size:1.75rem;color:var(--primary);margin-bottom:10px}
.denied p{color:var(--muted);margin-bottom:22px}
.tabs{display:flex;gap:3px;margin-bottom:18px;border-bottom:2px solid var(--border)}
.tab{padding:8px 16px;cursor:pointer;font-weight:600;font-size:.84rem;color:var(--muted);border-radius:8px 8px 0 0;transition:all .18s;border:none;background:none;font-family:'DM Sans',sans-serif}
.tab.on{color:var(--primary);border-bottom:2px solid var(--primary);margin-bottom:-2px;background:var(--card)}
.tabpanel{display:none}.tabpanel.on{display:block}
.qa{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:22px}
.info-box{background:#f0f9f4;border:1px solid #b7dfcc;border-radius:8px;padding:10px 14px;font-size:.81rem;color:#1a5c3a;margin-bottom:12px}
.warn-box{background:#fef9ec;border:1px solid #f0d080;border-radius:8px;padding:10px 14px;font-size:.81rem;color:#7a5800;margin-bottom:12px}
.sec-alert{background:#fde8e8;border:1px solid #f5a0a0;border-radius:8px;padding:10px 14px;font-size:.81rem;color:#7a1a1a;margin-bottom:8px}
.partner-card{background:var(--card);border-radius:var(--r);padding:18px 22px;box-shadow:var(--shadow);text-align:center}
.partner-name{font-family:'DM Serif Display',serif;font-size:1.05rem;color:var(--primary);margin-bottom:6px}
.partner-amt{font-family:'DM Serif Display',serif;font-size:1.9rem;margin-bottom:4px}
.chart-svg{width:100%;overflow:visible}
.caja-row{background:var(--card);border-radius:var(--r);padding:14px 18px;box-shadow:var(--shadow);margin-bottom:10px;border-left:4px solid var(--success)}
.caja-row.cerrada{border-left-color:var(--muted);opacity:.85}
.caja-header{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:10px}
.caja-user{font-weight:700;color:var(--primary);font-size:.96rem}
.caja-fecha{font-size:.76rem;color:var(--muted)}
.caja-medios{display:flex;gap:8px;flex-wrap:wrap}
.caja-item{display:flex;flex-direction:column;align-items:center;background:var(--bg);border-radius:8px;padding:8px 14px;min-width:90px;border:1px solid var(--border)}
.caja-item .ci-label{font-size:.68rem;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.4px;margin-bottom:3px}
.caja-item .ci-val{font-family:'DM Serif Display',serif;font-size:1.1rem;color:var(--primary)}
.estado-abierta{display:inline-flex;align-items:center;gap:5px;background:#d5f5e3;color:#1a7a42;padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:700}
.estado-cerrada{display:inline-flex;align-items:center;gap:5px;background:#f0f0f0;color:#666;padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:700}
.sec-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:10px;font-size:.68rem;font-weight:700}
.sec-badge.ok{background:#d5f5e3;color:#1a7a42}
.sec-badge.warn{background:#fef3cd;color:#9a6700}
.sec-badge.danger{background:#fde8e8;color:#c0392b}

/* ── App móvil supervisor ──────────────────────────────────────────── */
.app-wrap{max-width:480px;margin:0 auto;padding:16px 14px;min-height:100vh}
.app-header{text-align:center;padding:18px 0 10px;margin-bottom:6px}
.app-title{font-family:'DM Serif Display',serif;font-size:1.5rem;color:var(--primary);margin-bottom:3px}
.app-date{font-size:.78rem;color:var(--muted)}
.caja-card{background:var(--card);border-radius:16px;box-shadow:var(--shadow);padding:18px;margin-bottom:14px}
.caja-card h2{font-family:'DM Serif Display',serif;font-size:1.05rem;color:var(--primary);margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between}
.app-input-row{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px}
.app-field{display:flex;flex-direction:column;gap:5px}
.app-field label{font-size:.68rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}
.app-field input,.app-field select,.app-field textarea{padding:11px 12px;border:2px solid var(--border);border-radius:10px;font-family:'DM Sans',sans-serif;font-size:.95rem;background:var(--bg);outline:none;transition:border-color .18s;-webkit-appearance:none}
.app-field input:focus,.app-field select:focus{border-color:var(--primary);background:#fff}
.app-field input[type=number]{font-size:1.1rem;font-weight:600;color:var(--primary)}
.btn-app{width:100%;padding:14px;border-radius:12px;font-family:'DM Sans',sans-serif;font-size:1rem;font-weight:700;cursor:pointer;border:none;margin-top:6px;transition:all .18s;letter-spacing:.2px}
.btn-app-g{background:var(--success);color:#fff}.btn-app-g:active{background:#1f9149;transform:scale(.98)}
.btn-app-r{background:var(--danger);color:#fff}.btn-app-r:active{background:#a93226;transform:scale(.98)}
.btn-app-b{background:var(--info);color:#fff}.btn-app-b:active{background:#1a5e8a;transform:scale(.98)}
.btn-app-o{background:var(--accent);color:#fff}.btn-app-o:active{background:#b8955a;transform:scale(.98)}
.cierre-row{background:var(--card);border-radius:12px;padding:14px 16px;margin-bottom:10px;box-shadow:var(--shadow)}
.cierre-row.ok{border-left:4px solid var(--success)}.cierre-row.mal{border-left:4px solid var(--danger)}.cierre-row.pend{border-left:4px solid var(--warning)}
.cierre-nombre{font-weight:700;font-size:.96rem;color:var(--primary)}
.cierre-fecha{font-size:.72rem;color:var(--muted)}
.medios-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-top:10px}
.medio-chip{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:6px 8px;text-align:center}
.medio-chip .mc-lbl{font-size:.6rem;color:var(--muted);font-weight:700;text-transform:uppercase;letter-spacing:.4px;display:block}
.medio-chip .mc-val{font-family:'DM Serif Display',serif;font-size:.96rem;color:var(--primary);font-weight:600;display:block}
.diff-badge{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:20px;font-size:.75rem;font-weight:700;margin-top:8px}
.diff-ok{background:#d5f5e3;color:#1a7a42}.diff-mal{background:#fde8e8;color:#c0392b}.diff-pend{background:#fef3cd;color:#9a6700}
.tab-bar{position:fixed;bottom:0;left:0;right:0;background:var(--primary);display:flex;z-index:100;box-shadow:0 -2px 12px rgba(0,0,0,.15)}
.tab-item{flex:1;display:flex;flex-direction:column;align-items:center;padding:10px 4px 14px;cursor:pointer;color:rgba(255,255,255,.55);font-size:.62rem;font-weight:600;text-decoration:none;letter-spacing:.3px;transition:color .15s}
.tab-item.act{color:var(--accent)}.tab-item span{font-size:1.3rem;margin-bottom:2px}
.app-total-box{background:var(--primary);border-radius:14px;padding:16px 18px;color:#fff;margin-bottom:14px;text-align:center}
.app-total-box .atl{font-size:.72rem;opacity:.7;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.app-total-box .atv{font-family:'DM Serif Display',serif;font-size:2rem}
.app-flash{padding:12px 16px;border-radius:10px;margin-bottom:12px;font-size:.88rem;font-weight:600;text-align:center}
.app-fok{background:#d5f5e3;color:#1a7a42}.app-ferr{background:#fde8e8;color:#c0392b}
@media(max-width:680px){.stats{grid-template-columns:1fr 1fr}.arow{flex-direction:column;align-items:flex-start}nav .user-pill{display:none}.wrap{padding:18px 12px}.nav-links a{padding:5px 8px;font-size:.78rem}}
"""

ASISTENTE_JS = """
<div id="ai-btn" onclick="toggleChat()" style="position:fixed;bottom:24px;right:24px;z-index:1000;width:52px;height:52px;border-radius:50%;background:#1A3A2A;cursor:pointer;display:flex;align-items:center;justify-content:center;box-shadow:0 4px 16px rgba(0,0,0,0.25)">
  <span style="font-size:22px">&#x1F916;</span>
</div>
<div id="ai-panel" style="display:none;position:fixed;bottom:88px;right:24px;z-index:999;width:360px;max-width:calc(100vw - 32px);background:#fff;border-radius:16px;box-shadow:0 8px 40px rgba(0,0,0,0.18);overflow:hidden;flex-direction:column">
  <div style="background:#1A3A2A;padding:14px 16px;display:flex;align-items:center;justify-content:space-between">
    <div style="display:flex;align-items:center;gap:10px">
      <div style="width:36px;height:36px;border-radius:50%;background:#C8A96E;display:flex;align-items:center;justify-content:center;font-size:18px">&#x1F916;</div>
      <div>
        <div style="color:#fff;font-weight:600;font-size:.9rem">Asistente Estudio Carlon</div>
        <div style="color:rgba(255,255,255,.55);font-size:.72rem">Consultas del sistema y contables</div>
      </div>
    </div>
    <button onclick="toggleChat()" style="background:none;border:none;color:rgba(255,255,255,.7);cursor:pointer;font-size:18px;padding:4px">&#x2715;</button>
  </div>
  <div id="ai-msgs" style="height:300px;overflow-y:auto;padding:14px;display:flex;flex-direction:column;gap:10px;background:#F7F5F0">
    <div style="background:#fff;border-radius:4px 14px 14px 14px;padding:11px 13px;font-size:.83rem;color:#1C1C1C;max-width:90%;line-height:1.6;box-shadow:0 1px 4px rgba(0,0,0,0.06)">
      Hola! Puedo ayudarte con el sistema o con consultas contables. Que necesitas?
    </div>
  </div>
  <div style="padding:10px 12px;background:#fff;border-top:1px solid #E4DDD0;display:flex;gap:8px;align-items:flex-end">
    <textarea id="ai-input" placeholder="Escribi tu consulta..." rows="1"
      style="flex:1;border:1.5px solid #E4DDD0;border-radius:12px;padding:8px 12px;font-family:'DM Sans',sans-serif;font-size:.84rem;resize:none;outline:none;line-height:1.45;max-height:90px;overflow-y:auto;background:#F7F5F0;color:#1C1C1C"
      onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendAI()}"
      oninput="this.style.height='auto';this.style.height=this.scrollHeight+'px'"></textarea>
    <button id="ai-send" onclick="sendAI()" style="width:36px;height:36px;border-radius:50%;background:#1A3A2A;border:none;cursor:pointer;display:flex;align-items:center;justify-content:center;flex-shrink:0">
      <svg width="15" height="15" viewBox="0 0 24 24" fill="white"><path d="M2 21l21-9L2 3v7l15 2-15 2z"/></svg>
    </button>
  </div>
</div>
<script>
var aiAbierto=false,aiHistorial=[];
function addMsg(txt,tipo){
  var box=document.getElementById('ai-msgs');
  var d=document.createElement('div');
  d.className=tipo==='user'?'':'';
  d.style.cssText=tipo==='user'?'background:#1A3A2A;color:#fff;border-radius:14px 4px 14px 14px;padding:10px 13px;font-size:.83rem;max-width:90%;align-self:flex-end;margin-left:auto':'background:#fff;border-radius:4px 14px 14px 14px;padding:11px 13px;font-size:.83rem;color:#1C1C1C;max-width:90%;line-height:1.6;box-shadow:0 1px 4px rgba(0,0,0,.06);white-space:pre-wrap';
  d.textContent=txt;box.appendChild(d);box.scrollTop=box.scrollHeight;
}
function sendAI(){
  var inp=document.getElementById('ai-input');var q=inp.value.trim();if(!q)return;
  inp.value='';inp.style.height='auto';addMsg(q,'user');
  var btn=document.getElementById('ai-send');btn.style.opacity='0.4';btn.disabled=true;
  aiHistorial.push({role:'user',content:q});
  fetch('/asistente',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mensajes:aiHistorial})})
  .then(r=>r.json()).then(data=>{
    btn.style.opacity='1';btn.disabled=false;
    if(data.respuesta){addMsg(data.respuesta,'bot');aiHistorial.push({role:'assistant',content:data.respuesta});}
    else addMsg('No puedo responder eso ahora. Busca en AFIP o Google.','bot');
  }).catch(()=>{btn.style.opacity='1';btn.disabled=false;addMsg('Sin conexion.','bot');});
}
function toggleChat(){
  aiAbierto=!aiAbierto;var p=document.getElementById('ai-panel');
  p.style.display=aiAbierto?'flex':'none';
  if(aiAbierto)setTimeout(()=>document.getElementById('ai-input').focus(),100);
}
</script>
"""

def nav_html(active=""):
    user=session.get("user","");rol=session.get("rol","secretaria");disp=session.get("display",user)
    links_admin=[("/panel","Panel"),("/clientes","Clientes"),("/deudas","Deudores"),("/gastos","Gastos"),("/caja","Caja"),("/reportes","Reportes"),("/sueldos","Sueldos"),("/agenda","Agenda"),("/tareas","Tareas"),("/novedades","Novedades"),("/seguridad","Seguridad"),("/configuracion","Config")]
    links_sup=[("/app","📱 Mi App")]  # supervisor solo ve la app movil
    links_sec=[("/panel_sec","Inicio"),("/clientes","Clientes"),("/deudas","Deudores"),("/gastos","Gastos"),("/caja","Caja"),("/sueldos","Sueldos"),("/agenda","Agenda"),("/tareas","Tareas"),("/novedades","Novedades")]
    if rol=="admin": links=links_admin
    elif rol=="supervisor": links=links_sup
    else: links=links_sec
    items="".join(f'<a href="{h}" class="{"act" if active==l else ""}">{l}</a>' for h,l in links)
    items+='<a href="/logout" class="logout">Salir</a>'
    badge_txt="Admin" if rol=="admin" else ("Sup." if rol=="supervisor" else "Sec.")
    badge_cls="admin" if rol=="admin" else ("sec" if rol=="supervisor" else "sec")
    badge=f'<span class="rbadge {badge_cls}">{badge_txt}</span>'
    return f'<nav><span class="brand">&#x2726; Estudio Carlon</span><div class="nav-links">{items}</div><div class="user-pill">&#x1F464; {disp} {badge}</div></nav>{ASISTENTE_JS}'

def page(title,body,active=""):
    return f'<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{title} - Estudio Carlon</title><style>{CSS}</style></head><body>{nav_html(active)}<div class="wrap">{body}</div></body></html>'

def fmt(n):
    try: return f"${float(n):,.0f}".replace(",",".")
    except: return f"${n}"

def now_ar(): return datetime.now().strftime("%d/%m/%Y %H:%M")

def denied():
    b='<div class="denied"><div class="di">&#x1F512;</div><h2>Acceso restringido</h2><p>No tenes permiso para esta seccion.</p><a href="/clientes" class="btn btn-p">Volver</a></div>'
    return page("Acceso denegado",b),403

def login_req(f):
    @wraps(f)
    def w(*a,**kw):
        if not session.get("user"): return redirect("/")
        return f(*a,**kw)
    return w

def admin_req(f):
    @wraps(f)
    def w(*a,**kw):
        if not session.get("user"): return redirect("/")
        if session.get("rol")!="admin": return denied()
        return f(*a,**kw)
    return w

def conectar(): return psycopg2.connect(DB_URL)

def init_db():
    conn=conectar();c=conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS usuarios(id SERIAL PRIMARY KEY,usuario TEXT UNIQUE,clave TEXT,rol TEXT DEFAULT 'secretaria',nombre_display TEXT,totp_secret TEXT,totp_habilitado BOOLEAN DEFAULT FALSE,activo BOOLEAN DEFAULT TRUE)")
    c.execute("CREATE TABLE IF NOT EXISTS clientes(id SERIAL PRIMARY KEY,nombre TEXT,cuit TEXT,telefono TEXT,email TEXT,abono REAL DEFAULT 0)")
    c.execute("CREATE TABLE IF NOT EXISTS cuentas(id SERIAL PRIMARY KEY,cliente_id INTEGER,periodo TEXT,debe REAL DEFAULT 0,haber REAL DEFAULT 0)")
    c.execute("CREATE TABLE IF NOT EXISTS pagos(id SERIAL PRIMARY KEY,cliente_id INTEGER,periodo TEXT,monto REAL,medio TEXT,observaciones TEXT,facturado BOOLEAN DEFAULT FALSE,fecha TEXT,usuario TEXT,emitido_por TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS gastos(id SERIAL PRIMARY KEY,fecha TEXT,categoria TEXT,descripcion TEXT,monto REAL,usuario TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS auditoria(id SERIAL PRIMARY KEY,fecha TEXT,usuario TEXT,accion TEXT,detalle TEXT,cliente_id INTEGER,cliente_nombre TEXT)")
    c.execute("""CREATE TABLE IF NOT EXISTS cierres_caja(
        id SERIAL PRIMARY KEY,fecha TEXT,usuario TEXT,
        efectivo REAL DEFAULT 0,cheque REAL DEFAULT 0,dolares REAL DEFAULT 0,
        transferencia_nat REAL DEFAULT 0,transferencia_mai REAL DEFAULT 0,otro REAL DEFAULT 0,
        total_fisico REAL DEFAULT 0,total_general REAL DEFAULT 0,
        detalle_pagos TEXT,cerrado BOOLEAN DEFAULT FALSE,hora_cierre TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS agenda_vencimientos(
        id SERIAL PRIMARY KEY,vencimiento_id TEXT,mes INTEGER,anio INTEGER,
        estado TEXT DEFAULT 'pendiente',nota TEXT DEFAULT '',
        usuario TEXT,fecha_actualizacion TEXT,
        UNIQUE(vencimiento_id,mes,anio))""")
    c.execute("""CREATE TABLE IF NOT EXISTS tareas(
        id SERIAL PRIMARY KEY,titulo TEXT,descripcion TEXT,
        usuario TEXT,asignado_a TEXT,estado TEXT DEFAULT 'pendiente',
        prioridad TEXT DEFAULT 'normal',fecha_creacion TEXT,
        fecha_actualizacion TEXT,fecha_vencimiento TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS sueldos_estado(
        id SERIAL PRIMARY KEY,
        cliente_id INTEGER REFERENCES clientes(id) ON DELETE CASCADE,
        mes INTEGER NOT NULL,
        anio INTEGER NOT NULL,
        estado_recibo TEXT DEFAULT 'pendiente',
        estado_931 TEXT DEFAULT 'pendiente',
        estado_vep TEXT DEFAULT 'pendiente',
        fecha_vto_931 TEXT,
        observaciones TEXT,
        fecha_actualizacion TEXT,
        usuario TEXT,
        UNIQUE(cliente_id,mes,anio))""")
    c.execute("""CREATE TABLE IF NOT EXISTS empleados(
        id SERIAL PRIMARY KEY,
        cliente_id INTEGER REFERENCES clientes(id) ON DELETE CASCADE,
        nombre TEXT NOT NULL,
        cuil TEXT,
        categoria TEXT,
        convenio TEXT,
        fecha_ingreso TEXT,
        activo BOOLEAN DEFAULT TRUE,
        observaciones TEXT,
        fecha_alta TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS seguridad_eventos(
        id SERIAL PRIMARY KEY,tipo TEXT,detalle TEXT,ip TEXT,
        usuario TEXT,fecha TEXT,resuelto BOOLEAN DEFAULT FALSE)""")
    c.execute("""CREATE TABLE IF NOT EXISTS ips_bloqueadas(
        id SERIAL PRIMARY KEY,ip TEXT UNIQUE,motivo TEXT,
        fecha TEXT,desbloqueada BOOLEAN DEFAULT FALSE)""")
    c.execute("""CREATE TABLE IF NOT EXISTS config_seguridad(
        id SERIAL PRIMARY KEY,clave TEXT UNIQUE,valor TEXT)""")
    conn.commit()
    # Config por defecto
    defaults = [
        ("max_intentos_login","5"),
        ("bloqueo_mins","30"),
        ("paises_permitidos","AR"),
        ("alerta_whatsapp","1"),
        ("2fa_obligatorio","0"),
        ("session_timeout_mins","120"),
    ]
    for k,v in defaults:
        c.execute("INSERT INTO config_seguridad(clave,valor) VALUES(%s,%s) ON CONFLICT(clave) DO NOTHING",(k,v))
    # Crear admin por defecto si no hay ningún usuario admin
    c.execute("SELECT COUNT(*) FROM usuarios WHERE rol='admin'")
    if c.fetchone()[0] == 0:
        c.execute("SELECT COUNT(*) FROM usuarios")
        total = c.fetchone()[0]
        if total == 0:
            # Base limpia: crear admin natasha
            c.execute("INSERT INTO usuarios(usuario,clave,rol,nombre_display,activo) VALUES(%s,%s,%s,%s,TRUE)",
                      ("natasha", generate_password_hash("carlon2026"), "admin", "Natasha Carlon"))
        else:
            # Hay usuarios pero ninguno es admin: promover el primero
            c.execute("UPDATE usuarios SET rol='admin', activo=TRUE WHERE id=(SELECT id FROM usuarios ORDER BY id LIMIT 1)")
            # Asegurarse que todos tengan activo definido
            c.execute("UPDATE usuarios SET activo=TRUE WHERE activo IS NULL")
            c.execute("UPDATE usuarios SET totp_habilitado=FALSE WHERE totp_habilitado IS NULL")
    conn.commit();conn.close()

def actualizar_db():
    conn=conectar();c=conn.cursor()
    for ddl in [
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS email TEXT",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS abono REAL DEFAULT 0",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nombre_display TEXT",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS rol TEXT DEFAULT 'secretaria'",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS totp_secret TEXT",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS totp_habilitado BOOLEAN DEFAULT FALSE",
        "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE",
        "UPDATE usuarios SET activo=TRUE WHERE activo IS NULL",
        "UPDATE usuarios SET totp_habilitado=FALSE WHERE totp_habilitado IS NULL",
        "ALTER TABLE usuarios ALTER COLUMN activo SET DEFAULT TRUE",
        "ALTER TABLE usuarios ALTER COLUMN totp_habilitado SET DEFAULT FALSE",
        "ALTER TABLE auditoria ADD COLUMN IF NOT EXISTS cliente_nombre TEXT",
        "ALTER TABLE pagos ADD COLUMN IF NOT EXISTS observaciones TEXT",
        "ALTER TABLE pagos ADD COLUMN IF NOT EXISTS facturado BOOLEAN DEFAULT FALSE",
        "ALTER TABLE pagos ADD COLUMN IF NOT EXISTS emitido_por TEXT",
        "ALTER TABLE pagos ADD COLUMN IF NOT EXISTS concepto TEXT",
        "ALTER TABLE pagos ADD COLUMN IF NOT EXISTS periodos_incluidos TEXT",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS condicion_fiscal TEXT DEFAULT 'Responsable Inscripto'",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS actividad TEXT",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS envio_wa_facturas BOOLEAN DEFAULT FALSE",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS responsable_inscripto BOOLEAN DEFAULT FALSE",
        "ALTER TABLE clientes ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE",
        "UPDATE clientes SET activo=TRUE WHERE activo IS NULL",
    ]:
        try: c.execute(ddl)
        except: conn.rollback()
    conn.commit();conn.close()

def generar_deuda_mensual():
    conn=conectar();c=conn.cursor()
    periodo=datetime.now().strftime("%m/%Y")
    c.execute("SELECT id,abono FROM clientes")
    for cid,abono in c.fetchall():
        c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cid,periodo))
        if not c.fetchone():
            c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,%s,0)",(cid,periodo,abono or 0))
    conn.commit();conn.close()

def registrar_auditoria(accion,detalle,cliente_id=None,cliente_nombre=""):
    try:
        conn=conectar();c=conn.cursor()
        c.execute("INSERT INTO auditoria(fecha,usuario,accion,detalle,cliente_id,cliente_nombre) VALUES(%s,%s,%s,%s,%s,%s)",
                  (now_ar(),session.get("display",session.get("user","?")),accion,detalle,cliente_id,cliente_nombre))
        conn.commit();conn.close()
    except: pass

def svg_barras(datos,color="#C8A96E"):
    if not datos: return '<p style="color:var(--muted);font-size:.84rem">Sin datos</p>'
    W,H=400,150;mx=max(v for _,v in datos) or 1;n=len(datos)
    gap=int(W/n);bw=int(gap*0.6);bars=labels=vals=""
    for i,(per,val) in enumerate(datos):
        x=int(i*gap+gap*0.2);bh=int(val/mx*(H-30));y=H-20-bh
        bars+=f'<rect x="{x}" y="{y}" width="{bw}" height="{bh}" rx="3" fill="{color}" opacity=".85"/>'
        labels+=f'<text x="{x+bw//2}" y="{H-4}" text-anchor="middle" font-size="9" fill="#888">{per}</text>'
        if val>0: vals+=f'<text x="{x+bw//2}" y="{y-4}" text-anchor="middle" font-size="8" fill="{color}" font-weight="600">{fmt(val)}</text>'
    return f'<svg viewBox="0 0 {W} {H}" class="chart-svg">{bars}{labels}{vals}</svg>'

init_db();actualizar_db();generar_deuda_mensual()

# ══════════════════════════════════════════════════════════════════════════════
#  SETUP INICIAL — crear admin si no existe
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/setup", methods=["GET","POST"])
def setup():
    # Solo funciona si no hay ningún admin en la base
    conn=conectar();c=conn.cursor()
    c.execute("SELECT COUNT(*) FROM usuarios WHERE rol='admin'")
    hay_admin = c.fetchone()[0] > 0

    msg = ""
    if hay_admin:
        conn.close()
        # Hay admin: solo mostrar quiénes son (sin exponer clave)
        conn2=conectar();c2=conn2.cursor()
        c2.execute("SELECT usuario,nombre_display FROM usuarios WHERE rol='admin'")
        admins = c2.fetchall();conn2.close()
        lista = "".join(f"<li><b>{a[1] or a[0]}</b> → usuario: <code>{a[0]}</code></li>" for a in admins)
        return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Setup</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard" style="max-width:420px;width:94%">
<p class="ltitle">✅ Sistema configurado</p>
<p class="lsub">Ya existe un administrador</p>
<div class="info-box" style="text-align:left"><ul style="padding-left:16px">{lista}</ul></div>
<a href="/" class="btn btn-p" style="width:100%;justify-content:center;margin-top:14px">Ir al Login</a>
</div></div></body></html>'''

    if request.method=="POST":
        usuario = request.form.get("usuario","").strip()
        clave   = request.form.get("clave","").strip()
        nombre  = request.form.get("nombre","").strip() or usuario
        clave2  = request.form.get("clave2","").strip()
        if not usuario or not clave:
            msg = "Completá todos los campos"
        elif clave != clave2:
            msg = "Las contraseñas no coinciden"
        elif len(clave) < 6:
            msg = "Mínimo 6 caracteres"
        else:
            # Ver si el usuario ya existe → actualizarlo a admin
            c.execute("SELECT id FROM usuarios WHERE usuario=%s",(usuario,))
            row = c.fetchone()
            if row:
                c.execute("UPDATE usuarios SET clave=%s,rol='admin',nombre_display=%s,activo=TRUE WHERE id=%s",
                          (generate_password_hash(clave), nombre, row[0]))
            else:
                c.execute("INSERT INTO usuarios(usuario,clave,rol,nombre_display,activo) VALUES(%s,%s,'admin',%s,TRUE)",
                          (usuario, generate_password_hash(clave), nombre))
            conn.commit();conn.close()
            return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Listo</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard">
<p class="ltitle">✅ Admin creado</p>
<p class="lsub">Ya podés ingresar con <b>{usuario}</b></p>
<a href="/" class="btn btn-p" style="width:100%;justify-content:center;margin-top:14px">Ir al Login</a>
</div></div></body></html>'''

    conn.close()
    err = f'<div class="flash ferr">{msg}</div>' if msg else ""
    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Setup Inicial</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard" style="max-width:420px;width:94%">
<p class="ltitle">🔧 Setup Inicial</p>
<p class="lsub">Creá el usuario administrador del sistema</p>
{err}
<form method="post">
  <div class="fg" style="margin-bottom:12px"><label>Nombre completo</label>
    <input name="nombre" placeholder="Natasha Carlon"></div>
  <div class="fg" style="margin-bottom:12px"><label>Usuario (para login)</label>
    <input name="usuario" placeholder="natasha" required></div>
  <div class="fg" style="margin-bottom:12px"><label>Contraseña</label>
    <input name="clave" type="password" placeholder="Mínimo 6 caracteres" required></div>
  <div class="fg" style="margin-bottom:18px"><label>Confirmar contraseña</label>
    <input name="clave2" type="password" placeholder="Repetir contraseña" required></div>
  <button class="btn btn-p" style="width:100%;justify-content:center">Crear Administrador</button>
</form>
<div style="margin-top:12px;text-align:center">
  <a href="/" style="color:var(--muted);font-size:.8rem">&larr; Volver al login</a>
</div>
</div></div></body></html>'''

# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN con protección
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/", methods=["GET","POST"])
def login():
    ip = get_ip()
    # Anti-bot: rate limit global por IP (10 req/min)
    error = ""

    if request.method == "POST":
        # Verificar bloqueo
        bloqueado, mins = verificar_bloqueo_ip(ip)
        if bloqueado:
            registrar_evento_seguridad("INTENTO_IP_BLOQUEADA", f"IP {ip} intentó acceder bloqueada", ip)
            error = f"IP bloqueada por {mins} minuto(s) por exceso de intentos. Contacta al administrador."
        else:
            user = request.form.get("usuario","").strip()
            clave = request.form.get("clave","")
            totp_code = request.form.get("totp_code","").strip()
            conn = conectar(); c = conn.cursor()
            c.execute("SELECT clave,rol,nombre_display,totp_secret,totp_habilitado,activo FROM usuarios WHERE usuario=%s",(user,))
            data = c.fetchone(); conn.close()
            if data and check_password_hash(data[0], clave):
                # activo puede ser None (campo nuevo) → tratar None como activo=True
                if data[5] is False:
                    error = "Usuario desactivado. Contactá al administrador."
                elif data[4] is True and data[3] and len(str(data[3])) > 10:
                    # 2FA: solo si totp_habilitado=TRUE explícito y tiene secret válido
                    session["pending_2fa_user"] = user
                    session["pending_2fa_rol"] = data[1]
                    session["pending_2fa_display"] = data[2] or user
                    return redirect("/verificar_2fa")
                else:
                    limpiar_intento(ip)
                    session["user"]=user;session["rol"]=data[1] or "secretaria"
                    session["display"]=data[2] or user
                    registrar_auditoria("LOGIN","Inicio de sesion")
                    registrar_evento_seguridad("LOGIN_OK",f"Login exitoso",ip,data[2] or user)
                    if session["rol"]=="admin": return redirect("/panel")
                    elif session["rol"]=="supervisor": return redirect("/app")
                    else: return redirect("/panel_sec")
            else:
                count = registrar_intento_fallido(ip)
                restantes = max(0, MAX_INTENTOS - count)
                registrar_evento_seguridad("LOGIN_FALLIDO",f"Usuario: {user} - IP: {ip}",ip,user)
                if restantes > 0:
                    error = f"Usuario o contraseña incorrectos. Intentos restantes: {restantes}"
                else:
                    error = f"IP bloqueada por {BLOQUEO_MINS} minutos."

    err = f'<div class="flash ferr">{error}</div>' if error else ""
    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Ingresar</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard">
<p class="ltitle">Bienvenida</p>
<p class="lsub">Estudio Contable Carlon</p>{err}
<form method="post">
  <div class="fg" style="margin-bottom:12px"><label>Usuario</label>
    <input name="usuario" placeholder="tu usuario" autocomplete="username"></div>
  <div class="fg" style="margin-bottom:18px"><label>Contraseña</label>
    <input name="clave" type="password" placeholder="..." autocomplete="current-password"></div>
  <button class="btn btn-p" style="width:100%;justify-content:center">Ingresar</button>
</form>
</div></div></body></html>'''

@app.route("/reset_2fa", methods=["GET","POST"])
def reset_2fa():
    """Ruta de emergencia: desactiva 2FA de todos los usuarios sin necesitar login"""
    conn = conectar(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM usuarios WHERE rol='admin'")
    n_admin = c.fetchone()[0]

    msg = ""
    if request.method == "POST":
        clave_reset = request.form.get("clave_reset","").strip()
        # Clave de emergencia: primeros 8 chars del secret key + "reset"
        clave_esperada = (app.secret_key[:8] + "reset").lower()
        if clave_reset.lower() == clave_esperada:
            c.execute("UPDATE usuarios SET totp_habilitado=FALSE, totp_secret=NULL")
            conn.commit()
            registrar_evento_seguridad("RESET_2FA","Reset de emergencia 2FA ejecutado","","sistema")
            conn.close()
            return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Reset OK</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard">
<p class="ltitle">✅ 2FA Desactivado</p>
<p class="lsub">Ya podés entrar con usuario y contraseña normalmente</p>
<a href="/" class="btn btn-p" style="width:100%;justify-content:center;margin-top:14px">Ir al Login</a>
</div></div></body></html>'''
        else:
            msg = f"Clave incorrecta. Revisá el valor de SECRET_KEY en las variables de entorno."

    conn.close()
    # Calcular y mostrar la clave esperada (solo si no hay nadie logueado — es emergencia)
    clave_mostrar = (app.secret_key[:8] + "reset").lower()
    err = f'<div class="flash ferr">{msg}</div>' if msg else ""
    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Reset 2FA</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard" style="max-width:420px;width:94%">
<p class="ltitle">🔓 Desactivar 2FA</p>
<p class="lsub">Herramienta de emergencia</p>
{err}
<div class="warn-box" style="text-align:left;margin-bottom:16px">
  <b>Clave de emergencia:</b><br>
  <code style="font-size:1.1rem;letter-spacing:2px;font-weight:700">{clave_mostrar}</code><br>
  <span style="font-size:.74rem">Copiala y pegala abajo</span>
</div>
<form method="post">
  <div class="fg" style="margin-bottom:18px"><label>Clave de emergencia</label>
    <input name="clave_reset" placeholder="pegá la clave de arriba" required autofocus></div>
  <button class="btn btn-r" style="width:100%;justify-content:center">Desactivar 2FA en todos los usuarios</button>
</form>
<div style="margin-top:12px;text-align:center">
  <a href="/" style="color:var(--muted);font-size:.8rem">&larr; Volver al login</a>
</div>
</div></div></body></html>'''

@app.route("/verificar_2fa", methods=["GET","POST"])
def verificar_2fa():
    if "pending_2fa_user" not in session:
        return redirect("/")
    error = ""
    if request.method == "POST":
        code = request.form.get("code","").strip()
        user = session["pending_2fa_user"]
        conn = conectar(); c = conn.cursor()
        c.execute("SELECT totp_secret FROM usuarios WHERE usuario=%s",(user,))
        row = c.fetchone(); conn.close()
        if row:
            totp = pyotp.TOTP(row[0])
            if totp.verify(code):
                session["user"] = user
                session["rol"] = session.pop("pending_2fa_rol")
                session["display"] = session.pop("pending_2fa_display")
                session.pop("pending_2fa_user", None)
                registrar_auditoria("LOGIN","Inicio de sesion con 2FA")
                return redirect("/panel" if session["rol"]=="admin" else "/clientes")
            else:
                error = "Código incorrecto"
        else:
            error = "Error de sesión"
    err = f'<div class="flash ferr">{error}</div>' if error else ""
    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Verificar 2FA</title><style>{CSS}</style></head><body>
<div class="lwrap"><div class="lcard">
<p class="ltitle">🔐 Verificación</p>
<p class="lsub">Ingresá el código de tu app Google Authenticator</p>{err}
<form method="post">
  <div class="fg" style="margin-bottom:18px"><label>Código de 6 dígitos</label>
    <input name="code" type="text" inputmode="numeric" pattern="[0-9]*" maxlength="6"
     placeholder="000000" autocomplete="one-time-code" autofocus
     style="font-size:1.4rem;letter-spacing:8px;text-align:center"></div>
  <button class="btn btn-p" style="width:100%;justify-content:center">Verificar</button>
</form>
<div style="margin-top:12px;text-align:center">
  <a href="/" style="color:var(--muted);font-size:.8rem">&larr; Volver al login</a>
</div>
</div></div></body></html>'''

@app.route("/logout")
def logout():
    registrar_auditoria("LOGOUT","Cierre de sesion")
    registrar_evento_seguridad("LOGOUT","Cierre de sesion",get_ip())
    session.clear();return redirect("/")

# ══════════════════════════════════════════════════════════════════════════════
#  PANEL SECRETARIA
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/panel_sec")
@login_req
def panel_sec():
    if session.get("rol")=="admin": return redirect("/panel")
    conn=conectar();c=conn.cursor()
    usuario=session.get("display","")
    hoy=datetime.now()
    fecha_hoy=hoy.strftime("%d/%m/%Y")
    mes=hoy.month; anio=hoy.year

    # Totales del dia para esta secretaria
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE fecha LIKE %s AND emitido_por=%s",
              (fecha_hoy+"%",usuario))
    cobrado_hoy=c.fetchone()[0]

    # Clientes atendidos hoy
    c.execute("SELECT COUNT(DISTINCT cliente_id) FROM pagos WHERE fecha LIKE %s AND emitido_por=%s",
              (fecha_hoy+"%",usuario))
    clientes_hoy=c.fetchone()[0]

    # Vencimientos del mes
    c.execute("SELECT vencimiento_id,estado FROM agenda_vencimientos WHERE mes=%s AND anio=%s",(mes,anio))
    # F.931 vencimientos from clients with employees
    c.execute("""SELECT cl.nombre,cl.cuit,se.estado_931,se.estado_vep
                 FROM clientes cl
                 JOIN empleados e ON e.cliente_id=cl.id AND e.activo=TRUE
                 LEFT JOIN sueldos_estado se ON se.cliente_id=cl.id AND se.mes=%s AND se.anio=%s
                 WHERE cl.activo IS NOT FALSE
                 GROUP BY cl.id,cl.nombre,cl.cuit,se.estado_931,se.estado_vep
                 ORDER BY cl.nombre""",(mes,anio))
    clientes_con_sueldos=c.fetchall()
    venc_estados={r[0]:r[1] for r in c.fetchall()}
    proximos=[]
    for v in VENCIMIENTOS_IMPOSITIVOS:
        dias_rest=v["dia"]-hoy.day
        est=venc_estados.get(v["id"],"pendiente")
        if est in ("pendiente","borrador") and 0<=dias_rest<=7:
            proximos.append((v["nombre"],v["dia"],dias_rest,est))

    # Clientes RI que no han enviado facturas este mes
    mes_str=hoy.strftime("%m/%Y")
    c.execute("""SELECT COUNT(*) FROM clientes WHERE envio_wa_facturas=TRUE
                 AND activo IS NOT FALSE""")
    n_ri=c.fetchone()[0]

    # Tareas pendientes de esta secretaria
    c.execute("""SELECT id,titulo,estado,prioridad,fecha_vencimiento FROM tareas
                 WHERE (asignado_a=%s OR usuario=%s) AND estado!='completada'
                 ORDER BY CASE prioridad WHEN 'urgente' THEN 1 WHEN 'alta' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END,
                          fecha_vencimiento NULLS LAST LIMIT 10""",(usuario,usuario))
    mis_tareas=c.fetchall(); conn.close()

    # Vencimientos alertas
    alerta_venc=""
    if proximos:
        items="".join(f'<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid var(--border)"><span style="font-size:.82rem;font-weight:600">{v[0]}</span><span class="{"sec-badge danger" if v[2]<=2 else "sec-badge warn"}">{("VENCE HOY" if v[2]==0 else f"en {v[2]} días")}</span></div>' for v in proximos)
        alerta_venc=f'<div class="warn-box" style="margin-bottom:14px"><b>⏰ Vencimientos próximos ({len(proximos)})</b>{items}</div>'

    # Recordatorio RI
    rec_ri=(f'<div class="info-box" style="margin-bottom:14px">'
            f'📂 Hay <b>{n_ri}</b> clientes RI — recordá pedir facturas de compras este mes</div>' if n_ri>0 else "")

    # Tareas HTML
    PRIO_COLOR={"urgente":"var(--danger)","alta":"var(--warning)","normal":"var(--info)","baja":"var(--muted)"}
    tareas_html=""
    for t in mis_tareas:
        tid,titulo,est,prio,fvenc=t
        col=PRIO_COLOR.get(prio,"var(--info)")
        est_badge=('<span class="sec-badge warn">Borrador</span>' if est=="borrador"
                   else '<span class="sec-badge">Pendiente</span>')
        venc_txt=(f'<span style="font-size:.7rem;color:var(--muted)"> · Vence: {fvenc}</span>' if fvenc else "")
        tareas_html+=(f'<div class="logrow" style="justify-content:space-between;align-items:center">'
                     f'<div><div class="log-dot" style="background:{col}"></div></div>'
                     f'<div style="flex:1;margin-left:8px">'
                     f'<span style="font-weight:600;font-size:.85rem">{titulo}</span>{venc_txt}'
                     f'<br>{est_badge}</div>'
                     f'<div style="display:flex;gap:4px">'
                     f'<a href="/tareas?editar={tid}" class="btn btn-xs btn-o">✏️</a>'
                     f'<form method="post" action="/tareas/completar/{tid}" style="display:inline">'
                     f'<button class="btn btn-xs btn-g" title="Completar">✓</button></form>'
                     f'</div></div>')
    if not tareas_html:
        tareas_html='<p style="color:var(--muted);font-size:.84rem;padding:8px 0">Sin tareas pendientes ✨</p>'

    mes_nombre=["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                "Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes]

    body=f"""
    <h1 class="page-title">Bienvenida, {usuario.split()[0] if usuario else ""}!</h1>
    <p class="page-sub">{mes_nombre} {anio} · {fecha_hoy}</p>

    <!-- Dolar + Reloj -->
    <div style="background:var(--card);border-radius:var(--r);padding:12px 18px;box-shadow:var(--shadow);margin-bottom:16px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px">
      <div style="display:flex;gap:18px;flex-wrap:wrap;align-items:center">
        <div><span style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase">Dolar BNA Oficial</span>
          <div style="display:flex;gap:12px;margin-top:2px">
            <div><span style="font-size:.68rem;color:var(--muted)">Compra</span> <span id="ps-cmp" style="font-weight:700;color:var(--success)">---</span></div>
            <div><span style="font-size:.68rem;color:var(--muted)">Venta</span> <span id="ps-vta" style="font-weight:700;color:var(--danger)">---</span></div>
            <div><span style="font-size:.68rem;color:var(--muted)">Divisa</span> <span id="ps-div" style="font-weight:700;color:var(--warning)">---</span></div>
          </div>
        </div>
      </div>
      <div style="text-align:right">
        <div id="ps-reloj" style="font-family:'DM Serif Display',serif;font-size:1.1rem;color:var(--primary)"></div>
        <div id="ps-fecha" style="font-size:.74rem;color:var(--muted)"></div>
      </div>
    </div>

    <!-- Stats del dia -->
    <div class="stats" style="margin-bottom:16px">
      <div class="scard g"><div class="sicon">💰</div><div class="slabel">Cobrado hoy</div><div class="sval">{fmt(cobrado_hoy)}</div></div>
      <div class="scard b"><div class="sicon">👥</div><div class="slabel">Clientes hoy</div><div class="sval">{clientes_hoy}</div></div>
      <div class="scard o"><div class="sicon">⏰</div><div class="slabel">Vencimientos próximos</div><div class="sval">{len(proximos)}</div></div>
    </div>

    {alerta_venc}
    {rec_ri}

    <!-- Grid principal -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px" class="twocol">

      <!-- Tareas pendientes -->
      <div class="fcard" style="margin-bottom:0">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
          <h3 style="margin-bottom:0">📝 Mis tareas</h3>
          <a href="/tareas" class="btn btn-o btn-sm">Ver todas</a>
        </div>
        {tareas_html}
        <div style="margin-top:12px">
          <a href="/tareas?nueva=1" class="btn btn-p btn-sm">+ Nueva tarea</a>
        </div>
      </div>

      <!-- Novedades rápidas -->
      <div class="fcard" style="margin-bottom:0">
        <h3>📋 Accesos rápidos</h3>
        <div style="display:flex;flex-direction:column;gap:7px">
          <a href="/clientes" class="btn btn-p btn-sm">👥 Clientes</a>
          <a href="/caja" class="btn btn-o btn-sm">🏦 Mi Caja</a>
          <a href="/agenda" class="btn btn-o btn-sm">📅 Agenda Vencimientos</a>
          <a href="/novedades" class="btn btn-o btn-sm">📰 Novedades</a>
          <a href="/wa_masivo" class="btn btn-wa btn-sm">📱 WA Masivo</a>
          <a href="https://www.arca.gob.ar/landing/default.asp" target="_blank" class="btn btn-arca btn-sm">ARCA Login</a>
          <a href="https://servicioscf.afip.gob.ar/publico/sitio/contenido/novedad/listado.aspx" target="_blank" class="btn btn-o btn-sm">Novedades AFIP</a>
        </div>
      </div>

    </div>

    <script>
    fetch('https://dolarapi.com/v1/dolares/oficial').then(r=>r.json()).then(d=>{{
      var f=n=>'$'+n.toLocaleString('es-AR',{{minimumFractionDigits:2}});
      document.getElementById('ps-cmp').textContent=f(d.compra);
      document.getElementById('ps-vta').textContent=f(d.venta);
    }}).catch(()=>{{}});
    fetch('https://dolarapi.com/v1/dolares/tarjeta').then(r=>r.json()).then(d=>{{
      document.getElementById('ps-div').textContent='$'+d.venta.toLocaleString('es-AR',{{minimumFractionDigits:2}});
    }}).catch(()=>{{}});
    function tick(){{
      var n=new Date(),pad=x=>x.toString().padStart(2,"0");
      var dias=["Domingo","Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"];
      document.getElementById('ps-reloj').textContent=pad(n.getHours())+":"+pad(n.getMinutes())+":"+pad(n.getSeconds());
      document.getElementById('ps-fecha').textContent=dias[n.getDay()]+" "+pad(n.getDate())+"/"+(pad(n.getMonth()+1))+"/"+n.getFullYear();
    }}
    tick();setInterval(tick,1000);
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page("Inicio", body, "")


# ══════════════════════════════════════════════════════════════════════════════
#  PANEL PRINCIPAL con gráficos Chart.js
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/panel")
@login_req
def panel():
    if session.get("rol")!="admin": return redirect("/clientes")
    conn=conectar();c=conn.cursor()
    c.execute("SELECT COALESCE(SUM(debe),0) FROM cuentas");td=c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(haber),0) FROM cuentas");th=c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM clientes");nc=c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT cliente_id) FROM cuentas WHERE (debe-haber)>0");nd=c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(monto),0) FROM gastos");tg=c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Natasha%'");cobro_nat=c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Maira%'");cobro_mai=c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM clientes WHERE abono IS NULL OR abono=0");n_sin_abono=c.fetchone()[0]
    hoy=datetime.now().strftime("%d/%m/%Y")
    c.execute("SELECT COUNT(DISTINCT usuario) FROM pagos WHERE fecha LIKE %s",(f"%{hoy}%",));sec_cobraron=c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT usuario) FROM cierres_caja WHERE fecha=%s AND cerrado=TRUE",(hoy,));sec_cerraron=c.fetchone()[0]
    cajas_pendientes=max(0,sec_cobraron-sec_cerraron)
    # Alertas de seguridad sin resolver
    c.execute("SELECT COUNT(*) FROM seguridad_eventos WHERE resuelto=FALSE AND tipo IN ('BLOQUEO_IP','ACCESO_PAIS_BLOQUEADO','LOGIN_FALLIDO')");alertas_sec=c.fetchone()[0]
    # Datos para gráficos
    c.execute("SELECT periodo,COALESCE(SUM(haber),0) FROM cuentas WHERE periodo ~ '^[0-9]{{2}}/[0-9]{{4}}$' GROUP BY periodo ORDER BY SPLIT_PART(periodo,'/',2) DESC,SPLIT_PART(periodo,'/',1) DESC LIMIT 8")
    raw_ing=list(reversed(c.fetchall()))
    periodos=[r[0] for r in raw_ing];ingresos_m=[float(r[1]) for r in raw_ing]
    gastos_m=[]
    for per in periodos:
        c.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE fecha LIKE %s",(f"%{per.split('/')[1]}%",))
        gastos_m.append(float(c.fetchone()[0]))
    nat_m,mai_m=[],[]
    for per in periodos:
        c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE periodo=%s AND medio ILIKE '%%Natasha%%'",(per,));nat_m.append(float(c.fetchone()[0]))
        c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE periodo=%s AND medio ILIKE '%%Maira%%'",(per,));mai_m.append(float(c.fetchone()[0]))
    c.execute("SELECT medio,COALESCE(SUM(monto),0) FROM pagos GROUP BY medio ORDER BY SUM(monto) DESC")
    medios_raw=c.fetchall();total_med=sum(float(r[1]) for r in medios_raw) or 1
    medios_labels=[r[0] for r in medios_raw];medios_data=[round(float(r[1])/total_med*100,1) for r in medios_raw]
    c.execute("SELECT categoria,COALESCE(SUM(monto),0) FROM gastos GROUP BY categoria ORDER BY SUM(monto) DESC LIMIT 7")
    gastos_cat=c.fetchall();gcat_labels=[r[0] for r in gastos_cat];gcat_data=[float(r[1]) for r in gastos_cat]
    cum_ing,cum_gas=[],[]; si,sg=0.0,0.0
    for i in range(len(ingresos_m)):
        si+=ingresos_m[i];sg+=gastos_m[i];cum_ing.append(round(si));cum_gas.append(round(sg))
    c.execute("SELECT cl.nombre,SUM(COALESCE(cu.debe,0)-COALESCE(cu.haber,0)) d FROM cuentas cu JOIN clientes cl ON cl.id=cu.cliente_id WHERE cl.activo IS NOT FALSE GROUP BY cl.nombre HAVING SUM(COALESCE(cu.debe,0)-COALESCE(cu.haber,0))>0 ORDER BY d DESC LIMIT 6")
    top=c.fetchall()
    c.execute("SELECT COALESCE(fecha,''),COALESCE(usuario,''),COALESCE(accion,''),COALESCE(detalle,''),cliente_nombre FROM auditoria ORDER BY id DESC LIMIT 8")
    actividad=c.fetchall();conn.close()
    deuda=td-th;rend=th-tg;pct=int(th/td*100) if td>0 else 0
    total_s=cobro_nat+cobro_mai;pct_nat=int(cobro_nat/total_s*100) if total_s>0 else 50;pct_mai=100-pct_nat
    alertas=""
    if n_sin_abono>0:
        alertas+=f'<div class="warn-box"><b>{n_sin_abono}</b> clientes sin honorario · <a href="/clientes" style="color:#7a5800;font-weight:600">Ver</a></div>'
    if cajas_pendientes>0:
        alertas+=f'<div class="sec-alert"><b>{cajas_pendientes}</b> secretaria(s) sin cerrar caja hoy · <a href="/caja" style="color:#7a1a1a;font-weight:600">Ver caja</a></div>'
    if alertas_sec>0:
        alertas+=f'<div class="sec-alert">🔴 <b>{alertas_sec}</b> alerta(s) de seguridad pendientes · <a href="/seguridad" style="color:#7a1a1a;font-weight:600">Ver seguridad</a></div>'
    mx_deu=top[0][1] if top else 1
    barras_deu="".join(
        '<div class="chartrow"><span class="cl">'+str(n or "?")+'</span>'
        '<div class="cbg"><div class="cfill" style="width:'+str(int(float(s or 0)/mx_deu*100))+'%"></div></div>'
        '<span class="cv">'+fmt(s or 0)+'</span></div>'
        for n,s in top if s
    ) or '<p style="color:var(--muted);font-size:.84rem;padding:12px 0">Sin deudores</p>'
    _act_rows=[]
    for a in actividad:
        a = tuple(a) + (None,) * (5 - len(a))
        _act_rows.append(f'<div class="logrow"><div class="log-dot"></div><span class="log-time">{a[0]}</span><span class="log-user">{a[1]}</span><span class="log-msg"><b>{a[2]}</b> - {a[3]}{" · "+a[4] if a[4] else ""}</span></div>')
    act_html="".join(_act_rows) or '<p style="color:var(--muted);font-size:.84rem;padding:10px 0">Sin actividad</p>'
    body=f"""
    <h1 class="page-title">Panel General</h1>
    <p class="page-sub">Hola, <b>{session.get("display","")}</b> - {now_ar()}</p>
    {alertas}
    <div style="background:var(--card);border-radius:var(--r);padding:12px 20px;box-shadow:var(--shadow);margin-bottom:14px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px"><div style="display:flex;gap:20px;flex-wrap:wrap;align-items:center"><div><span style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px">Dolar BNA Oficial</span><div style="display:flex;gap:14px;align-items:baseline;margin-top:2px"><div><span style="font-size:.63rem;color:var(--muted)">Divisa Cmp</span> <span id="dolar-cmp" style="font-weight:700;font-size:.95rem;color:var(--success)">---</span></div><div><span style="font-size:.63rem;color:var(--muted)">Divisa Vta</span> <span id="dolar-vta" style="font-weight:700;font-size:.95rem;color:var(--danger)">---</span></div><div><span style="font-size:.63rem;color:var(--muted)">Billete Cmp</span> <span id="dolar-bil-cmp" style="font-weight:700;font-size:.95rem;color:var(--success)">---</span></div><div><span style="font-size:.63rem;color:var(--muted)">Billete Vta</span> <span id="dolar-bil-vta" style="font-weight:700;font-size:.95rem;color:var(--danger)">---</span></div></div></div></div><div style="text-align:right"><div id="reloj" style="font-family:'DM Serif Display',serif;font-size:1.15rem;color:var(--primary);font-weight:600"></div><div id="fecha-hoy" style="font-size:.74rem;color:var(--muted)"></div></div></div>
    <script>
    // Reloj en tiempo real
    function tick(){{
      var n=new Date();
      var d=n.getDate().toString().padStart(2,"0");
      var m=(n.getMonth()+1).toString().padStart(2,"0");
      var y=n.getFullYear();
      var H=n.getHours().toString().padStart(2,"0");
      var M=n.getMinutes().toString().padStart(2,"0");
      var S=n.getSeconds().toString().padStart(2,"0");
      var dias=["Domingo","Lunes","Martes","Miercoles","Jueves","Viernes","Sabado"];
      document.getElementById("reloj").textContent=H+":"+M+":"+S;
      document.getElementById("fecha-hoy").textContent=dias[n.getDay()]+" "+d+"/"+m+"/"+y;
    }}
    tick(); setInterval(tick,1000);
    // Cotizacion dolar BNA via DolarApi.ar
    fetch("https://dolarapi.com/v1/dolares/oficial")
      .then(r=>r.json())
      .then(d=>{{
        
        document.getElementById("dolar-cmp").textContent="$"+d.compra.toLocaleString("es-AR",{{minimumFractionDigits:2}});
        document.getElementById("dolar-vta").textContent="$"+d.venta.toLocaleString("es-AR",{{minimumFractionDigits:2}});
      }})
      .catch(()=>{{}});
    fetch("https://dolarapi.com/v1/dolares/bolsa")
      .then(r=>r.json())
      .then(d=>{{
        var bilCmp=document.getElementById("dolar-bil-cmp");
        var bilVta=document.getElementById("dolar-bil-vta");
        if(bilCmp) bilCmp.textContent="$"+d.compra.toLocaleString("es-AR",{{minimumFractionDigits:2}});
        if(bilVta) bilVta.textContent="$"+d.venta.toLocaleString("es-AR",{{minimumFractionDigits:2}});
      }})
      .catch(()=>{{}});
    </script>
    <div class="stats">
      <div class="scard"><div class="sicon">&#x1F4B0;</div><div class="slabel">Total Facturado</div><div class="sval">{fmt(td)}</div></div>
      <div class="scard g"><div class="sicon">&#x2705;</div><div class="slabel">Total Cobrado</div><div class="sval">{fmt(th)}</div></div>
      <div class="scard r"><div class="sicon">&#x1F534;</div><div class="slabel">Deuda Pendiente</div><div class="sval">{fmt(deuda)}</div></div>
      <div class="scard o"><div class="sicon">&#x1F4B8;</div><div class="slabel">Total Gastos</div><div class="sval">{fmt(tg)}</div></div>
      <div class="scard {'g' if rend>=0 else 'r'}"><div class="sicon">&#x1F4CA;</div><div class="slabel">Rendimiento Real</div><div class="sval">{fmt(rend)}</div></div>
      <div class="scard b"><div class="sicon">&#x1F465;</div><div class="slabel">Clientes</div><div class="sval">{nc}</div></div>
    </div>
    <div class="fcard" style="margin-bottom:18px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <span style="font-weight:600;color:var(--primary)">Cobrado vs Facturado</span>
        <span style="font-weight:700;color:var(--success)">{pct}%</span>
      </div>
      <div class="progwrap"><div class="progbar" style="width:{pct}%"></div></div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px" class="twocol">
      <div class="partner-card">
        <div class="partner-name">Natasha Carlon</div>
        <div class="partner-amt" style="color:var(--primary)">{fmt(cobro_nat)}</div>
        <div style="font-size:.72rem;color:var(--muted)">{pct_nat}% del total cobrado</div>
      </div>
      <div class="partner-card">
        <div class="partner-name">Maira Carlon</div>
        <div class="partner-amt" style="color:var(--info)">{fmt(cobro_mai)}</div>
        <div style="font-size:.72rem;color:var(--muted)">{pct_mai}% del total cobrado</div>
      </div>
    </div>
    <div class="fcard" style="margin-bottom:18px">
      <h3>📊 Ingresos vs Gastos — últimos meses</h3>
      <div style="position:relative;height:220px"><canvas id="ch1"></canvas></div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px" class="twocol">
      <div class="fcard" style="margin-bottom:0">
        <h3>Cobros por socia</h3>
        <div style="position:relative;height:185px"><canvas id="ch2"></canvas></div>
      </div>
      <div class="fcard" style="margin-bottom:0">
        <h3>Medios de pago</h3>
        <div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap">
          <div style="position:relative;height:160px;width:160px;flex-shrink:0"><canvas id="ch3"></canvas></div>
          <div id="leg3" style="font-size:.74rem;color:var(--muted);display:flex;flex-direction:column;gap:5px"></div>
        </div>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px" class="twocol">
      <div class="fcard" style="margin-bottom:0">
        <h3>Gastos por categoría</h3>
        <div id="gcat" style="display:flex;flex-direction:column;gap:7px"></div>
      </div>
      <div class="fcard" style="margin-bottom:0">
        <h3>Rendimiento acumulado</h3>
        <div style="position:relative;height:170px"><canvas id="ch4"></canvas></div>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px" class="twocol">
      <div class="fcard" style="margin-bottom:0"><h3>Top Deudores</h3>{barras_deu}</div>
      <div class="fcard" style="margin-bottom:0"><h3>Actividad Reciente</h3>{act_html}</div>
    </div>
    <div class="qa">
      <a href="/clientes" class="btn btn-p">Clientes</a>
      <a href="/deudas" class="btn btn-a">Deudores ({nd})</a>
      <a href="/gastos" class="btn btn-o">Gastos</a>
      <a href="/caja" class="btn btn-o">Caja</a>
      <a href="/reportes" class="btn btn-b">Reportes</a>
      <a href="/agenda" class="btn btn-o">Agenda</a>
      <a href="/seguridad" class="btn btn-r">🔒 Seguridad</a>
      <a href="/configuracion" class="btn btn-o">⚙️ Config</a>
    </div>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
    <script>
    const P={json.dumps(periodos)},IM={json.dumps(ingresos_m)},GM={json.dumps(gastos_m)};
    const NM={json.dumps(nat_m)},MM={json.dumps(mai_m)};
    const ML={json.dumps(medios_labels)},MD={json.dumps(medios_data)};
    const GL={json.dumps(gcat_labels)},GD={json.dumps(gcat_data)};
    const CI={json.dumps(cum_ing)},CG={json.dumps(cum_gas)};
    const MC=['#185FA5','#0F6E56','#1D9E75','#E67E22','#7B68EE','#E24B4A','#888780'];
    const gc='rgba(0,0,0,0.06)',tc='rgba(0,0,0,0.42)';
    const fK=v=>'$'+(Math.abs(v)>=1000?(v/1000).toFixed(0)+'k':Math.round(v));
    const fF=v=>'$'+Math.round(v).toLocaleString('es-AR');
    const base={{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>' '+ctx.dataset.label+': '+fF(ctx.raw)}}}}}},scales:{{x:{{ticks:{{color:tc,font:{{size:10}}}},grid:{{color:gc}}}},y:{{ticks:{{color:tc,font:{{size:10}},callback:fK}},grid:{{color:gc}}}}}}}};
    new Chart(document.getElementById('ch1'),{{type:'bar',data:{{labels:P,datasets:[{{label:'Cobrado',data:IM,backgroundColor:'#185FA5',borderRadius:4}},{{label:'Gastos',data:GM,backgroundColor:'#E24B4A',borderRadius:4}}]}},options:base}});
    new Chart(document.getElementById('ch2'),{{type:'bar',data:{{labels:P,datasets:[{{label:'Natasha',data:NM,backgroundColor:'#185FA5',borderRadius:3}},{{label:'Maira',data:MM,backgroundColor:'#0F6E56',borderRadius:3}}]}},options:base}});
    new Chart(document.getElementById('ch3'),{{type:'doughnut',data:{{labels:ML,datasets:[{{data:MD,backgroundColor:MC.slice(0,ML.length),borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,cutout:'60%',plugins:{{legend:{{display:false}}}}}}}});
    const lg=document.getElementById('leg3');
    ML.forEach((l,i)=>{{lg.innerHTML+=`<span style="display:flex;align-items:center;gap:5px"><span style="width:9px;height:9px;border-radius:2px;background:${{MC[i]}};flex-shrink:0;display:inline-block"></span>${{l}} <b>${{MD[i]}}%</b></span>`}});
    const gDiv=document.getElementById('gcat'),mxG=Math.max(...GD)||1;
    GL.forEach((l,i)=>{{gDiv.innerHTML+=`<div style="display:flex;align-items:center;gap:7px;font-size:.78rem"><span style="width:100px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis">${{l}}</span><div style="flex:1;background:var(--border);border-radius:3px;height:7px"><div style="width:${{Math.round(GD[i]/mxG*100)}}%;height:100%;background:#E67E22;border-radius:3px"></div></div><span style="width:70px;text-align:right;font-weight:600">${{fK(GD[i])}}</span></div>`}});
    new Chart(document.getElementById('ch4'),{{type:'line',data:{{labels:P,datasets:[{{label:'Ingresos',data:CI,borderColor:'#1D9E75',backgroundColor:'rgba(29,158,117,0.07)',fill:true,tension:0.35,pointRadius:3,pointBackgroundColor:'#1D9E75'}},{{label:'Gastos',data:CG,borderColor:'#E24B4A',backgroundColor:'rgba(226,75,74,0.05)',fill:true,tension:0.35,pointRadius:3,pointBackgroundColor:'#E24B4A',borderDash:[5,3]}}]}},options:base}});
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page("Panel", body, "Panel")

# ══════════════════════════════════════════════════════════════════════════════
#  PANEL DE SEGURIDAD
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/seguridad", methods=["GET","POST"])
@admin_req
def seguridad():
    conn=conectar();c=conn.cursor();flash=""

    if request.method=="POST":
        accion = request.form.get("accion","")
        if accion == "desbloquear_ip":
            ip_d = request.form.get("ip","").strip()
            desbloquear_ip_admin(ip_d)
            c.execute("UPDATE ips_bloqueadas SET desbloqueada=TRUE WHERE ip=%s",(ip_d,))
            conn.commit()
            c.execute("UPDATE seguridad_eventos SET resuelto=TRUE WHERE ip=%s AND tipo='BLOQUEO_IP'",(ip_d,))
            conn.commit()
            flash = f'<div class="flash fok">IP {ip_d} desbloqueada</div>'
            registrar_auditoria("DESBLOQUEO_IP",f"Admin desbloqueó {ip_d}")
        elif accion == "resolver_evento":
            eid = request.form.get("eid","")
            c.execute("UPDATE seguridad_eventos SET resuelto=TRUE WHERE id=%s",(eid,))
            conn.commit()
            flash = '<div class="flash fok">Evento marcado como resuelto</div>'
        elif accion == "bloquear_ip_manual":
            ip_m = request.form.get("ip_manual","").strip()
            motivo = request.form.get("motivo","Bloqueo manual").strip()
            if ip_m:
                LOGIN_INTENTOS[ip_m] = {"count":99,"ts":time.time(),"blocked_until":time.time()+86400*7}
                c.execute("INSERT INTO ips_bloqueadas(ip,motivo,fecha) VALUES(%s,%s,%s) ON CONFLICT(ip) DO UPDATE SET motivo=%s,fecha=%s,desbloqueada=FALSE",
                          (ip_m,motivo,now_ar(),motivo,now_ar()))
                conn.commit()
                registrar_evento_seguridad("BLOQUEO_MANUAL",f"Admin bloqueó {ip_m}: {motivo}",ip_m)
                flash = f'<div class="flash fok">IP {ip_m} bloqueada manualmente</div>'
        elif accion == "enviar_test_wa":
            enviar_whatsapp(f"✅ TEST SEGURIDAD - Sistema Estudio Carlon\nEste mensaje llega a tu celular personal (admin).\n📅 {now_ar()}")
            flash = '<div class="flash fok">Mensaje de prueba enviado por WhatsApp</div>'

    # Cargar datos de seguridad
    c.execute("SELECT tipo,detalle,ip,usuario,fecha,id,resuelto FROM seguridad_eventos ORDER BY id DESC LIMIT 100")
    eventos = c.fetchall()
    c.execute("SELECT ip,motivo,fecha,desbloqueada FROM ips_bloqueadas ORDER BY id DESC LIMIT 50")
    ips_bl = c.fetchall()
    c.execute("SELECT COUNT(*) FROM seguridad_eventos WHERE resuelto=FALSE");total_pend = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM seguridad_eventos WHERE tipo='LOGIN_FALLIDO' AND fecha LIKE %s",(f"%{datetime.now().strftime('%d/%m/%Y')}%",));hoy_fallidos = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM seguridad_eventos WHERE tipo='BLOQUEO_IP'");total_bloq = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM seguridad_eventos WHERE tipo='LOGIN_OK' AND fecha LIKE %s",(f"%{datetime.now().strftime('%d/%m/%Y')}%",));hoy_ok = c.fetchone()[0]
    conn.close()

    # IPs actualmente bloqueadas en memoria
    ips_mem = [(ip,d) for ip,d in LOGIN_INTENTOS.items() if d.get("blocked_until",0) > time.time()]

    filas_ev = ""
    colores = {"BLOQUEO_IP":"red","ACCESO_PAIS_BLOQUEADO":"red","LOGIN_FALLIDO":"orange","LOGIN_OK":"green","LOGOUT":"green","LOGIN_MANUAL":"orange"}
    for ev in eventos:
        tipo,det,ip_ev,usr_ev,fec,eid,res = ev
        col = colores.get(tipo,"")
        resuelto_badge = '<span class="sec-badge ok">✓</span>' if res else f'<form method="post" style="display:inline"><input type=hidden name=accion value=resolver_evento><input type=hidden name=eid value={eid}><button class="btn btn-xs btn-o">Resolver</button></form>'
        filas_ev += f'<div class="logrow"><div class="log-dot {col}"></div><span class="log-time">{fec}</span><span class="log-user" style="min-width:90px;font-size:.7rem">{tipo}</span><span class="log-msg">{det} {" · "+usr_ev if usr_ev else ""} {" · IP:"+ip_ev if ip_ev else ""}</span>{resuelto_badge}</div>'

    filas_ips = ""
    for ip_b,mot,fec_b,desbl in ips_bl:
        estado = '<span class="sec-badge ok">Desbloqueada</span>' if desbl else '<span class="sec-badge danger">Bloqueada</span>'
        btn_desbl = "" if desbl else f'<form method="post" style="display:inline"><input type=hidden name=accion value=desbloquear_ip><input type=hidden name=ip value="{ip_b}"><button class="btn btn-xs btn-g">Desbloquear</button></form>'
        filas_ips += f'<tr><td class="nm">{ip_b}</td><td class="mu">{mot}</td><td class="mu">{fec_b}</td><td>{estado}</td><td>{btn_desbl}</td></tr>'

    ips_mem_html = ""
    for ip_m,d in ips_mem:
        mins_rest = int((d["blocked_until"] - time.time()) / 60) + 1
        ips_mem_html += f'<div class="logrow"><div class="log-dot red"></div><span class="log-time">Activa</span><span class="log-user">{ip_m}</span><span class="log-msg">{d["count"]} intentos · {mins_rest} min restantes</span><form method="post" style="display:inline"><input type=hidden name=accion value=desbloquear_ip><input type=hidden name=ip value="{ip_m}"><button class="btn btn-xs btn-g">Desbloquear</button></form></div>'

    body = f"""
    <h1 class="page-title">🔒 Panel de Seguridad</h1>
    <p class="page-sub">Monitoreo de accesos, alertas y bloqueos</p>
    {flash}
    <div class="stats">
      <div class="scard r"><div class="sicon">⚠️</div><div class="slabel">Alertas pendientes</div><div class="sval">{total_pend}</div></div>
      <div class="scard o"><div class="sicon">🔴</div><div class="slabel">Fallos hoy</div><div class="sval">{hoy_fallidos}</div></div>
      <div class="scard p"><div class="sicon">🛡️</div><div class="slabel">IPs bloqueadas</div><div class="sval">{total_bloq}</div></div>
      <div class="scard g"><div class="sicon">✅</div><div class="slabel">Logins ok hoy</div><div class="sval">{hoy_ok}</div></div>
    </div>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px" class="twocol">
      <div class="fcard" style="margin-bottom:0">
        <h3>🔴 IPs bloqueadas en este momento</h3>
        {ips_mem_html or '<p style="color:var(--muted);font-size:.84rem">Ninguna IP bloqueada</p>'}
      </div>
      <div class="fcard" style="margin-bottom:0">
        <h3>🛡️ Bloquear IP manualmente</h3>
        <form method="post">
          <input type="hidden" name="accion" value="bloquear_ip_manual">
          <div class="fgrid" style="grid-template-columns:1fr">
            <div class="fg"><label>Dirección IP</label><input name="ip_manual" placeholder="192.168.1.100"></div>
            <div class="fg"><label>Motivo</label><input name="motivo" placeholder="Actividad sospechosa" value="Bloqueo manual admin"></div>
          </div>
          <div style="display:flex;gap:8px;flex-wrap:wrap">
            <button class="btn btn-r btn-sm">Bloquear IP</button>
            <form method="post" style="display:inline"><input type=hidden name=accion value=enviar_test_wa><button class="btn btn-wa btn-sm">📱 Test WhatsApp</button></form>
          </div>
        </form>
      </div>
    </div>

    <div class="fcard">
      <h3>📋 Historial de IPs bloqueadas</h3>
      <div class="dtable"><table>
        <thead><tr><th>IP</th><th>Motivo</th><th>Fecha</th><th>Estado</th><th>Acción</th></tr></thead>
        <tbody>{filas_ips or "<tr><td colspan=5 style='color:var(--muted);text-align:center;padding:20px'>Sin registros</td></tr>"}</tbody>
      </table></div>
    </div>

    <div class="fcard">
      <h3>🔍 Registro de eventos de seguridad</h3>
      <div class="search"><span>🔍</span><input id="bus" placeholder="Filtrar eventos..." oninput="filtEv(this.value)"></div>
      <div id="ev-lista">
        {filas_ev or '<p style="color:var(--muted);font-size:.84rem">Sin eventos registrados</p>'}
      </div>
    </div>
    <script>
    function filtEv(q){{q=q.toLowerCase();document.querySelectorAll('#ev-lista .logrow').forEach(r=>r.style.display=r.textContent.toLowerCase().includes(q)?'flex':'none')}}
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page("Seguridad", body, "🔒 Seguridad")

# ══════════════════════════════════════════════════════════════════════════════
#  PANEL DE CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/configuracion", methods=["GET","POST"])
@admin_req
def configuracion():
    conn=conectar();c=conn.cursor();flash=""

    if request.method=="POST":
        accion = request.form.get("accion","")

        if accion == "crear_usuario":
            usuario=request.form.get("usuario","").strip()
            clave=request.form.get("clave","").strip()
            rol=request.form.get("rol","secretaria")
            display=request.form.get("nombre_display","").strip() or usuario
            if not usuario or not clave:
                flash='<div class="flash ferr">Completá usuario y contraseña</div>'
            elif len(clave) < 6:
                flash='<div class="flash ferr">Contraseña mínimo 6 caracteres</div>'
            else:
                c.execute("SELECT id FROM usuarios WHERE usuario=%s",(usuario,))
                if c.fetchone():
                    flash='<div class="flash ferr">Ese usuario ya existe</div>'
                else:
                    c.execute("INSERT INTO usuarios(usuario,clave,rol,nombre_display,activo,totp_habilitado) VALUES(%s,%s,%s,%s,TRUE,FALSE)",
                              (usuario,generate_password_hash(clave),rol,display))
                    conn.commit()
                    registrar_auditoria("NUEVO USUARIO",f"@{usuario} Rol:{rol}")
                    registrar_evento_seguridad("NUEVO_USUARIO",f"Creado @{usuario} rol:{rol}",get_ip())
                    flash=f'<div class="flash fok">✅ Usuario {display} creado correctamente</div>'

        elif accion == "cambiar_clave":
            uid=request.form.get("uid")
            nueva=request.form.get("nueva_clave","").strip()
            confirmar=request.form.get("confirmar_clave","").strip()
            if nueva != confirmar:
                flash='<div class="flash ferr">Las contraseñas no coinciden</div>'
            elif len(nueva) < 6:
                flash='<div class="flash ferr">Mínimo 6 caracteres</div>'
            else:
                c.execute("UPDATE usuarios SET clave=%s WHERE id=%s",(generate_password_hash(nueva),uid))
                conn.commit()
                c.execute("SELECT nombre_display,usuario FROM usuarios WHERE id=%s",(uid,))
                u=c.fetchone()
                registrar_auditoria("CAMBIO_CLAVE",f"Clave cambiada para @{u[1] if u else uid}")
                registrar_evento_seguridad("CAMBIO_CLAVE",f"Admin cambió clave de @{u[1] if u else uid}",get_ip())
                flash='<div class="flash fok">✅ Contraseña actualizada</div>'

        elif accion == "cambiar_mis_datos":
            nd2=request.form.get("nuevo_display","").strip()
            nc2=request.form.get("nueva_clave_admin","").strip()
            nc2_conf=request.form.get("confirmar_clave_admin","").strip()
            user_actual=session.get("user")
            if nd2:
                c.execute("UPDATE usuarios SET nombre_display=%s WHERE usuario=%s",(nd2,user_actual))
                session["display"]=nd2
            if nc2:
                if nc2 != nc2_conf:
                    flash='<div class="flash ferr">Las contraseñas no coinciden</div>'
                elif len(nc2) < 6:
                    flash='<div class="flash ferr">Mínimo 6 caracteres</div>'
                else:
                    c.execute("UPDATE usuarios SET clave=%s WHERE usuario=%s",(generate_password_hash(nc2),user_actual))
            conn.commit()
            if not flash:
                flash='<div class="flash fok">✅ Datos actualizados</div>'

        elif accion == "borrar_usuario":
            uid=request.form.get("uid")
            c.execute("SELECT usuario,nombre_display FROM usuarios WHERE id=%s",(uid,))
            u=c.fetchone()
            if u and u[0]!=session.get("user"):
                c.execute("DELETE FROM usuarios WHERE id=%s",(uid,))
                conn.commit()
                registrar_auditoria("BAJA_USUARIO",f"Eliminado @{u[0]}")
                flash='<div class="flash fok">Usuario eliminado</div>'
            else:
                flash='<div class="flash ferr">No podés eliminar tu propio usuario</div>'

        elif accion == "activar_2fa":
            uid = request.form.get("uid")
            c.execute("SELECT usuario,totp_secret FROM usuarios WHERE id=%s",(uid,))
            u = c.fetchone()
            if u:
                secret = u[1] or pyotp.random_base32()
                c.execute("UPDATE usuarios SET totp_secret=%s,totp_habilitado=TRUE WHERE id=%s",(secret,uid))
                conn.commit()
                flash=f'<div class="flash fok">✅ 2FA activado para @{u[0]}</div>'

        elif accion == "desactivar_2fa":
            uid = request.form.get("uid")
            c.execute("UPDATE usuarios SET totp_habilitado=FALSE WHERE id=%s",(uid,))
            conn.commit()
            flash='<div class="flash fok">2FA desactivado</div>'

        elif accion == "guardar_config_seg":
            keys = ["max_intentos_login","bloqueo_mins","paises_permitidos","alerta_whatsapp","session_timeout_mins"]
            for k in keys:
                v = request.form.get(k,"")
                if v:
                    c.execute("UPDATE config_seguridad SET valor=%s WHERE clave=%s",(v,k))
            conn.commit()
            flash='<div class="flash fok">✅ Configuración de seguridad guardada</div>'

        elif accion == "editar_usuario":
            uid = request.form.get("uid")
            nuevo_display = request.form.get("nuevo_display","").strip()
            nuevo_rol = request.form.get("nuevo_rol","secretaria")
            if uid:
                # No permitir quitarse rol admin a uno mismo si es el único admin
                c.execute("SELECT usuario FROM usuarios WHERE id=%s",(uid,))
                row_u = c.fetchone()
                if row_u and row_u[0] == session.get("user") and nuevo_rol != "admin":
                    flash='<div class="flash ferr">No podés quitarte el rol admin a vos mismo</div>'
                else:
                    c.execute("UPDATE usuarios SET nombre_display=%s, rol=%s WHERE id=%s",
                              (nuevo_display, nuevo_rol, uid))
                    conn.commit()
                    registrar_auditoria("EDICION_USUARIO",f"Editado uid:{uid} display:{nuevo_display} rol:{nuevo_rol}")
                    flash='<div class="flash fok">✅ Usuario actualizado correctamente</div>'

        elif accion == "activar_desactivar_usuario":
            uid = request.form.get("uid")
            activo_val = request.form.get("activo_val","1")
            c.execute("UPDATE usuarios SET activo=%s WHERE id=%s",(activo_val=="1",uid))
            conn.commit()
            flash='<div class="flash fok">Estado de usuario actualizado</div>'

    # Cargar datos
    c.execute("SELECT id,usuario,rol,nombre_display,totp_habilitado,activo FROM usuarios ORDER BY rol,usuario")
    lista = c.fetchall()
    c.execute("SELECT clave,valor FROM config_seguridad")
    config = {r[0]:r[1] for r in c.fetchall()}
    conn.close()

    # Generar QR codes para 2FA por usuario
    qr_html = ""
    cards = ""
    for u in lista:
        uid,uname,urol,udisp,totp_on,activo = u
        badge = '<span class="badge badm">Admin</span>' if urol=="admin" else '<span class="badge bsec">Secretaria</span>'
        activo_badge = '<span class="sec-badge ok">Activo</span>' if activo is not False else '<span class="sec-badge danger">Inactivo</span>'
        totp_badge = '<span class="sec-badge ok">2FA ✓</span>' if totp_on is True else '<span class="sec-badge warn">2FA off</span>'
        es_yo = uname == session.get("user")

        btn_del = '<span style="font-size:.73rem;color:var(--muted)">sos vos</span>' if es_yo else \
            f'<form method="post" style="display:inline" onsubmit="return confirm(\'Eliminar a {udisp or uname}?\')"><input type=hidden name=accion value=borrar_usuario><input type=hidden name=uid value={uid}><button class="btn btn-xs btn-r">🗑</button></form>'

        btn_2fa = f'<form method="post" style="display:inline"><input type=hidden name=accion value={"desactivar_2fa" if totp_on else "activar_2fa"}><input type=hidden name=uid value={uid}><button class="btn btn-xs {"btn-o" if totp_on else "btn-b"}">{"Desact. 2FA" if totp_on else "Activar 2FA"}</button></form>'

        _activo_real = activo is not False
        btn_act = f'<form method="post" style="display:inline"><input type=hidden name=accion value=activar_desactivar_usuario><input type=hidden name=uid value={uid}><input type=hidden name=activo_val value={"0" if _activo_real else "1"}><button class="btn btn-xs {"btn-o" if _activo_real else "btn-g"}">{"Deshabilitar" if _activo_real else "Habilitar"}</button></form>' if not es_yo else ""

        cards += f'''<div class="ucard {"adm" if urol=="admin" else ""}">
          <div>
            <div style="font-weight:600;font-size:.96rem;color:var(--primary)">{udisp or uname} {badge}</div>
            <div style="font-size:.75rem;color:var(--muted)">@{uname}</div>
            <div style="margin-top:4px;display:flex;gap:5px">{activo_badge}{totp_badge}</div>
          </div>
          <div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center">
            <button onclick="abrirClave({uid},'{uname}')" class="btn btn-xs btn-o">🔑 Clave</button>
            <button onclick="abrirEditar({uid},'{udisp or uname}','{urol}')" class="btn btn-xs btn-b">✏️ Editar</button>
            {btn_2fa}{btn_act}{btn_del}
          </div>
        </div>'''

    body = f"""
    <h1 class="page-title">⚙️ Configuración</h1>
    <p class="page-sub">Usuarios, contraseñas, seguridad y 2FA</p>
    {flash}

    <div class="tabs">
      <button class="tab on" onclick="showTab('t1',this)">👥 Usuarios</button>
      <button class="tab" onclick="showTab('t2',this)">➕ Nuevo Usuario</button>
      <button class="tab" onclick="showTab('t3',this)">🔒 Mi Cuenta</button>
      <button class="tab" onclick="showTab('t4',this)">🛡️ Config. Seguridad</button>
    </div>

    <div id="t1" class="tabpanel on">
      <div class="fcard">
        <h3>Usuarios del sistema ({len(lista)})</h3>
        {cards or '<p style="color:var(--muted)">Sin usuarios</p>'}
      </div>
      <div class="fcard">
        <h3>ℹ️ Permisos por rol</h3>
        <div class="info-box" style="margin-bottom:8px"><b>Admin:</b> Panel financiero, gráficos, reportes, usuarios, configuración, seguridad, todos los módulos.</div>
        <div style="background:#f0f4ff;border:1px solid #b0c4ee;border-radius:8px;padding:10px 14px;font-size:.8rem;color:#1a3a8a"><b>Secretaria:</b> Clientes, pagos, recibos, gastos, caja, agenda. Sin acceso a panel, reportes, usuarios ni configuración.</div>
      </div>
    </div>

    <div id="t2" class="tabpanel">
      <div class="fcard">
        <h3>➕ Crear nuevo usuario</h3>
        <form method="post">
          <input type="hidden" name="accion" value="crear_usuario">
          <div class="fgrid">
            <div class="fg"><label>Nombre completo</label><input name="nombre_display" placeholder="María González"></div>
            <div class="fg"><label>Usuario (login)</label><input name="usuario" placeholder="mariag" required></div>
            <div class="fg"><label>Contraseña</label><input name="clave" type="password" placeholder="Mínimo 6 caracteres" required></div>
            <div class="fg"><label>Rol</label>
              <select name="rol">
                <option value="secretaria">Secretaria</option>
                <option value="supervisor">Supervisora (App móvil)</option>
                <option value="admin">Administrador</option>
              </select>
            </div>
          </div>
          <div class="info-box">🔒 Las contraseñas se guardan encriptadas con hash seguro (bcrypt). Nunca se almacenan en texto plano.</div>
          <button class="btn btn-p">Crear Usuario</button>
        </form>
      </div>
    </div>

    <div id="t3" class="tabpanel">
      <div class="fcard">
        <h3>🔒 Mis datos — {session.get("display","")}</h3>
        <form method="post">
          <input type="hidden" name="accion" value="cambiar_mis_datos">
          <div class="fgrid">
            <div class="fg"><label>Nombre para mostrar</label><input name="nuevo_display" value="{session.get('display','')}"></div>
            <div class="fg"><label>Nueva contraseña (vacío = no cambia)</label><input name="nueva_clave_admin" type="password" placeholder="Nueva contraseña"></div>
            <div class="fg"><label>Confirmar nueva contraseña</label><input name="confirmar_clave_admin" type="password" placeholder="Repetir contraseña"></div>
          </div>
          <button class="btn btn-a btn-sm">Guardar mis datos</button>
        </form>
      </div>
    </div>

    <div id="t4" class="tabpanel">
      <div class="fcard">
        <h3>🛡️ Configuración de seguridad</h3>
        <form method="post">
          <input type="hidden" name="accion" value="guardar_config_seg">
          <div class="fgrid">
            <div class="fg">
              <label>Intentos máx. login antes de bloquear</label>
              <input name="max_intentos_login" type="number" value="{config.get('max_intentos_login','5')}" min="2" max="20">
            </div>
            <div class="fg">
              <label>Minutos de bloqueo por IP</label>
              <input name="bloqueo_mins" type="number" value="{config.get('bloqueo_mins','30')}" min="5" max="1440">
            </div>
            <div class="fg">
              <label>Países permitidos (cod. ISO, ej: AR)</label>
              <input name="paises_permitidos" value="{config.get('paises_permitidos','AR')}" placeholder="AR,UY,CL">
            </div>
            <div class="fg">
              <label>Timeout de sesión (minutos)</label>
              <input name="session_timeout_mins" type="number" value="{config.get('session_timeout_mins','120')}" min="15" max="480">
            </div>
            <div class="fg">
              <label>Alertas WhatsApp</label>
              <select name="alerta_whatsapp">
                <option value="1" {"selected" if config.get('alerta_whatsapp','1')=='1' else ""}>Habilitadas</option>
                <option value="0" {"selected" if config.get('alerta_whatsapp','1')=='0' else ""}>Deshabilitadas</option>
              </select>
            </div>
          </div>
          <div class="info-box" style="margin-bottom:14px">
            📱 Alertas de seguridad → tu celular personal (<b>{WHATSAPP_ADMIN}</b>).<br>
                    📱 Mensajes a clientes → celular del estudio (<b>{WHATSAPP_ESTUDIO}</b>).<br>
            🌎 Bloqueo por país usando ip-api.com.<br>
            🔐 Datos sensibles (CUIT, teléfono, email) encriptados con Fernet AES-128.<br>
            🤖 Rate limiting activo: anti-bots automático.
          </div>
          <button class="btn btn-p">Guardar configuración</button>
        </form>
      </div>

      <div class="fcard">
        <h3>🔐 Google Authenticator (2FA)</h3>
        <div class="info-box" style="margin-bottom:14px">
          Para activar el 2FA en un usuario:<br>
          1. Hacé clic en "Activar 2FA" en la lista de usuarios (pestaña Usuarios)<br>
          2. El usuario debe abrir Google Authenticator → Agregar cuenta → Escanear QR<br>
          3. Ve a <a href="/configuracion/qr_2fa?uid=ID" style="color:var(--primary)">Ver QR de usuario</a> para obtener el código QR
        </div>
        <a href="/configuracion/mi_2fa" class="btn btn-b">Ver mi QR de 2FA</a>
      </div>
    </div>

    <!-- Modal cambiar contraseña -->
    <div class="mo" id="mc"><div class="modal">
      <h3>🔑 Cambiar Contraseña</h3>
      <p class="msub" id="mc-sub"></p>
      <form method="post">
        <input type="hidden" name="accion" value="cambiar_clave">
        <input type="hidden" name="uid" id="mc-uid">
        <div class="fg" style="margin-bottom:12px">
          <label>Nueva contraseña</label>
          <input name="nueva_clave" id="mc-clave1" type="password" placeholder="Mínimo 6 caracteres" required>
        </div>
        <div class="fg" style="margin-bottom:14px">
          <label>Confirmar contraseña</label>
          <input name="confirmar_clave" id="mc-clave2" type="password" placeholder="Repetir contraseña" required>
        </div>
        <div id="mc-match" style="font-size:.78rem;margin-bottom:8px"></div>
        <div class="mact">
          <button type="button" class="btn btn-o" onclick="closeM('mc')">Cancelar</button>
          <button type="submit" class="btn btn-p" id="mc-submit">Guardar</button>
        </div>
      </form>
    </div></div>

    <!-- Modal editar usuario -->
    <div class="mo" id="me"><div class="modal">
      <h3>✏️ Editar Usuario</h3>
      <p class="msub" id="me-sub"></p>
      <form method="post">
        <input type="hidden" name="accion" value="editar_usuario">
        <input type="hidden" name="uid" id="me-uid">
        <div class="fg" style="margin-bottom:12px">
          <label>Nombre completo</label>
          <input name="nuevo_display" id="me-display" placeholder="María González">
        </div>
        <div class="fg" style="margin-bottom:14px">
          <label>Rol</label>
          <select name="nuevo_rol" id="me-rol">
            <option value="secretaria">Secretaria</option>
            <option value="supervisor">Supervisora (App móvil)</option>
            <option value="admin">Administrador</option>
          </select>
        </div>
        <div class="mact">
          <button type="button" class="btn btn-o" onclick="closeM('me')">Cancelar</button>
          <button type="submit" class="btn btn-p">Guardar cambios</button>
        </div>
      </form>
    </div></div>

    <script>
    function showTab(id,btn){{
      document.querySelectorAll('.tabpanel').forEach(p=>p.classList.remove('on'));
      document.querySelectorAll('.tab').forEach(b=>b.classList.remove('on'));
      document.getElementById(id).classList.add('on');
      btn.classList.add('on');
    }}
    function abrirClave(id,u){{
      document.getElementById('mc-sub').textContent='Cambiar clave de @'+u;
      document.getElementById('mc-uid').value=id;
      document.getElementById('mc-clave1').value='';
      document.getElementById('mc-clave2').value='';
      document.getElementById('mc-match').textContent='';
      document.getElementById('mc').classList.add('on');
      setTimeout(()=>document.getElementById('mc-clave1').focus(),100);
    }}
    function abrirEditar(id,nombre,rol){{
      document.getElementById('me-sub').textContent='@'+nombre;
      document.getElementById('me-uid').value=id;
      document.getElementById('me-display').value=nombre;
      document.getElementById('me-rol').value=rol;
      document.getElementById('me').classList.add('on');
      setTimeout(()=>document.getElementById('me-display').focus(),100);
    }}
    function selRango(){{
      var desde=document.getElementById('per-desde').value;
      var hasta=document.getElementById('per-hasta').value;
      if(!desde||!hasta){{alert('Ingresa desde y hasta');return;}}
      var d=new Date(desde+'-01'), h=new Date(hasta+'-01');
      document.querySelectorAll('#periodos-check input[name=periodos_sel]').forEach(function(chk){{
        var p=chk.value;  // formato MM/YYYY
        var parr=p.split('/');var pm=parseInt(parr[0]);var py=parseInt(parr[1]);
        var pdate=new Date(py,pm-1,1);
        chk.checked = pdate>=d && pdate<=h;
        chk.closest('label').style.background=chk.checked?'#d5f5e3':'var(--bg)';
      }});
    }}
    // Update label color on checkbox change
    document.addEventListener('change',function(e){{
      if(e.target.name==='periodos_sel'){{
        e.target.closest('label').style.background=e.target.checked?'#d5f5e3':'var(--bg)';
      }}
    }});
    function closeM(id){{document.getElementById(id).classList.remove('on')}}
    document.querySelectorAll('.mo').forEach(m=>m.addEventListener('click',e=>{{if(e.target===m)m.classList.remove('on')}}))
    // Validar que las claves coincidan en tiempo real
    function chkMatch(){{
      var c1=document.getElementById('mc-clave1').value;
      var c2=document.getElementById('mc-clave2').value;
      var el=document.getElementById('mc-match');
      var btn=document.getElementById('mc-submit');
      if(!c2){{el.textContent='';btn.disabled=false;return;}}
      if(c1===c2){{el.style.color='var(--success)';el.textContent='✓ Las contraseñas coinciden';btn.disabled=false;}}
      else{{el.style.color='var(--danger)';el.textContent='✗ No coinciden';btn.disabled=true;}}
    }}
    document.getElementById('mc-clave1').addEventListener('input',chkMatch);
    document.getElementById('mc-clave2').addEventListener('input',chkMatch);
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page("Configuración", body, "⚙️ Config")

@app.route("/configuracion/mi_2fa")
@login_req
def mi_2fa():
    conn=conectar();c=conn.cursor()
    c.execute("SELECT totp_secret,totp_habilitado FROM usuarios WHERE usuario=%s",(session.get("user"),))
    row=c.fetchone();conn.close()
    secret = row[0] if row else None
    habilitado = row[1] if row else False
    if not secret:
        # Generar y guardar secret
        secret = pyotp.random_base32()
        conn=conectar();c=conn.cursor()
        c.execute("UPDATE usuarios SET totp_secret=%s WHERE usuario=%s",(secret,session.get("user")))
        conn.commit();conn.close()
    totp_uri = pyotp.totp.TOTP(secret).provisioning_uri(
        name=session.get("user","usuario"),
        issuer_name="Estudio Carlon"
    )
    qr=qrcode.make(totp_uri)
    qb=BytesIO();qr.save(qb);qb.seek(0)
    import base64 as b64
    qr_b64 = b64.b64encode(qb.read()).decode()
    estado_badge = '<span class="sec-badge ok">2FA ACTIVO ✓</span>' if habilitado else '<span class="sec-badge warn">2FA no activado aún</span>'
    body = f"""
    <h1 class="page-title">🔐 Mi Google Authenticator</h1>
    <p class="page-sub">Configuración del segundo factor de autenticación</p>
    <div class="fcard" style="max-width:500px">
      <h3>Escanear código QR</h3>
      {estado_badge}
      <div style="margin:20px 0;text-align:center">
        <img src="data:image/png;base64,{qr_b64}" style="border:8px solid #fff;border-radius:8px;box-shadow:var(--shadow);max-width:200px">
      </div>
      <div class="info-box">
        <b>Pasos:</b><br>
        1. Instalá Google Authenticator en tu celular<br>
        2. Tocá el + → "Escanear código QR"<br>
        3. Apuntá la cámara al QR de arriba<br>
        4. Se va a agregar "Estudio Carlon - {session.get("user","")}"<br>
        5. El código cambia cada 30 segundos
      </div>
      <div style="background:#f5f3ff;border:1px solid #c4b5f7;border-radius:8px;padding:10px 14px;font-size:.82rem;margin-top:12px">
        <b>Clave manual:</b> <code style="font-size:.9rem;letter-spacing:2px">{secret}</code><br>
        <span style="color:var(--muted);font-size:.74rem">Usá esto si no podés escanear el QR</span>
      </div>
      <div style="margin-top:16px">
        <a href="/configuracion" class="btn btn-p">&larr; Volver a configuración</a>
      </div>
    </div>"""
    return page("Mi 2FA", body)

# ══════════════════════════════════════════════════════════════════════════════
#  ASISTENTE IA
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/asistente", methods=["POST"])
@login_req
def asistente():
    try:
        data = request.get_json()
        mensajes = data.get("mensajes", [])
        if not ANTHROPIC_API_KEY: return jsonify({"respuesta": None})
        payload = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 400,
            "system": "Sos el asistente del Estudio Contable Carlon de Santiago del Estero, Argentina. Respondé preguntas sobre el sistema, contabilidad e impuestos en español. Sé conciso.",
            "messages": mensajes[-8:]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01"}
        )
        resp = urllib.request.urlopen(req, timeout=15)
        result = json.loads(resp.read())
        return jsonify({"respuesta": result["content"][0]["text"]})
    except:
        return jsonify({"respuesta": None})

# ══════════════════════════════════════════════════════════════════════════════
#  CLIENTES
# ══════════════════════════════════════════════════════════════════════════════
CONDICIONES_FISCALES = [
    "Responsable Inscripto","Monotributista","Exento","No Responsable",
    "Consumidor Final","Sujeto No Categorizado","Proveedor del Exterior"
]

@app.route("/clientes", methods=["GET","POST"])
@login_req
def clientes():
    conn=conectar();c=conn.cursor();flash=""
    if request.method=="POST":
        nombre=request.form.get("nombre","").strip()
        cuit=request.form.get("cuit","").strip()
        tel=request.form.get("telefono","").strip()
        email=request.form.get("email","").strip()
        abono=request.form.get("abono",0) or 0
        condicion=request.form.get("condicion_fiscal","Responsable Inscripto")
        actividad=request.form.get("actividad","").strip()
        ri = condicion == "Responsable Inscripto"
        wa_fact = request.form.get("envio_wa_facturas","0")=="1"
        c.execute("""INSERT INTO clientes(nombre,cuit,telefono,email,abono,condicion_fiscal,actividad,responsable_inscripto,envio_wa_facturas)
                     VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                  (nombre,enc(cuit),enc(tel),enc(email),abono,condicion,actividad,ri,wa_fact))
        conn.commit()
        periodo=datetime.now().strftime("%m/%Y")
        c.execute("SELECT id FROM clientes WHERE nombre=%s ORDER BY id DESC LIMIT 1",(nombre,))
        row=c.fetchone()
        if row:
            c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,%s,0)",(row[0],periodo,float(abono) if abono else 0))
            conn.commit();registrar_auditoria("NUEVO CLIENTE",f"CUIT:{cuit} Cond:{condicion} Hon:{fmt(abono)}",row[0],nombre)
        flash=f'<div class="flash fok">✅ Cliente {nombre} agregado</div>'

    tab_cl=request.args.get("tab","activos")
    if tab_cl=="baja":
        c.execute("SELECT id,nombre,cuit,telefono,email,abono,condicion_fiscal,actividad,responsable_inscripto,envio_wa_facturas FROM clientes WHERE activo=FALSE ORDER BY nombre")
    else:
        c.execute("SELECT id,nombre,cuit,telefono,email,abono,condicion_fiscal,actividad,responsable_inscripto,envio_wa_facturas FROM clientes WHERE activo IS NOT FALSE ORDER BY nombre")
    data_raw=c.fetchall()
    # count baja
    c.execute("SELECT COUNT(*) FROM clientes WHERE activo=FALSE");n_baja=c.fetchone()[0]
    conn.close();es_admin=session.get("rol")=="admin"
    rows=""
    for d in data_raw:
        cid,nombre,cuit_enc,tel_enc,email_enc,abono,condicion,actividad,ri,wa_f=d
        cuit_d=dec(cuit_enc);tel_d=dec(tel_enc);email_d=dec(email_enc)
        cuit_limpio=(cuit_d or "").replace("-","").replace(" ","")
        cond_color={"Responsable Inscripto":"#185FA5","Monotributista":"#1D9E75","Exento":"#7B68EE"}.get(condicion or "","#888")
        cond_badge=f'<span style="font-size:.65rem;padding:2px 6px;border-radius:8px;background:#f0f4ff;color:{cond_color};font-weight:700">{condicion or "---"}</span>'
        ri_icon="🟢 " if ri else ""
        wa_icon='<span title="Recibe WA facturas" style="font-size:.8rem"> 📱</span>' if wa_f else ""
        btn_arca=('<a href="https://www.arca.gob.ar/landing/default.asp" target="_blank" class="btn btn-xs btn-arca" title="Ingresar ARCA">ARCA</a>'
                 +('<a href="https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do?nroCuit='+cuit_limpio+'" target="_blank" class="btn btn-xs" style="background:#6a1b9a;color:#fff;padding:3px 7px;font-size:.68rem;border-radius:6px">Const.</a>' if cuit_limpio else ""))
        btn_iibb=f'<a href="http://dgronline.dgrsantiago.gob.ar/dgronline/HPreImpCons005Libre.aspx?cuit={cuit_limpio}" target="_blank" class="btn btn-xs" style="background:#6a1b9a;color:#fff;padding:3px 8px;font-size:.71rem;border-radius:6px" title="IIBB Rentas SGO">IIBB</a>' if cuit_limpio else ""
        if tab_cl=="baja":
            if es_admin:
                btn_del=('<a href="/reactivar_cliente/'+str(cid)+'" class="btn btn-xs btn-g">Reactivar</a>'
                         +'<a href="/borrar_cliente/'+str(cid)+'" class="btn btn-xs btn-r" title="Eliminar definitivo (admin)">&#128465;</a>')
            else:
                btn_del='<a href="/reactivar_cliente/'+str(cid)+'" class="btn btn-xs btn-g">Reactivar</a>'
        else:
            # Todos pueden dar de baja, solo admin puede eliminar
            btn_del='<a href="/baja_cliente/'+str(cid)+'" class="btn btn-xs btn-o" title="Dar de baja">Dar de baja</a>'
        rows+=f'''<tr data-search="{nombre.lower()} {(cuit_d or "").lower()} {(email_d or "").lower()} {(actividad or "").lower()}">
          <td class="nm">{nombre}{ri_icon}{wa_icon}<br><span style="font-size:.7rem;color:var(--muted)">{actividad or ""}</span></td>
          <td class="mu">{cuit_d or "---"}<br>{cond_badge}</td>
          <td class="mu">{tel_d or "---"}</td>
          <td class="mu">{email_d or "---"}</td>
          <td>{fmt(abono or 0)}</td>
          <td><div style="display:flex;gap:4px;flex-wrap:wrap">
            <a href="/cuenta/{cid}" class="btn btn-xs btn-p">Cuenta</a>
            <a href="/editar_cliente/{cid}" class="btn btn-xs btn-o">Editar</a>
            {btn_arca}{btn_iibb}{btn_del}
          </div></td>
        </tr>'''

    btn_wa_masivo=('<a href="/wa_masivo" class="btn btn-wa btn-sm">📱 WA Recordatorio General</a>'
                  +' <a href="/wa_facturas_preview?tipo=facturas" class="btn btn-wa btn-sm">📂 WA Facturas</a>'
                  +' <a href="/wa_facturas_preview?tipo=cobro" class="btn btn-wa btn-sm">💰 WA Cobros</a>')
    cond_opts="".join(f'<option value="{cf}">{cf}</option>' for cf in CONDICIONES_FISCALES)
    modal='<div class="mo" id="mb"><div class="modal"><h3>Eliminar cliente?</h3><p class="msub" id="mb-nm"></p><p style="font-size:.81rem;color:var(--muted)">Se eliminan todos sus registros.</p><div class="mact"><button class="btn btn-o" onclick="closeM(&apos;mb&apos;)">Cancelar</button><a id="mb-ok" href="#" class="btn btn-r">Eliminar</a></div></div></div>' if es_admin else ""
    n_ri=sum(1 for d in data_raw if d[8])
    n_wa=sum(1 for d in data_raw if d[9])
    form_nuevo_cli='<div class="fcard"><h3>Nuevo Cliente</h3><form method="post">' if tab_cl=="activos" else ""
    form_cierre_cli='      </div>\n    </form></div>' if tab_cl=="activos" else ""
    # Pre-compute tab classes to avoid quote conflicts in f-string
    tab_act_cls = "on" if tab_cl=="activos" else ""
    tab_bja_cls = "on" if tab_cl=="baja" else ""
    n_activos = len(data_raw) if tab_cl=="activos" else ""
    sub_txt = ("Dados de baja: "+str(len(data_raw))+" clientes"
               if tab_cl=="baja" else
               str(len(data_raw))+" clientes activos · "+str(n_ri)+" Resp. Inscriptos · "+str(n_wa)+" con WA")
    wa_btns = btn_wa_masivo if tab_cl=="activos" else ""
    form_open = '<div class="fcard"><h3>Nuevo Cliente</h3><form method="post"><div class="fgrid">' if tab_cl=="activos" else ""
    form_close = '</div></form></div>' if tab_cl=="activos" else ""
    body=f"""
    <h1 class="page-title">Clientes</h1>
    <p class="page-sub">{sub_txt}</p>
    {flash}
    <div class="tabs" style="margin-bottom:0">
      <button class="tab {tab_act_cls}" onclick="window.location.href='/clientes?tab=activos'">👥 Activos</button>
      <button class="tab {tab_bja_cls}" onclick="window.location.href='/clientes?tab=baja'">🚫 Dados de baja ({n_baja})</button>
    </div>
    <div style="display:flex;gap:10px;margin:14px 0;flex-wrap:wrap">
      {wa_btns}
      <a href="/exportar/excel/clientes" class="btn btn-g btn-sm">📊 Excel Clientes</a>
    </div>
    {form_open}
        <div class="fg"><label>Nombre / Razón Social</label><input name="nombre" required placeholder="Garcia Juan"></div>
        <div class="fg"><label>CUIT</label><input name="cuit" placeholder="20-12345678-9" id="cuit-inp"></div>
        <div class="fg"><label>Condición Fiscal</label><select name="condicion_fiscal">{cond_opts}</select></div>
        <div class="fg"><label>Actividad Principal</label><input name="actividad" placeholder="Ej: Comercio minorista ropa"></div>
        <div class="fg"><label>Teléfono WhatsApp (sin 0 ni 15)</label><input name="telefono" placeholder="3855123456" id="tel-inp"></div>
        <div class="fg"><label>Email</label><input name="email" type="email" placeholder="cliente@email.com"></div>
        <div class="fg"><label>Honorarios $ / mes</label><input name="abono" type="number" placeholder="0"></div>
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:12px">
        <input type="checkbox" name="envio_wa_facturas" value="1" id="wa-chk" style="width:auto">
        <label style="font-size:.84rem;cursor:pointer" for="wa-chk">📱 Enviar recordatorio WhatsApp de facturas de compras a fin de mes</label>
      </div>
      <div id="wa-hint" style="display:none" class="info-box" style="margin-bottom:10px">✅ Este cliente recibirá WhatsApp automático a fin de mes.</div>
      <div class="info-box" style="margin-bottom:12px">🔒 CUIT, teléfono y email se guardan encriptados.</div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn btn-p">Guardar Cliente</button>
        <button type="button" class="btn btn-arca btn-sm" onclick="buscarArca()">Ver constancia ARCA</button>
        <button type="button" class="btn btn-sm" style="background:#6a1b9a;color:#fff" onclick="buscarIIBB()">Ver constancia IIBB</button>
    {form_close}
    <div class="search"><span>🔍</span><input id="bus" placeholder="Buscar por nombre, CUIT, actividad..." oninput="filt(this.value)"></div>
    <div class="dtable"><table>
      <thead><tr><th>Nombre</th><th>CUIT / Cond.</th><th>Teléfono</th><th>Email</th><th>Honorarios</th><th>Acciones</th></tr></thead>
      <tbody id="tb">{rows}</tbody>
    </table></div>{modal}
    <script>
    function filt(q){{q=q.toLowerCase();document.querySelectorAll('#tb tr').forEach(r=>r.style.display=r.dataset.search.includes(q)?'':'none')}}
    function confBorrar(id,nm,e){{e.preventDefault();document.getElementById('mb-nm').textContent=nm;document.getElementById('mb-ok').href='/borrar_cliente/'+id;document.getElementById('mb').classList.add('on')}}
    function selRango(){{
      var desde=document.getElementById('per-desde').value;
      var hasta=document.getElementById('per-hasta').value;
      if(!desde||!hasta){{alert('Ingresa desde y hasta');return;}}
      var d=new Date(desde+'-01'), h=new Date(hasta+'-01');
      document.querySelectorAll('#periodos-check input[name=periodos_sel]').forEach(function(chk){{
        var p=chk.value;  // formato MM/YYYY
        var parr=p.split('/');var pm=parseInt(parr[0]);var py=parseInt(parr[1]);
        var pdate=new Date(py,pm-1,1);
        chk.checked = pdate>=d && pdate<=h;
        chk.closest('label').style.background=chk.checked?'#d5f5e3':'var(--bg)';
      }});
    }}
    // Update label color on checkbox change
    document.addEventListener('change',function(e){{
      if(e.target.name==='periodos_sel'){{
        e.target.closest('label').style.background=e.target.checked?'#d5f5e3':'var(--bg)';
      }}
    }});
    function closeM(id){{document.getElementById(id).classList.remove('on')}}
    document.querySelectorAll('.mo').forEach(m=>m.addEventListener('click',e=>{{if(e.target===m)m.classList.remove('on')}}))
    var waChk=document.getElementById('wa-chk');if(waChk)waChk.addEventListener('change',function(){{document.getElementById('wa-hint').style.display=this.checked?'block':'none'}})
    function buscarArca(){{var c=document.getElementById('cuit-inp').value.replace(/-/g,'').replace(/ /g,'');if(!c){{alert('Ingresá el CUIT');return;}}window.open('https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do?nroCuit='+c,'_blank')}}
    function buscarIIBB(){{var c=document.getElementById('cuit-inp').value.replace(/-/g,'').replace(/ /g,'');if(!c){{alert('Ingresá el CUIT');return;}}window.open('http://dgronline.dgrsantiago.gob.ar/dgronline/HPreImpCons005Libre.aspx?cuit='+c,'_blank')}}
    </script>"""
    return page("Clientes",body,"Clientes")

@app.route("/editar_cliente/<int:id>", methods=["GET","POST"])
@login_req
def editar_cliente(id):
    conn=conectar();c=conn.cursor()
    if request.method=="POST":
        nombre=request.form.get("nombre","").strip()
        cuit=request.form.get("cuit","").strip()
        tel=request.form.get("telefono","").strip()
        email=request.form.get("email","").strip()
        abono=request.form.get("abono",0) or 0
        condicion=request.form.get("condicion_fiscal","Responsable Inscripto")
        actividad=request.form.get("actividad","").strip()
        ri = condicion == "Responsable Inscripto"
        wa_fact = request.form.get("envio_wa_facturas","0")=="1"
        c.execute("""UPDATE clientes SET nombre=%s,cuit=%s,telefono=%s,email=%s,abono=%s,
                     condicion_fiscal=%s,actividad=%s,responsable_inscripto=%s,envio_wa_facturas=%s
                     WHERE id=%s""",
                  (nombre,enc(cuit),enc(tel),enc(email),abono,condicion,actividad,ri,wa_fact,id))
        conn.commit()
        registrar_auditoria("EDICION CLIENTE",f"Actualizado {nombre} | {condicion}",id,nombre)
        conn.close();return redirect("/clientes")
    c.execute("SELECT id,nombre,cuit,telefono,email,abono,condicion_fiscal,actividad,responsable_inscripto,envio_wa_facturas FROM clientes WHERE id=%s",(id,))
    d=c.fetchone();conn.close()
    if not d: return redirect("/clientes")
    cuit_d=dec(d[2]);tel_d=dec(d[3]);email_d=dec(d[4])
    condicion=d[6] or "Responsable Inscripto";actividad=d[7] or "";ri=d[8];wa_f=d[9]
    cuit_limpio=(cuit_d or "").replace("-","").replace(" ","")
    cond_opts="".join(f'<option value="{cf}" {"selected" if cf==condicion else ""}>{cf}</option>' for cf in CONDICIONES_FISCALES)
    wa_checked="checked" if wa_f else ""
    body=f"""
    <a href="/clientes" class="btn btn-o btn-sm" style="margin-bottom:18px">&larr; Volver a Clientes</a>
    <h1 class="page-title">Editar Cliente</h1>
    <p class="page-sub">{d[1]}</p>
    <div style="display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap">
      {"<a href=\'https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do?nroCuit="+cuit_limpio+"\' target=\'_blank\' class=\'btn btn-arca btn-sm\'>Ver Constancia ARCA</a>" if cuit_limpio else ""}
      {"<a href=\'http://dgronline.dgrsantiago.gob.ar/dgronline/HPreImpCons005Libre.aspx?cuit="+cuit_limpio+"\' target=\'_blank\' class=\'btn btn-sm\' style=\'background:#6a1b9a;color:#fff\'>Ver Constancia IIBB</a>" if cuit_limpio else ""}
    </div>
    <div class="fcard"><form method="post">
      <div class="fgrid">
        <div class="fg"><label>Nombre / Razón Social</label><input name="nombre" value="{d[1] or ""}" required></div>
        <div class="fg"><label>CUIT</label><input name="cuit" value="{cuit_d or ""}"></div>
        <div class="fg"><label>Condición Fiscal</label><select name="condicion_fiscal">{cond_opts}</select></div>
        <div class="fg"><label>Actividad Principal</label><input name="actividad" value="{actividad}" placeholder="Ej: Comercio minorista ropa"></div>
        <div class="fg"><label>Teléfono WhatsApp</label><input name="telefono" value="{tel_d or ""}"></div>
        <div class="fg"><label>Email</label><input name="email" type="email" value="{email_d or ""}"></div>
        <div class="fg"><label>Honorarios $ / mes</label><input name="abono" type="number" value="{d[5] or 0}"></div>
      </div>
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px">
        <input type="checkbox" name="envio_wa_facturas" value="1" id="wa-chk" {wa_checked} style="width:auto">
        <label style="font-size:.84rem;cursor:pointer" for="wa-chk">Enviar recordatorio WA de facturas a fin de mes</label>
      </div>
      <div style="display:flex;gap:8px"><button class="btn btn-p">Guardar Cambios</button><a href="/clientes" class="btn btn-o">Cancelar</a></div>
    </form></div>"""
    return page(f"Editar - {d[1]}",body,"Clientes")

@app.route("/wa_masivo", methods=["GET","POST"])
@login_req
def wa_masivo():
    conn=conectar();c=conn.cursor()
    c.execute("""SELECT cl.id,cl.nombre,cl.telefono,cl.condicion_fiscal,
                        COALESCE(SUM(cu.debe-cu.haber),0) saldo
                 FROM clientes cl
                 LEFT JOIN cuentas cu ON cu.cliente_id=cl.id
                 WHERE cl.telefono IS NOT NULL AND cl.telefono != ''
                 GROUP BY cl.id,cl.nombre,cl.telefono,cl.condicion_fiscal
                 ORDER BY cl.nombre""")
    clientes_raw=c.fetchall(); conn.close()
    flash=""

    MSG_DEFAULT = (
        "Estimado/a le recordamos ponerse al dia con sus honorarios. "
        "De esta manera evitamos el freno de sus presentaciones y la acumulacion de deudas.\n\n"
        "Puede solicitar el detalle por este medio o llegarse al estudio.\n\n"
        "En caso de transferencia el alias de la contadora es:\n"
        "estudio.conta.carlon\n"
        "Titular Natasha Alexis Carlon"
    )

    if request.method=="POST":
        accion=request.form.get("accion","")
        mensaje=request.form.get("mensaje",MSG_DEFAULT).strip()
        seleccionados=request.form.getlist("sel")
        if accion=="enviar_seleccionados" and seleccionados:
            if not CALLMEBOT_APIKEY:
                flash='<div class="flash ferr">Configura CALLMEBOT_APIKEY en Render para envio automatico. Usa el boton WA manual por cliente.</div>'
            else:
                enviados=0; errores=[]
                for cid in seleccionados:
                    row=next((r for r in clientes_raw if str(r[0])==cid),None)
                    if not row: continue
                    _,nombre,tel_enc,_,_=row
                    tel_d=(dec(tel_enc) or "").replace(" ","").replace("-","").replace("+","").strip()
                    if not tel_d: continue
                    num=f"54{tel_d}" if not tel_d.startswith("54") else tel_d
                    nombre_c=nombre.split()[0] if nombre else "cliente"
                    msg_final=f"Estimado/a {nombre_c},\n\n"+mensaje
                    ok=enviar_whatsapp_estudio(tel_d, msg_final)
                    if ok: enviados+=1
                    else: errores.append(nombre)
                    time.sleep(0.6)
                registrar_auditoria("WA_MASIVO_GRAL",f"{enviados} enviados, {len(errores)} errores")
                err_txt=(f" · Errores: {', '.join(errores[:4])}" if errores else "")
                flash=f'<div class="flash fok">✅ {enviados} mensajes enviados{err_txt}</div>'

    # Build client rows with checkboxes
    filas=""
    for row in clientes_raw:
        cid,nombre,tel_enc,cond,saldo=row
        tel_d=dec(tel_enc) if tel_enc else "---"
        tel_limpio=(tel_d or "").replace(" ","").replace("-","")
        tiene_deuda = saldo>0
        deuda_badge=(f'<span style="color:var(--danger);font-size:.74rem;font-weight:700">{fmt(saldo)}</span>' if tiene_deuda else "")
        cond_badge=f'<span style="font-size:.68rem;background:#f0f4ff;color:#185FA5;padding:2px 6px;border-radius:6px;font-weight:600">{cond or "---"}</span>'
        # WA manual link with default message
        nombre_c=nombre.split()[0] if nombre else "cliente"
        msg_prev=f"Estimado/a {nombre_c}, "+MSG_DEFAULT
        wa_link=(f"https://wa.me/54{tel_limpio}?text={urllib.parse.quote(msg_prev)}" if tel_limpio and tel_d!="---" else "#")
        filas+=(f'''<div class="arow" style="align-items:center">
          <div style="display:flex;align-items:center;gap:10px">
            <input type="checkbox" name="sel" value="{cid}" id="chk{cid}"
              {"checked" if tiene_deuda else ""}
              style="width:18px;height:18px;cursor:pointer;accent-color:var(--primary)">
            <label for="chk{cid}" style="cursor:pointer">
              <span style="font-weight:600;color:var(--primary)">{nombre}</span>
              {cond_badge} {deuda_badge}
              <br><span style="font-size:.73rem;color:var(--muted)">📱 {tel_d}</span>
            </label>
          </div>
          <a href="{wa_link}" target="_blank" class="btn btn-wa btn-xs">WA manual</a>
        </div>''')

    n_deuda=sum(1 for r in clientes_raw if r[4]>0)

    body=f"""
    <a href="/clientes" class="btn btn-o btn-sm" style="margin-bottom:14px">&larr; Volver</a>
    <h1 class="page-title">📱 WA Masivo — Recordatorio General</h1>
    <p class="page-sub">Seleccioná los clientes y editá el mensaje antes de enviar</p>
    {flash}
    <form method="post">
      <input type="hidden" name="accion" value="enviar_seleccionados">
      <div class="fcard" style="margin-bottom:16px">
        <h3>✏️ Mensaje a enviar</h3>
        <div class="info-box" style="margin-bottom:10px;font-size:.8rem">
          El mensaje se envía como: <b>"Estimado/a [Nombre], [tu mensaje]"</b>
        </div>
        <textarea name="mensaje" rows="7"
          style="width:100%;padding:10px;border:1.5px solid var(--border);border-radius:8px;
          font-family:'DM Sans',sans-serif;font-size:.87rem;resize:vertical;outline:none;
          line-height:1.6;background:var(--bg)">{MSG_DEFAULT}</textarea>
      </div>
      <div class="fcard" style="margin-bottom:16px">
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:14px">
          <h3 style="margin-bottom:0">{len(clientes_raw)} clientes con WhatsApp</h3>
          <div style="display:flex;gap:8px;flex-wrap:wrap">
            <button type="button" onclick="selectAll(true)" class="btn btn-o btn-sm">Seleccionar todos</button>
            <button type="button" onclick="selectAll(false)" class="btn btn-o btn-sm">Deseleccionar</button>
            <button type="button" onclick="selectDeuda()" class="btn btn-a btn-sm">Solo con deuda ({n_deuda})</button>
          </div>
        </div>
        {filas}
      </div>
      <div style="display:flex;gap:10px;flex-wrap:wrap">
        <button type="submit" class="btn btn-wa">📱 Enviar a seleccionados (automático)</button>
      </div>
      {"<div class=warn-box style=margin-top:14px>⚠️ Para envio automático necesitas CALLMEBOT_APIKEY. El botón <b>WA manual</b> por cliente funciona siempre.</div>" if not CALLMEBOT_APIKEY else ""}
    </form>
    <script>
    function selectAll(v){{document.querySelectorAll('input[name=sel]').forEach(c=>c.checked=v)}}
    function selectDeuda(){{
      document.querySelectorAll('.arow').forEach(row=>{{
        var chk=row.querySelector('input[name=sel]');
        var tieneDeuda=row.querySelector('span[style*="danger"]');
        if(chk) chk.checked=!!tieneDeuda;
      }})
    }}
    </script>"""
    return page("WA Masivo", body, "Clientes")

@app.route("/wa_facturas_preview", methods=["GET","POST"])
@login_req
def wa_facturas_preview():
    conn=conectar();c=conn.cursor()
    tipo = request.args.get("tipo","facturas")
    hoy=datetime.now()
    mes_nombre=["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                "Agosto","Septiembre","Octubre","Noviembre","Diciembre"][hoy.month]
    periodo_actual=hoy.strftime("%m/%Y")

    # Cargar todos con telefono
    c.execute("""SELECT cl.id,cl.nombre,cl.telefono,cl.condicion_fiscal,cl.actividad,cl.abono,
                        cl.envio_wa_facturas,cl.responsable_inscripto,
                        COALESCE(SUM(cu.debe-cu.haber),0) saldo
                 FROM clientes cl
                 LEFT JOIN cuentas cu ON cu.cliente_id=cl.id
                 WHERE cl.telefono IS NOT NULL AND cl.telefono != ''
                 GROUP BY cl.id,cl.nombre,cl.telefono,cl.condicion_fiscal,cl.actividad,
                          cl.abono,cl.envio_wa_facturas,cl.responsable_inscripto
                 ORDER BY cl.nombre""")
    todos=c.fetchall(); conn.close()

    # Filtrar segun tipo
    if tipo=="facturas":
        # Responsables inscriptos o con WA de facturas activo
        data=[d for d in todos if d[7] or d[6]]
    elif tipo=="cobro":
        data=[d for d in todos if d[8]>0]
    else:
        data=todos

    # Mensajes por defecto
    MSG_FACTURAS=(
        "Estimado/a {nombre}, le recordamos que se acerca el vencimiento de su Declaracion Jurada.\n\n"
        "Para poder presentarla correctamente y cumplir en tiempo y forma con los plazos impositivos, "
        "le solicitamos que nos haga llegar a la brevedad sus *facturas de compras de {mes}*.\n\n"
        "Esto nos permite registrar correctamente sus creditos fiscales y evitar atrasos en las "
        "presentaciones ante ARCA/AFIP.\n\n"
        "Puede enviarnos las facturas por este WhatsApp, por mail o acercarse al estudio.\n\n"
        "Quedamos a su disposicion. Muchas gracias!\n"
        "— *Estudio Contable Carlon*"
    )
    MSG_COBRO=(
        "Estimado/a {nombre}, le recordamos que registra un saldo pendiente de *{saldo}* "
        "en concepto de honorarios profesionales correspondientes al periodo {periodo}.\n\n"
        "Le pedimos que regularice su situacion a la brevedad para poder continuar "
        "con la gestion de sus presentaciones impositivas sin inconvenientes.\n\n"
        "En caso de transferencia:\n"
        "Alias: *estudio.conta.carlon*\n"
        "Titular: Natasha Alexis Carlon\n"
        "CBU: 0110420630042013452529\n\n"
        "Puede solicitar el detalle de la deuda por este medio o acercarse al estudio.\n"
        "— *Estudio Contable Carlon*"
    )

    flash=""
    if request.method=="POST":
        accion=request.form.get("accion","")
        seleccionados=request.form.getlist("sel")
        mensaje_custom=request.form.get("mensaje_custom","").strip()

        if accion in ("enviar_seleccionados","enviar_todos") and (seleccionados or accion=="enviar_todos"):
            if not CALLMEBOT_APIKEY:
                flash='<div class="flash ferr">Configura CALLMEBOT_APIKEY en Render para envio automatico. Usa WA manual por cliente.</div>'
            else:
                ids_sel={int(x) for x in seleccionados} if accion=="enviar_seleccionados" else {d[0] for d in data}
                enviados=0; errores=[]
                for d in data:
                    cid,nombre,tel_enc,condicion,actividad,abono,wa_f,ri,saldo=d
                    if cid not in ids_sel: continue
                    tel_d=(dec(tel_enc) or "").replace(" ","").replace("-","").replace("+","").strip()
                    if not tel_d: continue
                    num=f"54{tel_d}" if not tel_d.startswith("54") else tel_d
                    nombre_c=nombre.split()[0] if nombre else "cliente"
                    if mensaje_custom:
                        msg=f"Estimado/a {nombre_c},\n\n"+mensaje_custom
                    elif tipo=="facturas":
                        msg=MSG_FACTURAS.replace("{nombre}",nombre_c).replace("{mes}",mes_nombre)
                    else:
                        msg=MSG_COBRO.replace("{nombre}",nombre_c).replace("{saldo}",fmt(saldo)).replace("{periodo}",periodo_actual)
                    ok=enviar_whatsapp_estudio(tel_d, msg)
                    if ok: enviados+=1
                    else: errores.append(nombre)
                    time.sleep(0.6)
                registrar_auditoria("WA_MASIVO",f"Tipo:{tipo} {mes_nombre}: {enviados} enviados")
                err_txt=(f" · Errores: {', '.join(errores[:4])}" if errores else "")
                flash=f'<div class="flash fok">✅ {enviados} mensajes enviados{err_txt}</div>'

    # Build client list with checkboxes
    filas=""
    for d in data:
        cid,nombre,tel_enc,condicion,actividad,abono,wa_f,ri,saldo=d
        tel_d=dec(tel_enc) if tel_enc else "---"
        tel_limpio=(tel_d or "").replace(" ","").replace("-","")
        nombre_c=nombre.split()[0] if nombre else "cliente"
        # Default message for manual WA
        if tipo=="facturas":
            msg_prev=MSG_FACTURAS.replace("{nombre}",nombre_c).replace("{mes}",mes_nombre).replace("\n","\n")
        else:
            msg_prev=MSG_COBRO.replace("{nombre}",nombre_c).replace("{saldo}",fmt(saldo)).replace("{periodo}",periodo_actual).replace("\n","\n")
        wa_link=(f"https://wa.me/54{tel_limpio}?text={urllib.parse.quote(msg_prev)}"
                 if tel_limpio and tel_d!="---" else "#")
        # badges
        cond_col={"Responsable Inscripto":"#185FA5","Monotributista":"#1D9E75"}.get(condicion or "","#888")
        cond_b=f'<span style="font-size:.67rem;background:#f0f4ff;color:{cond_col};padding:2px 6px;border-radius:6px;font-weight:700">{condicion or "---"}</span>'
        saldo_b=(f'<span style="color:var(--danger);font-size:.75rem;font-weight:700"> · Debe: {fmt(saldo)}</span>' if saldo>0 else "")
        # checkbox - pre-checked by default
        filas+=(f'''<div class="arow" style="align-items:center">
          <div style="display:flex;align-items:center;gap:10px">
            <input type="checkbox" name="sel" value="{cid}" id="c{cid}" checked
              style="width:18px;height:18px;cursor:pointer;accent-color:var(--primary)">
            <label for="c{cid}" style="cursor:pointer">
              <span style="font-weight:600;color:var(--primary)">{nombre}</span>
              {cond_b}{saldo_b}<br>
              <span style="font-size:.73rem;color:var(--muted)">📱 {tel_d}</span>
            </label>
          </div>
          <a href="{wa_link}" target="_blank" class="btn btn-wa btn-xs">📱 WA manual</a>
        </div>''')

    # Tab titles
    tabs=('<div style="display:flex;gap:6px;margin-bottom:16px;flex-wrap:wrap">'          +f'<a href="/wa_facturas_preview?tipo=facturas" class="btn {"btn-wa" if tipo=="facturas" else "btn-o"} btn-sm">📂 Facturas Compras (RI)</a>'          +f'<a href="/wa_facturas_preview?tipo=cobro" class="btn {"btn-wa" if tipo=="cobro" else "btn-o"} btn-sm">💰 Cobro Honorarios</a>'          +'<a href="/wa_masivo" class="btn btn-o btn-sm">📋 Mensaje General</a>'          +'</div>')

    tipo_titulo={"facturas":"📂 Recordatorio Facturas de Compras","cobro":"💰 Recordatorio de Cobro Honorarios"}.get(tipo,"📱 WhatsApp")
    msg_defecto=(MSG_FACTURAS.replace("{nombre}","[Nombre]").replace("{mes}",mes_nombre)
                 if tipo=="facturas" else
                 MSG_COBRO.replace("{nombre}","[Nombre]").replace("{saldo}","$XXX").replace("{periodo}",periodo_actual))

    no_data_msg=('<div class="warn-box">⚠️ No hay clientes para este mensaje. '                 +("Habilitá el envío de WA en la ficha de cada cliente Responsable Inscripto." if tipo=="facturas"
                   else "No hay clientes con deuda pendiente.")                 +'</div>' if not data else "")

    body=f"""
    <a href="/clientes" class="btn btn-o btn-sm" style="margin-bottom:14px">&larr; Volver</a>
    <h1 class="page-title">{tipo_titulo}</h1>
    <p class="page-sub">{mes_nombre} {hoy.year} · {len(data)} clientes</p>
    {flash}
    {tabs}
    {no_data_msg}
    <form method="post">
      <input type="hidden" name="accion" value="enviar_seleccionados">
      <div class="fcard" style="margin-bottom:14px">
        <h3>✏️ Mensaje (editable — dejar vacío para usar el predeterminado)</h3>
        <div style="background:#f0f9f4;border:1px solid #b7dfcc;border-radius:8px;padding:10px 14px;font-size:.8rem;color:#1a5c3a;margin-bottom:10px;white-space:pre-line">{msg_defecto}</div>
        <div class="fg">
          <label>Personalizar mensaje (opcional — si escribís acá reemplaza el predeterminado)</label>
          <textarea name="mensaje_custom" rows="4" placeholder="Dejá vacío para usar el mensaje de arriba..."
            style="padding:9px;border:1.5px solid var(--border);border-radius:8px;font-family:'DM Sans',sans-serif;
            font-size:.85rem;width:100%;resize:vertical;outline:none;background:var(--bg);line-height:1.55"></textarea>
        </div>
      </div>
      <div class="fcard" style="margin-bottom:14px">
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:12px">
          <h3 style="margin-bottom:0">Seleccionar destinatarios</h3>
          <div style="display:flex;gap:7px;flex-wrap:wrap">
            <button type="button" onclick="document.querySelectorAll('input[name=sel]').forEach(c=>c.checked=true)" class="btn btn-o btn-sm">✓ Todos</button>
            <button type="button" onclick="document.querySelectorAll('input[name=sel]').forEach(c=>c.checked=false)" class="btn btn-o btn-sm">✗ Ninguno</button>
          </div>
        </div>
        {filas}
      </div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center">
        <button type="submit" class="btn btn-wa">📱 Enviar a seleccionados (automático)</button>
      </div>
      {"<div class=warn-box style=margin-top:12px>⚠️ Para envio masivo automático configurá CALLMEBOT_APIKEY en Render. El botón WA manual siempre funciona.</div>" if not CALLMEBOT_APIKEY else ""}
    </form>"""
    return page("WhatsApp", body, "Clientes")


@app.route("/baja_cliente/<int:id>")
@login_req
def baja_cliente(id):
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(id,));row=c.fetchone();nombre=row[0] if row else "?"
    c.execute("UPDATE clientes SET activo=FALSE WHERE id=%s",(id,))
    conn.commit();conn.close()
    registrar_auditoria("BAJA_CLIENTE",f"{nombre} dado de baja",id,nombre)
    return redirect("/clientes")

@app.route("/reactivar_cliente/<int:id>")
@login_req
def reactivar_cliente(id):
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(id,));row=c.fetchone();nombre=row[0] if row else "?"
    c.execute("UPDATE clientes SET activo=TRUE WHERE id=%s",(id,))
    conn.commit();conn.close()
    registrar_auditoria("REACTIVAR_CLIENTE",f"{nombre} reactivado",id,nombre)
    return redirect("/clientes?tab=baja")

@app.route("/borrar_cliente/<int:id>")
@admin_req
def borrar_cliente(id):
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(id,));row=c.fetchone();nombre=row[0] if row else "?"
    c.execute("DELETE FROM cuentas WHERE cliente_id=%s",(id,));c.execute("DELETE FROM pagos WHERE cliente_id=%s",(id,));c.execute("DELETE FROM clientes WHERE id=%s",(id,))
    conn.commit();conn.close();registrar_auditoria("BAJA CLIENTE","Cliente eliminado",id,nombre);return redirect("/clientes")


@app.route("/borrar_pagos_masivo/<int:cliente_id>", methods=["POST"])
@login_req
def borrar_pagos_masivo(cliente_id):
    """Elimina multiples periodos de una vez"""
    periodos_raw=request.form.get("periodos_borrar","")
    if not periodos_raw:
        return redirect(f"/cuenta/{cliente_id}")
    periodos=[p.replace("-","/").strip() for p in periodos_raw.split(",") if p.strip()]
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(cliente_id,))
    row=c.fetchone(); nombre=row[0] if row else "?"
    eliminados=0
    for per in periodos:
        c.execute("DELETE FROM pagos WHERE cliente_id=%s AND periodo=%s",(cliente_id,per))
        c.execute("DELETE FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cliente_id,per))
        eliminados+=1
    conn.commit();conn.close()
    registrar_auditoria("BORRAR_MASIVO",f"Eliminados {eliminados} periodos: {','.join(periodos[:5])}",cliente_id,nombre)
    return redirect(f"/cuenta/{cliente_id}")

@app.route("/borrar_pago/<int:cliente_id>/<path:periodo>")
@login_req
def borrar_pago(cliente_id, periodo):
    periodo = periodo.replace("-","/")
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(cliente_id,))
    row=c.fetchone(); nombre=row[0] if row else "?"
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE cliente_id=%s AND periodo=%s",(cliente_id,periodo))
    total_pago=c.fetchone()[0]
    # Delete pagos and reset cuentas
    c.execute("DELETE FROM pagos WHERE cliente_id=%s AND periodo=%s",(cliente_id,periodo))
    c.execute("DELETE FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cliente_id,periodo))
    conn.commit();conn.close()
    registrar_auditoria("BORRAR_PAGO",f"Eliminado pago {fmt(total_pago)} periodo {periodo}",cliente_id,nombre)
    return redirect(f"/cuenta/{cliente_id}")

@app.route("/editar_pago", methods=["POST"])
@login_req
def editar_pago():
    pago_id    = request.form.get("pago_id","").strip()
    cliente_id = request.form.get("cliente_id","").strip()
    nuevo_per  = request.form.get("nuevo_periodo","").strip()
    nuevo_monto= float(request.form.get("nuevo_monto",0) or 0)
    nuevo_medio= request.form.get("nuevo_medio","Efectivo")
    nuevo_obs  = request.form.get("nuevo_obs","").strip()
    if not pago_id or not cliente_id:
        return redirect(f"/cuenta/{cliente_id}")
    conn=conectar();c=conn.cursor()
    # Obtener pago original
    c.execute("SELECT periodo,monto,cliente_id FROM pagos WHERE id=%s",(pago_id,))
    row=c.fetchone()
    if not row:
        conn.close(); return redirect(f"/cuenta/{cliente_id}")
    per_orig,monto_orig,cid=row
    # Revertir haber del periodo original
    c.execute("UPDATE cuentas SET haber=GREATEST(COALESCE(haber,0)-%s,0) WHERE cliente_id=%s AND periodo=%s",
              (monto_orig,cid,per_orig))
    # Aplicar nuevo haber al nuevo periodo
    c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cid,nuevo_per))
    if c.fetchone():
        c.execute("UPDATE cuentas SET haber=COALESCE(haber,0)+%s WHERE cliente_id=%s AND periodo=%s",
                  (nuevo_monto,cid,nuevo_per))
    else:
        c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,0,%s)",
                  (cid,nuevo_per,nuevo_monto))
    # Actualizar el registro del pago
    c.execute("""UPDATE pagos SET periodo=%s,monto=%s,medio=%s,observaciones=%s
                 WHERE id=%s""",
              (nuevo_per,nuevo_monto,nuevo_medio,nuevo_obs,pago_id))
    conn.commit(); conn.close()
    registrar_auditoria("EDITAR_PAGO",
        f"Pago #{pago_id}: {per_orig}→{nuevo_per} | {fmt(monto_orig)}→{fmt(nuevo_monto)} | {nuevo_medio}",
        int(cliente_id))
    return redirect(f"/cuenta/{cliente_id}")


# ══════════════════════════════════════════════════════════════════════════════
#  CUENTA / PAGOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/registrar_periodos/<int:cliente_id>", methods=["POST"])
@login_req
def registrar_periodos(cliente_id):
    """Registra masivamente periodos historicos para un cliente"""
    desde=request.form.get("desde","").strip()  # MM/YYYY
    hasta=request.form.get("hasta","").strip()  # MM/YYYY
    monto_unitario=float(request.form.get("monto",0) or 0)
    medio=request.form.get("medio","Transferencia -> Natasha Carlon")
    es_historico = request.form.get("es_historico","0") == "1"
    if not desde or not hasta or monto_unitario<=0:
        return redirect(f"/cuenta/{cliente_id}")
    
    conn=conectar();c=conn.cursor()
    # Generate all periods between desde and hasta
    from datetime import datetime as _dt
    try:
        d=_dt.strptime(desde,"%m/%Y")
        h=_dt.strptime(hasta,"%m/%Y")
    except:
        conn.close(); return redirect(f"/cuenta/{cliente_id}")
    
    periodos=[]
    cur=d
    while cur<=h:
        periodos.append(cur.strftime("%m/%Y"))
        if cur.month==12: cur=cur.replace(year=cur.year+1,month=1)
        else: cur=cur.replace(month=cur.month+1)
    
    registrados=0
    for per in periodos:
        c.execute("SELECT id,COALESCE(haber,0) FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cliente_id,per))
        row=c.fetchone()
        if row:
            if row[1]<monto_unitario:
                c.execute("UPDATE cuentas SET debe=%s,haber=%s WHERE cliente_id=%s AND periodo=%s",
                          (monto_unitario,monto_unitario,cliente_id,per))
        else:
            c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,%s,%s)",
                      (cliente_id,per,monto_unitario,monto_unitario))
        # Check pagos
        c.execute("SELECT id FROM pagos WHERE cliente_id=%s AND periodo=%s",(cliente_id,per))
        if not c.fetchone():
            try:
                fecha_ins = "01/01/2000 00:00:00" if es_historico else now_ar()
                obs_ins = "Saldo inicial historico" if es_historico else "Registro masivo"
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por,concepto) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (cliente_id,per,monto_unitario,medio,obs_ins,False,fecha_ins,session.get("display","sistema"),session.get("display","sistema"),"Honorarios mensuales"))
            except:
                conn.rollback()
                fecha_ins = "01/01/2000 00:00:00" if es_historico else now_ar()
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (cliente_id,per,monto_unitario,medio,"Saldo inicial" if es_historico else "Masivo",False,fecha_ins,session.get("display","sistema"),session.get("display","sistema")))
        registrados+=1
    
    conn.commit()
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(cliente_id,))
    nom=c.fetchone()
    conn.close()
    registrar_auditoria("REGISTRO MASIVO",f"Periodos {desde}-{hasta} | {registrados} periodos | ${monto_unitario:,.0f}",
                        cliente_id, nom[0] if nom else "?")
    return redirect(f"/cuenta/{cliente_id}")

@app.route("/cuenta/<int:id>", methods=["GET","POST"])
@login_req
def cuenta(id):
    conn=conectar();c=conn.cursor();flash=""
    if request.method=="POST":
        tipo_reg=request.form.get("tipo_registro","simple")
        pago=float(request.form.get("pago",0) or 0)
        medio=request.form.get("medio","Efectivo")
        obs=request.form.get("observaciones","").strip()
        facturado=request.form.get("facturado","0")=="1"
        concepto=request.form.get("concepto","Honorarios mensuales").strip()
        periodos_sel=request.form.getlist("periodos_sel")
        periodo_simple=request.form.get("periodo","").strip()
        saldo_manual=float(request.form.get("saldo_manual",0) or 0)
        c.execute("SELECT nombre FROM clientes WHERE id=%s",(id,));nom=c.fetchone();nombre_cli=nom[0] if nom else "?"

        if tipo_reg=="multiple" and periodos_sel:
            # Pago de múltiples periodos
            n_per=len(periodos_sel)
            monto_por_per=round(pago/n_per,2) if n_per>0 else pago
            periodos_str=",".join(periodos_sel)
            for per in periodos_sel:
                c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(id,per))
                row=c.fetchone()
                if row: c.execute("UPDATE cuentas SET haber=COALESCE(haber,0)+%s WHERE cliente_id=%s AND periodo=%s",(monto_por_per,id,per))
                else: c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,0,%s)",(id,per,monto_por_per))
            # Si hay saldo a favor o deudor
            if saldo_manual!=0:
                per_saldo=periodos_sel[-1]
                if saldo_manual>0:  # saldo a favor del cliente
                    c.execute("UPDATE cuentas SET haber=COALESCE(haber,0)+%s WHERE cliente_id=%s AND periodo=%s",(saldo_manual,id,per_saldo))
                else:  # saldo deudor
                    c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(id,per_saldo))
                    if not c.fetchone():
                        c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,%s,0)",(id,per_saldo,abs(saldo_manual)))
            conn.commit()
            try:
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por,concepto,periodos_incluidos) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodos_sel[0],pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display",""),concepto,periodos_str))
            except:
                conn.rollback()
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodos_sel[0],pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display","")))
            conn.commit()
            registrar_auditoria("PAGO MULTIPLE",f"Periodos:{periodos_str} | Total:{fmt(pago)} | {medio}",id,nombre_cli)
            flash=f'<div class="flash fok">Pago de {fmt(pago)} registrado por {n_per} periodos ({periodos_str})</div>'

        elif tipo_reg=="concepto_libre":
            # Recibo por concepto libre (certificación, DJ, etc.)
            periodo_uso=periodo_simple or datetime.now().strftime("%m/%Y")
            c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(id,periodo_uso))
            row=c.fetchone()
            if row: c.execute("UPDATE cuentas SET haber=COALESCE(haber,0)+%s WHERE cliente_id=%s AND periodo=%s",(pago,id,periodo_uso))
            else: c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,0,%s)",(id,periodo_uso,pago))
            conn.commit()
            try:
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por,concepto,periodos_incluidos) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodo_uso,pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display",""),concepto,""))
            except:
                conn.rollback()
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodo_uso,pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display","")))
            conn.commit()
            registrar_auditoria("PAGO CONCEPTO",f"Concepto:{concepto} | {fmt(pago)} | {medio}",id,nombre_cli)
            flash=f'<div class="flash fok">{fmt(pago)} registrado — {concepto}</div>'

        else:
            # Pago simple normal
            periodo=periodo_simple
            c.execute("SELECT id FROM cuentas WHERE cliente_id=%s AND periodo=%s",(id,periodo))
            row=c.fetchone()
            if row: c.execute("UPDATE cuentas SET haber=COALESCE(haber,0)+%s WHERE cliente_id=%s AND periodo=%s",(pago,id,periodo))
            else: c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,0,%s)",(id,periodo,pago))
            conn.commit()
            try:
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por,concepto,periodos_incluidos) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodo,pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display",""),concepto,""))
            except:
                conn.rollback()
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (id,periodo,pago,medio,obs,facturado,now_ar(),session.get("display",""),session.get("display","")))
            conn.commit()
            registrar_auditoria("PAGO REGISTRADO",f"Periodo:{periodo} | Monto:{fmt(pago)} | Medio:{medio}",id,nombre_cli)
            flash=f'<div class="flash fok">Pago de {fmt(pago)} registrado - {medio}</div>'
    c.execute("SELECT nombre,cuit,telefono,email,abono FROM clientes WHERE id=%s",(id,))
    cli=c.fetchone()
    if not cli: return "Cliente no encontrado",404
    nombre,cuit_enc,tel_enc,email_enc,abono_cli=cli
    cuit=dec(cuit_enc);tel=dec(tel_enc);email=dec(email_enc)
    abono_cli=float(abono_cli or 0)
    c.execute("SELECT periodo,COALESCE(debe,0),COALESCE(haber,0) FROM cuentas WHERE cliente_id=%s ORDER BY id DESC",(id,))
    datos=c.fetchall()
    # Historial - columnas nuevas opcionales
    try:
        c2=conn.cursor()
        c2.execute("SELECT fecha,usuario,periodo,monto,medio,facturado,observaciones,emitido_por,id,COALESCE(concepto,'Honorarios mensuales'),COALESCE(periodos_incluidos,'') FROM pagos WHERE cliente_id=%s ORDER BY id DESC LIMIT 30",(id,))
        historial=c2.fetchall()
        c2.close()
    except Exception:
        try: conn.rollback()
        except: pass
        c3=conn.cursor()
        c3.execute("SELECT fecha,usuario,periodo,monto,medio,facturado,observaciones,emitido_por,id FROM pagos WHERE cliente_id=%s ORDER BY id DESC LIMIT 30",(id,))
        historial=c3.fetchall()
        c3.close()
    # Obtener periodos que tienen al menos un pago registrado (antes de cerrar conn)
    c4=conn.cursor()
    c4.execute("SELECT DISTINCT periodo FROM pagos WHERE cliente_id=%s",(id,))
    periodos_con_pago={row[0] for row in c4.fetchall()}
    c4.close()
    conn.close()
    total_deuda=sum(max(d[1]-d[2],0) for d in datos);total_pago=sum(d[2] for d in datos)
    cuit_limpio=(cuit or "").replace("-","").replace(" ","")
    telefono=(tel or "").replace(" ","").replace("+","").strip()
    filas=""
    for d in datos:
        saldo=d[1]-d[2]
        tiene_pago = d[0] in periodos_con_pago
        if d[2]>0 and saldo<=0:
            badge='<span class="badge bp">PAGADO</span>'
        elif d[2]>0 and saldo>0:
            badge='<span class="badge bpar">PARCIAL — debe '+fmt(saldo)+'</span>'
        elif d[1]>0:
            badge='<span class="badge bd">DEBE '+fmt(d[1])+'</span>'
        elif tiene_pago:
            badge='<span class="badge bp">PAGADO</span>'
        else:
            # d[1]==0, d[2]==0 sin pagos - usar abono del cliente
            if abono_cli>0:
                badge='<span class="badge bd">DEBE '+fmt(abono_cli)+'</span>'
            else:
                badge='<span class="badge" style="background:#f0f0f0;color:#888">Sin cargo</span>'
        pu=d[0].replace("/","-")
        if not d[0]: continue
        per_esc=d[0].replace('/','-')
        if saldo>0.5:
            bp=('<button data-per="'+d[0]+'" data-sal="'+str(round(saldo))+'"'
                +' class="btn btn-xs btn-g pagarBtn">Pagar</button>')
        else:
            bp='<span style="color:var(--success);font-size:.73rem;font-weight:600">Al dia</span>'
        ba='<a href="https://www.arca.gob.ar/landing/default.asp" target="_blank" class="btn btn-xs btn-arca">ARCA</a>'
        bc=('<a href="https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do?nroCuit='+cuit_limpio+'" target="_blank" class="btn btn-xs" style="background:#6a1b9a;color:#fff;padding:3px 8px;font-size:.71rem;border-radius:6px">Const.</a>' if cuit_limpio else "")
        if telefono and saldo>0.5:
            bw=('<button data-per="'+d[0]+'" data-sal="'+str(round(saldo))+'"'
                +' class="btn btn-xs btn-wa waBtn">WA</button>')
        else:
            bw=""
        bdel=('  <a href="/borrar_pago/'+str(id)+'/'+per_esc+'" class="btn btn-xs btn-r" onclick="return confirm(\'Eliminar pago de '+str(d[0])+'?\')" title="Eliminar">🗑</a>')
        # link recibo publico para compartir
        _base=os.getenv('BASE_URL','https://estudio-web-1.onrender.com')
        recibo_url=_base+'/recibo/'+str(id)+'/'+per_esc
        # WA compartir recibo
        if telefono:
            msg_rec=('Estimado/a '+nombre+', le enviamos el recibo de pago del periodo '+d[0]+'. '
                     +'Puede verlo aqui: '+recibo_url+' '
                     +'-- Estudio Contable Carlon')
            bwr=('<a href="https://wa.me/54'+telefono+'?text='+urllib.parse.quote(msg_rec)+'" '
                 +'target="_blank" class="btn btn-xs btn-wa" title="Enviar recibo por WA">📱 Recibo</a>')
        else:
            bwr=''
        # Email compartir recibo
        if email:
            asunto=urllib.parse.quote('Recibo de pago '+d[0]+' - Estudio Carlon')
            cuerpo=urllib.parse.quote('Estimado/a '+nombre+',\n\nAdjunto encontrara su recibo de pago del periodo '+d[0]+'.\n\nPuede verlo o descargarlo aqui:\n'+recibo_url+'\n\nMuchas gracias!\n-- Estudio Contable Carlon')
            bem=('<a href="mailto:'+email+'?subject='+asunto+'&body='+cuerpo+'" '
                 +'class="btn btn-xs btn-b" title="Enviar recibo por email">✉️ Email</a>')
        else:
            bem=''
        monto_fila=d[2] if d[2]>0 else d[1]
        filas+=('<div class="arow" style="position:relative">'
               +'<label style="display:flex;align-items:center;gap:6px;cursor:pointer">'
               +'<input type="checkbox" class="per-chk" data-per="'+per_esc+'" data-monto="'+str(round(monto_fila or 0))+'"'
               +' style="width:16px;height:16px;accent-color:var(--primary);cursor:pointer">'
               +'</label>'
               +'<span class="period">'+d[0]+'</span>'
               +'<span style="font-size:.86rem">'+fmt(monto_fila)+'</span>'
               +badge
               +'<div style="display:flex;gap:5px;flex-wrap:wrap">'
               +'<a href="/recibo/'+str(id)+'/'+per_esc+'" target="_blank" class="btn btn-xs btn-o">Ver</a>'
               +'<a href="/recibo/'+str(id)+'/'+per_esc+'?download=1" class="btn btn-xs btn-o">PDF</a>'
               +bwr+bem+bp+ba+bc+bw+bdel+'</div></div>')
    hist_rows=""
    for h in historial:
        # Safety: pad h to at least 11 elements
        h = tuple(h) + (None,) * (11 - len(h))
        fact_b=('<span style="color:var(--success);font-size:.69rem;font-weight:700">Facturado</span>'
                if h[5] else '<span style="color:var(--muted);font-size:.69rem">Sin factura</span>')
        emitido=h[7] or h[1]
        pid=h[8] if h[8] else 0
        btn_edit=('<button data-pid="'+str(pid)+'" data-per="'+h[2]+'" data-med="'+h[4].replace('"','&quot;')+'" data-mon="'+str(h[3])+'" data-obs="'+str(h[6] or "").replace('"','&quot;')+'" class="btn btn-xs btn-o editBtn" title="Editar">&#9998;</button>')
        concepto_p=h[9] if h[9] else "Honorarios mensuales"
        periodos_p=h[10] if h[10] else ""
        periodo_disp=str(h[2] or '')+(f' ({periodos_p})' if periodos_p and periodos_p!=h[2] else '')
        # Boton recibo consolidado si tiene multiples periodos
        if periodos_p and ',' in periodos_p:
            # Pago multiple - generar link consolidado
            pers_esc=periodos_p.replace('/','-')
            btn_recibo_hist=(
                '<a href="/recibo_consolidado/'+str(id)+'?periodos='+pers_esc+'&total='+str(round(float(h[3] or 0)))+'" '
                +'target="_blank" class="btn btn-xs btn-g" title="Descargar recibo consolidado">📄 Recibo</a>'
            )
        elif h[2]:
            per_hist=str(h[2]).replace('/','-')
            btn_recibo_hist=(
                '<a href="/recibo/'+str(id)+'/'+per_hist+'" '
                +'target="_blank" class="btn btn-xs btn-o" title="Ver recibo">📄</a>'
            )
        else:
            btn_recibo_hist=""
        hist_rows+=(
            '<div class="logrow" style="justify-content:space-between;align-items:center">'
            +'<div style="display:flex;gap:8px;align-items:flex-start;flex:1">'
            +'<div class="log-dot"></div>'
            +'<span class="log-time">'+str(h[0] or '')+'</span>'
            +'<span class="log-user">'+emitido+'</span>'
            +'<span class="log-msg"><b>'+periodo_disp+'</b>'
            +(f' · <span style="color:var(--info);font-size:.75rem;font-weight:600">{concepto_p}</span>' if concepto_p!="Honorarios mensuales" else "")
            +' · '+fmt(h[3] or 0)+' · '+str(h[4] or '')
            +(((" · "+str(h[6])) if h[6] else "")+' '+fact_b+'</span>')
            +'</div>'
            +'<div style="display:flex;gap:4px">'+btn_recibo_hist+btn_edit+'</div>'
            +'</div>')
    medios_opts="".join(f'<option value="{m}">{m}</option>' for m in MEDIOS_PAGO)
    medios_opts2=medios_opts
    medios_opts3=medios_opts
    # Periodos con deuda para selector multiple
    periodos_deudores=[d[0] for d in datos if d[1]-d[2]>0.5]
    # Generate last 24 months for manual selection
    from datetime import datetime as _dt
    _hoy=_dt.now()
    ultimos_meses=[]
    for _m in range(24):
        _mes=(_hoy.month - _m - 1) % 12 + 1
        _anio=_hoy.year - ((_hoy.month - _m - 1) // 12 + (1 if _hoy.month-_m-1 < 0 else 0))
        ultimos_meses.append(f"{_mes:02d}/{_anio}")
    # Show periods with debt first (checked), then unchecked months
    periodos_set=set(periodos_deudores)
    periodos_deudores_html="".join(
        '<label style="display:flex;align-items:center;gap:4px;background:'
        +("var(--bg);border:2px solid var(--danger)" if p in periodos_set else "var(--bg);border:1.5px solid var(--border)")
        +';border-radius:6px;padding:4px 8px;cursor:pointer;font-size:.82rem">'
        +'<input type="checkbox" name="periodos_sel" value="'+p+'"'
        +(' checked' if p in periodos_set else '')
        +' style="width:auto">'+p+'</label>'
        for p in ultimos_meses
    )
    # Also allow manual period range input
    periodos_deudores_html = (
        '<div style="margin-bottom:10px">'
        +'<div style="display:flex;gap:6px;margin-bottom:8px;align-items:center">'
        +'<span style="font-size:.78rem;font-weight:600;color:var(--muted)">Desde:</span>'
        +'<input id="per-desde" type="month" style="font-size:.82rem;padding:3px 6px;border:1.5px solid var(--border);border-radius:6px;width:130px">'
        +'<span style="font-size:.78rem;font-weight:600;color:var(--muted)">Hasta:</span>'
        +'<input id="per-hasta" type="month" style="font-size:.82rem;padding:3px 6px;border:1.5px solid var(--border);border-radius:6px;width:130px">'
        +'<button type="button" onclick="selRango()" class="btn btn-xs btn-p">Seleccionar rango</button>'
        +'</div>'
        +'<div id="periodos-check" style="display:flex;flex-wrap:wrap;gap:5px">'
        +periodos_deudores_html
        +'</div></div>'
    )
    tel_disp=(tel or "---") if (tel or "").strip() not in ("","nan","NaN","None") else "---"
    body=f"""
    <a href="/clientes" class="btn btn-o btn-sm" style="margin-bottom:18px">&larr; Clientes</a>
    <a href="/empleados/{id}" class="btn btn-p btn-sm">👷 Empleados</a>
    <h1 class="page-title">{nombre}</h1>
    <p class="page-sub">CUIT: {cuit or "---"} - Tel: {tel_disp} - {email or "---"}</p>
    <div class="stats" style="margin-bottom:16px">
      <div class="scard g"><div class="slabel">Total Cobrado</div><div class="sval">{fmt(total_pago)}</div></div>
      <div class="scard r"><div class="slabel">Deuda Pendiente</div><div class="sval">{fmt(total_deuda)}</div></div>
    </div>
    {flash}
    <div id="sel-bar" style="display:none;background:var(--primary);color:#fff;border-radius:var(--r);padding:12px 18px;margin-bottom:12px;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px">
      <span id="sel-info" style="font-size:.9rem"></span>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button onclick="selTodos()" class="btn btn-o btn-sm" style="background:rgba(255,255,255,.15);color:#fff;border-color:rgba(255,255,255,.3)">Seleccionar todos</button>
        <button onclick="deselTodos()" class="btn btn-o btn-sm" style="background:rgba(255,255,255,.15);color:#fff;border-color:rgba(255,255,255,.3)">Deseleccionar</button>
        <button onclick="abrirReciboConsolidado()" class="btn btn-g btn-sm">📄 Recibo consolidado</button>
        {'<button onclick="abrirWAConsolidado()" class="btn btn-wa btn-sm">📱 WA recibo</button>' if telefono else ""}
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 340px;gap:16px;align-items:start" class="twocol">
      <div>
        <form id="form-borrar-masivo" method="post" action="/borrar_pagos_masivo/{id}">
          <input type="hidden" name="periodos_borrar" id="input-periodos-borrar">
        </form>
        <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center;flex-wrap:wrap">
          <button onclick="selTodos()" class="btn btn-xs btn-o">☑ Todos</button>
          <button onclick="deselTodos()" class="btn btn-xs btn-o">☐ Ninguno</button>
          <button onclick="abrirReciboConsolidado()" class="btn btn-xs btn-g" id="btn-consolid" style="display:none">📄 Recibo consolidado</button>
          <button onclick="borrarSeleccionados()" class="btn btn-xs btn-r" id="btn-borrar-sel" style="display:none">🗑 Eliminar seleccionados</button>
        </div>
        {filas}
      </div>
      <div>
        <!-- Registro masivo historico (solo admin) -->
        {'<div class="fcard" style="border:2px dashed var(--warning);margin-bottom:12px">'
        '<h3 style="color:var(--warning)">⚡ Registro histórico masivo</h3>'
        '<p style="font-size:.8rem;color:var(--muted);margin-bottom:10px">Para registrar varios períodos históricos ya cobrados de una vez</p>'
        '<form method="post" action="/registrar_periodos/'+str(id)+'">'
        '<div style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end">'
        '<div class="fg" style="flex:1;min-width:100px"><label>Desde</label><input name="desde" placeholder="MM/AAAA" value="09/2024"></div>'
        '<div class="fg" style="flex:1;min-width:100px"><label>Hasta</label><input name="hasta" placeholder="MM/AAAA" value="'+datetime.now().strftime("%m/%Y")+'"></div>'
        '<div class="fg" style="flex:1;min-width:100px"><label>Monto por período $</label><input name="monto" type="number" value="'+str(int(abono_cli or 0))+'"></div>'
        '<div class="fg" style="flex:1;min-width:120px"><label>Medio</label><select name="medio">'+medios_opts+'</select></div>'
        '<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">'
        '<input type="checkbox" name="es_historico" value="1" id="chk-hist" checked style="width:auto">'
        '<label for="chk-hist" style="font-size:.82rem;font-weight:600;color:var(--warning);text-transform:none;cursor:pointer">'
        'Es registro histórico (NO suma a caja del día)</label></div>'
        '<button class="btn btn-o btn-sm" onclick="return confirm(\'Registrar todos los periodos del rango?\')">Registrar rango</button>'
        '</div></form></div>' if True else ""}
        <div class="fcard">
          <h3>Registrar Pago</h3>
          <div class="tabs" style="margin-bottom:12px">
            <button class="tab on" onclick="showRegTab('rt-simple',this)">Honorario mensual</button>
            <button class="tab" onclick="showRegTab('rt-multiple',this)">Varios periodos</button>
            <button class="tab" onclick="showRegTab('rt-concepto',this)">Concepto libre</button>
          </div>

          <!-- TAB SIMPLE -->
          <div id="rt-simple" class="regpanel on">
            <form method="post">
              <input type="hidden" name="tipo_registro" value="simple">
              <input type="hidden" name="concepto" value="Honorarios mensuales">
              <div class="fgrid" style="grid-template-columns:1fr">
                <div class="fg"><label>Periodo (MM/AAAA)</label><input name="periodo" value="{datetime.now().strftime('%m/%Y')}" required></div>
                <div class="fg"><label>Monto $</label><input name="pago" type="number" step="0.01" required></div>
                <div class="fg"><label>Medio de pago</label><select name="medio">{medios_opts}</select></div>
                <div class="fg"><label>Observaciones</label><input name="observaciones" placeholder="Opcional"></div>
                <div class="fg" style="flex-direction:row;align-items:center;gap:8px"><input type="checkbox" name="facturado" value="1" id="chk-fact" style="width:auto"> <label style="text-transform:none;font-size:.84rem;cursor:pointer" for="chk-fact">Emitir factura ARCA</label></div>
              </div>
              <button class="btn btn-g">Registrar</button>
            </form>
          </div>

          <!-- TAB MULTIPLES PERIODOS -->
          <div id="rt-multiple" class="regpanel">
            <form method="post">
              <input type="hidden" name="tipo_registro" value="multiple">
              <input type="hidden" name="concepto" value="Honorarios mensuales">
              <div class="info-box" style="margin-bottom:10px;font-size:.79rem">Seleccioná los períodos que incluye este pago. El monto se divide proporcionalmente.</div>
              <div class="fg" style="margin-bottom:10px">
                <label>Periodos a incluir (seleccioná los que corresponden)</label>
                <div style="margin-top:6px">
                  {periodos_deudores_html}
                </div>
              </div>
              <div class="fgrid" style="grid-template-columns:1fr">
                <div class="fg"><label>Total cobrado $</label><input name="pago" type="number" step="0.01" id="monto-multi" required></div>
                <div class="fg"><label>Saldo a favor (+) o deudor (-) que queda</label><input name="saldo_manual" type="number" step="0.01" placeholder="0" id="saldo-rest"></div>
                <div class="fg"><label>Medio de pago</label><select name="medio">{medios_opts2}</select></div>
                <div class="fg"><label>Observaciones</label><input name="observaciones" placeholder="Opcional"></div>
              </div>
              <div id="multi-resumen" class="info-box" style="font-size:.8rem;margin-bottom:10px"></div>
              <button class="btn btn-g">Registrar pago multiplo</button>
            </form>
          </div>

          <!-- TAB CONCEPTO LIBRE -->
          <div id="rt-concepto" class="regpanel">
            <form method="post">
              <input type="hidden" name="tipo_registro" value="concepto_libre">
              <div class="info-box" style="margin-bottom:10px;font-size:.79rem">Para recibos por certificaciones, DDJJ, trámites, etc.</div>
              <div class="fgrid" style="grid-template-columns:1fr">
                <div class="fg"><label>Concepto *</label>
                  <input name="concepto" list="conceptos-lista" placeholder="Ej: Certificación contable" required>
                  <datalist id="conceptos-lista">
                    <option value="Certificacion contable">
                    <option value="Declaracion Jurada Ganancias">
                    <option value="Declaracion Jurada Bienes Personales">
                    <option value="Tramite AFIP/ARCA">
                    <option value="Liquidacion de sueldos">
                    <option value="Confeccion balance">
                    <option value="Inscripcion monotributo">
                    <option value="Baja monotributo">
                    <option value="Honorarios adicionales">
                  </datalist>
                </div>
                <div class="fg"><label>Periodo de referencia</label><input name="periodo" value="{datetime.now().strftime('%m/%Y')}"></div>
                <div class="fg"><label>Monto $</label><input name="pago" type="number" step="0.01" required></div>
                <div class="fg"><label>Medio de pago</label><select name="medio">{medios_opts3}</select></div>
                <div class="fg"><label>Detalle adicional</label><input name="observaciones" placeholder="Detalle del servicio..."></div>
                <div class="fg" style="flex-direction:row;align-items:center;gap:8px"><input type="checkbox" name="facturado" value="1" id="chk-fact-c" style="width:auto"> <label style="text-transform:none;font-size:.84rem;cursor:pointer" for="chk-fact-c">Emitir factura ARCA</label></div>
              </div>
              <button class="btn btn-g">Emitir recibo</button>
            </form>
          </div>
          <style>.regpanel{{display:none}}.regpanel.on{{display:block}}</style>
          <script>
          function showRegTab(id,btn){{
            document.querySelectorAll(".regpanel").forEach(p=>p.classList.remove("on"));
            document.querySelectorAll(".tab").forEach(b=>b.classList.remove("on"));
            document.getElementById(id).classList.add("on");btn.classList.add("on");}}
          // Resumen multipago
          document.addEventListener("change",function(){{
            var chks=document.querySelectorAll("#periodos-check input:checked");
            var monto=parseFloat(document.getElementById("monto-multi")?.value||0)||0;
            var saldo=parseFloat(document.getElementById("saldo-rest")?.value||0)||0;
            var n=chks.length;
            var res=document.getElementById("multi-resumen");
            if(res&&n>0&&monto>0){{
              var por_per=Math.round(monto/n);
              var pers=Array.from(chks).map(c=>c.value).join(", ");
              res.innerHTML="<b>"+n+" periodos:</b> "+pers+"<br><b>Por periodo:</b> $"+por_per.toLocaleString("es-AR")
                +(saldo!==0?"<br><b>Saldo "+(saldo>0?"a favor":"deudor")+":</b> $"+Math.abs(saldo).toLocaleString("es-AR"):"");
            }} else if(res) res.innerHTML="";
          }});
          </script>
        </div>
        <div class="fcard"><h3>Historial</h3>{hist_rows or '<p style="color:var(--muted);font-size:.84rem">Sin pagos</p>'}</div>
      </div>
    </div>
    <!-- Modal pago rápido -->
    <div class="mo" id="mp"><div class="modal">
      <h3>💰 Registrar Pago</h3><p class="msub" id="mp-sub"></p>
      <form method="post">
        <input type="hidden" name="periodo" id="mp-per">
        <div class="fg" style="margin-bottom:12px"><label>Monto $</label>
          <input name="pago" id="mp-monto" type="number" step="0.01"></div>
        <div class="fg" style="margin-bottom:14px"><label>Medio de pago</label>
          <select name="medio">{medios_opts}</select></div>
        <div class="fg" style="flex-direction:row;align-items:center;gap:8px;margin-bottom:12px">
          <input type="checkbox" name="facturado" value="1" id="chk-fact-m" style="width:auto">
          <label style="font-size:.84rem;cursor:pointer" for="chk-fact-m">Emitir factura ARCA</label>
        </div>
        <div class="mact">
          <button type="button" class="btn btn-o" onclick="closeM('mp')">Cancelar</button>
          <button type="submit" class="btn btn-g">Registrar</button>
        </div>
      </form>
    </div></div>

    <!-- Modal WA editable -->
    <div class="mo" id="mwa"><div class="modal">
      <h3>📱 Enviar WhatsApp</h3>
      <p class="msub" id="mwa-sub"></p>
      <div class="fg" style="margin-bottom:14px">
        <label>Mensaje (podés editarlo antes de enviar)</label>
        <textarea id="mwa-txt" rows="6"
          style="padding:10px;border:1.5px solid var(--border);border-radius:8px;
          font-family:'DM Sans',sans-serif;font-size:.85rem;width:100%;
          resize:vertical;outline:none;background:var(--bg);line-height:1.55"></textarea>
      </div>
      <div class="info-box" style="margin-bottom:12px;font-size:.78rem">
        Al hacer clic en <b>Enviar</b> se abre WhatsApp con el mensaje listo.
        Podés modificarlo antes de enviarlo.
      </div>
      <div class="mact">
        <button type="button" class="btn btn-o" onclick="closeM('mwa')">Cancelar</button>
        <button type="button" class="btn btn-wa" onclick="enviarWA()">📱 Abrir WhatsApp</button>
      </div>
    </div></div>

    <!-- Modal editar pago -->
    <div class="mo" id="mep"><div class="modal">
      <h3>✏️ Editar Pago</h3>
      <p class="msub" id="mep-sub"></p>
      <form method="post" action="/editar_pago">
        <input type="hidden" name="pago_id" id="mep-id">
        <input type="hidden" name="cliente_id" value="{id}">
        <div class="fg" style="margin-bottom:11px">
          <label>Periodo (MM/AAAA)</label>
          <input name="nuevo_periodo" id="mep-per" placeholder="05/2026" required
            style="font-size:.9rem">
        </div>
        <div class="fg" style="margin-bottom:11px">
          <label>Monto $</label>
          <input name="nuevo_monto" id="mep-monto" type="number" step="0.01" required>
        </div>
        <div class="fg" style="margin-bottom:11px">
          <label>Medio de pago</label>
          <select name="nuevo_medio" id="mep-medio">{medios_opts}</select>
        </div>
        <div class="fg" style="margin-bottom:14px">
          <label>Observaciones</label>
          <input name="nuevo_obs" id="mep-obs" placeholder="Opcional">
        </div>
        <div class="info-box" style="margin-bottom:12px;font-size:.78rem">
          Solo se edita el registro del pago. El saldo se recalcula automáticamente.
        </div>
        <div class="mact">
          <button type="button" class="btn btn-o" onclick="closeM('mep')">Cancelar</button>
          <button type="submit" class="btn btn-a">Guardar cambios</button>
        </div>
      </form>
    </div></div>

    <script>
    var _waTel = {json.dumps(telefono or "")};
    var _waNom = {json.dumps(nombre or "")};
    var _waCuit = {json.dumps(cuit or "")};
    var _clienteId = {id};
    var _baseUrl = "https://estudio-web-1.onrender.com";

    // Checkbox selection for consolidated recibo
    function updateSel(){{
      var chks=document.querySelectorAll('.per-chk:checked');
      var btn=document.getElementById('btn-consolid');
      var bar=document.getElementById('sel-bar');
      var info=document.getElementById('sel-info');
      if(btn){{
        if(chks.length>=1){{
          btn.style.display='inline-flex';
          btn.textContent='Recibo consolidado ('+chks.length+' periodos)';
        }} else {{
          btn.style.display='none';
        }}
      }}
      if(bar) bar.style.display=chks.length>=1?'flex':'none';
      if(info) info.textContent=chks.length+' periodo(s) seleccionado(s)';
      var btnDel=document.getElementById('btn-borrar-sel');
      if(btnDel) btnDel.style.display=chks.length>=1?'inline-flex':'none';
    }}
    // Escuchar cambios en checkboxes
    document.addEventListener('change',function(e){{
      if(e.target && e.target.classList && e.target.classList.contains('per-chk')) updateSel();
    }});
    // Tambien escuchar clicks directos por si acaso
    document.addEventListener('click',function(e){{
      if(e.target && e.target.classList && e.target.classList.contains('per-chk')) setTimeout(updateSel,50);
    }});
    function borrarSeleccionados(){{
      var chks=document.querySelectorAll('.per-chk:checked');
      if(chks.length<1){{alert('Seleccioná al menos un período');return;}}
      var periodos=Array.from(chks).map(c=>c.dataset.per).join(',');
      var perDisplay=periodos.replace(/-/g,'/');
      if(!confirm('Eliminar definitivamente los períodos: '+perDisplay+'?\nEsto no se puede deshacer.'))return;
      document.getElementById('input-periodos-borrar').value=periodos;
      document.getElementById('form-borrar-masivo').submit();
    }}
    function selTodos(){{
      document.querySelectorAll('.per-chk').forEach(function(c){{c.checked=true;}});
      updateSel();
    }}
    function deselTodos(){{
      document.querySelectorAll('.per-chk').forEach(function(c){{c.checked=false;}});
      updateSel();
    }}
    // Correr al cargar para reflejar estado inicial
    setTimeout(updateSel, 200);
    function abrirReciboConsolidado(){{
      var chks=document.querySelectorAll('.per-chk:checked');
      if(chks.length<1){{alert('Selecciona al menos un periodo');return;}}
      var periodos=Array.from(chks).map(c=>c.dataset.per).join(',');
      var total=Array.from(chks).reduce(function(s,c){{return s+parseInt(c.dataset.monto||0)}},0);
      window.open('/recibo_consolidado/'+_clienteId+'?periodos='+periodos+'&total='+total,'_blank');
    }}
    function abrirWAConsolidado(){{
      var chks=document.querySelectorAll('.per-chk:checked');
      if(chks.length<1){{alert('Selecciona al menos un periodo');return;}}
      var periodos=Array.from(chks).map(c=>c.dataset.per).join(',');
      var total=Array.from(chks).reduce(function(s,c){{return s+parseInt(c.dataset.monto||0)}},0);
      var url='/recibo_consolidado/'+_clienteId+'?periodos='+periodos+'&total='+total;
      var _base=_baseUrl;
      var msg='Estimado/a '+_waNom+', le enviamos su recibo consolidado de pago por los periodos '+periodos.replace(/-/g,'/').replace(/,/g,', ')+'. Puede verlo aqui: '+_base+url+' -- Estudio Contable Carlon';
      window.open('https://wa.me/54'+_waTel.replace(/[^0-9]/g,'')+encodeURIComponent('?text=')+encodeURIComponent(msg),'_blank');
    }}
    function abrirPago(p,s){{
      document.getElementById('mp-sub').textContent=p+' - Saldo: $'+Math.round(s).toLocaleString('es-AR');
      document.getElementById('mp-per').value=p;
      document.getElementById('mp-monto').value=Math.round(s);
      document.getElementById('mp').classList.add('on');
      setTimeout(function(){{document.getElementById('mp-monto').select()}},120);
    }}
    document.addEventListener('click',function(e){{
      var b=e.target.closest('.pagarBtn');
      if(b){{abrirPago(b.dataset.per,parseInt(b.dataset.sal));}}
      var w=e.target.closest('.waBtn');
      if(w){{abrirWA(w.dataset.per,parseInt(w.dataset.sal));}}
    }});
    // Redirigir a ARCA al tildar Facturado
    var fChk=document.getElementById('chk-fact');
    if(fChk){{
      fChk.addEventListener('change',function(){{
        if(this.checked){{
          if(confirm('Se va a abrir ARCA para emitir la factura. Continuar?'))
            window.open('https://www.arca.gob.ar/landing/default.asp','_blank');
        }}
      }});
    }}
    // Tambien el checkbox del modal rapido
    var fChkM=document.getElementById('chk-fact-m');
    if(fChkM){{
      fChkM.addEventListener('change',function(){{
        if(this.checked){{
          if(confirm('Se va a abrir ARCA para emitir la factura. Continuar?'))
            window.open('https://www.arca.gob.ar/landing/default.asp','_blank');
        }}
      }});
    }}

    function abrirWA(periodo, saldo){{
      var fmt=function(n){{return '$'+Math.round(n).toLocaleString('es-AR');}};
      var msg = "Hola " + _waNom + "!\n\n"
        + "Le informamos que registra un saldo pendiente de *" + fmt(saldo) + "* "
        + "correspondiente al periodo " + periodo + " en concepto de honorarios.\n\n"
        + "Para regularizar puede transferir a:\n"
        + "Banco Nacion\n"
        + "CBU: 0110420630042013452529\n"
        + "Alias: ESTUDIO.CONTA.CARLON\n"
        + "Titular: Alexis Natasha Carlon\n\n"
        + "Muchas gracias!\n"
        + "-- *Estudio Contable Carlon*";
      document.getElementById('mwa-sub').textContent="Para: " + _waNom + " - " + periodo;
      document.getElementById('mwa-txt').value=msg;
      document.getElementById('mwa').classList.add('on');
    }}

    function enviarWA(){{
      var msg=document.getElementById('mwa-txt').value.trim();
      if(!msg)return;
      var num="54"+_waTel.replace(/[^0-9]/g,'');
      window.open('https://wa.me/'+num+'?text='+encodeURIComponent(msg),'_blank');
      closeM('mwa');
    }}

    function editarPago(pid,periodo,medio,monto,obs){{
      document.getElementById('mep-id').value=pid;
      document.getElementById('mep-sub').textContent='Pago #'+pid+' - '+periodo;
      document.getElementById('mep-per').value=periodo;
      document.getElementById('mep-monto').value=monto;
      var sel=document.getElementById('mep-medio');
      for(var i=0;i<sel.options.length;i++){{if(sel.options[i].value===medio){{sel.selectedIndex=i;break;}}}}
      document.getElementById('mep-obs').value=obs||'';
      document.getElementById('mep').classList.add('on');
      setTimeout(function(){{document.getElementById('mep-per').focus()}},100);
    }}
    document.addEventListener('click',function(e){{
      var eb=e.target.closest('.editBtn');
      if(eb){{
        editarPago(eb.dataset.pid,eb.dataset.per,eb.dataset.med,
                   eb.dataset.mon,eb.dataset.obs);
      }}
    }});

    function selRango(){{
      var desde=document.getElementById('per-desde').value;
      var hasta=document.getElementById('per-hasta').value;
      if(!desde||!hasta){{alert('Ingresa desde y hasta');return;}}
      var d=new Date(desde+'-01'), h=new Date(hasta+'-01');
      document.querySelectorAll('#periodos-check input[name=periodos_sel]').forEach(function(chk){{
        var p=chk.value;  // formato MM/YYYY
        var parr=p.split('/');var pm=parseInt(parr[0]);var py=parseInt(parr[1]);
        var pdate=new Date(py,pm-1,1);
        chk.checked = pdate>=d && pdate<=h;
        chk.closest('label').style.background=chk.checked?'#d5f5e3':'var(--bg)';
      }});
    }}
    // Update label color on checkbox change
    document.addEventListener('change',function(e){{
      if(e.target.name==='periodos_sel'){{
        e.target.closest('label').style.background=e.target.checked?'#d5f5e3':'var(--bg)';
      }}
    }});
    function closeM(id){{document.getElementById(id).classList.remove('on')}}
    document.querySelectorAll('.mo').forEach(m=>m.addEventListener('click',e=>{{if(e.target===m)m.classList.remove('on')}}))
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page(f"Cuenta - {nombre}", body, "Clientes")

# ══════════════════════════════════════════════════════════════════════════════
#  DEUDAS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/deudas")
@login_req
def deudas():
    conn=conectar();c=conn.cursor()
    c.execute("""SELECT cl.id,cl.nombre,cl.telefono,SUM(cu.debe-cu.haber) saldo
                 FROM cuentas cu JOIN clientes cl ON cl.id=cu.cliente_id
                 GROUP BY cl.id,cl.nombre,cl.telefono
                 HAVING SUM(cu.debe-cu.haber)>0
                 ORDER BY saldo DESC""")
    data=c.fetchall();conn.close()
    total=sum(r[3] for r in data)
    cards=""
    for d in data:
        cid,nombre,tel_enc,saldo=d
        tel_d=dec(tel_enc)
        telefono=(tel_d or "").replace(" ","").replace("+","").strip()
        wa_msg=f"Hola {nombre}, le recordamos que tiene una deuda pendiente de {fmt(saldo)} con el Estudio Contable Carlon. Por favor regularizar. Gracias."
        wa_link=f"https://wa.me/{telefono}?text={wa_msg.replace(' ','%20')}" if telefono else "#"
        cards+=f'<div class="dcard"><div><span class="dname">{nombre}</span></div><div class="damt">{fmt(saldo)}</div><div style="display:flex;gap:6px"><a href="/cuenta/{cid}" class="btn btn-xs btn-p">Ver cuenta</a>{"<a href="+chr(39)+wa_link+chr(39)+" target=_blank class=btn btn-xs btn-wa>📱 WA</a>" if telefono else ""}</div></div>'
    body=f"""
    <h1 class="page-title">Deudores</h1><p class="page-sub">{len(data)} clientes con saldo pendiente · Total: {fmt(total)}</p>
    {cards or '<div class="info-box">Sin deudores 🎉</div>'}"""
    return page("Deudores",body,"Deudores")

# ══════════════════════════════════════════════════════════════════════════════
#  GASTOS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/gastos", methods=["GET","POST"])
@login_req
def gastos():
    es_admin = session.get("rol") == "admin"
    conn=conectar();c=conn.cursor();flash=""

    if request.method=="POST":
        fecha=request.form.get("fecha",now_ar())
        cat=request.form.get("categoria","Otros")
        desc=request.form.get("descripcion","").strip()
        monto=float(request.form.get("monto",0) or 0)
        # Validar que secretaria no pueda cargar categorías admin
        if not es_admin and cat in CATEGORIAS_ADMIN:
            flash='<div class="flash ferr">No tenés permiso para esa categoría</div>'
        else:
            c.execute("INSERT INTO gastos(fecha,categoria,descripcion,monto,usuario) VALUES(%s,%s,%s,%s,%s)",
                      (fecha,cat,desc,monto,session.get("display","")))
            conn.commit();registrar_auditoria("GASTO",f"{cat}: {fmt(monto)} - {desc}")
            flash=f'<div class="flash fok">Gasto registrado: {fmt(monto)}</div>'

    # Secretarias solo ven sus propias categorías permitidas
    if es_admin:
        c.execute("SELECT id,fecha,categoria,descripcion,monto,usuario FROM gastos ORDER BY id DESC LIMIT 100")
    else:
        placeholders=",".join(["%s"]*len(CATEGORIAS_SEC))
        c.execute(f"SELECT id,fecha,categoria,descripcion,monto,usuario FROM gastos WHERE categoria IN ({placeholders}) ORDER BY id DESC LIMIT 80",
                  tuple(CATEGORIAS_SEC))

    data=c.fetchall()

    # Total visible para cada rol
    total=sum(d[4] for d in data)

    # Admin ve totales por grupo
    if es_admin:
        c.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE categoria=ANY(%s)",([c2 for c2 in CATEGORIAS_ADMIN],))
        total_admin_cats=c.fetchone()[0]
        c.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE categoria=ANY(%s)",([c2 for c2 in CATEGORIAS_SEC],))
        total_sec_cats=c.fetchone()[0]
        resumen_admin=f'''
        <div class="stats" style="margin-bottom:16px">
          <div class="scard r"><div class="slabel">Gastos Estudio</div><div class="sval">{fmt(total_sec_cats)}</div></div>
          <div class="scard p"><div class="slabel">Gastos Admin/Socias</div><div class="sval">{fmt(total_admin_cats)}</div></div>
          <div class="scard o"><div class="slabel">Total General</div><div class="sval">{fmt(total)}</div></div>
        </div>
        <div class="warn-box" style="margin-bottom:14px">🔒 Las categorías <b>Sueldo, Gastos Personales, Retiros y Tarjetas</b> son solo visibles para Administradores.</div>
        '''
    else:
        resumen_admin=""

    cats_disponibles = CATEGORIAS_GASTO if es_admin else CATEGORIAS_SEC
    opts="".join(f'<option value="{cat2}">{cat2}</option>' for cat2 in cats_disponibles)

    # Colorear filas según grupo
    COLORES_CAT={"Sueldo":"#7B68EE","Gastos Personales Natasha":"#C8A96E","Gastos Personales Maira":"#C8A96E",
                 "Tarjetas":"#E67E22","Retiro Natasha":"#C8A96E","Retiro Maira":"#C8A96E"}

    rows=""
    for d in data:
        color=COLORES_CAT.get(d[2],"var(--info)")
        rows+=f'<tr><td class="mu">{d[1]}</td><td><span class="bmedio" style="background:{"#f5f0ff" if d[2] in CATEGORIAS_ADMIN else "#f0f4ff"};color:{color}">{d[2]}</span></td><td>{d[3]}</td><td style="font-weight:600;color:var(--danger)">{fmt(d[4])}</td><td class="mu">{d[5]}</td></tr>'

    conn.close()

    subtitulo = f"Total visible: {fmt(total)}" if not es_admin else f"Total general: {fmt(total)}"
    body=f"""
    <h1 class="page-title">Gastos</h1><p class="page-sub">{subtitulo}</p>{flash}
    {resumen_admin}
    <div class="fcard"><h3>Registrar Gasto</h3><form method="post">
      <div class="fgrid">
        <div class="fg"><label>Fecha</label><input name="fecha" value="{now_ar()}"></div>
        <div class="fg"><label>Categoría</label><select name="categoria">{opts}</select></div>
        <div class="fg"><label>Descripción</label><input name="descripcion" placeholder="Detalle..."></div>
        <div class="fg"><label>Monto $</label><input name="monto" type="number" step="0.01" required></div>
      </div>
      <button class="btn btn-r">Registrar Gasto</button>
    </form></div>
    <div class="dtable"><table>
      <thead><tr><th>Fecha</th><th>Categoría</th><th>Descripción</th><th>Monto</th><th>Usuario</th></tr></thead>
      <tbody>{rows or "<tr><td colspan=5 style='color:var(--muted);text-align:center;padding:20px'>Sin gastos registrados</td></tr>"}</tbody>
    </table></div>"""
    return page("Gastos",body,"Gastos")

# ══════════════════════════════════════════════════════════════════════════════
MEDIOS_CAJA = [
    ("Efectivo",        "efectivo", "#27AE60"),
    ("Cheque",          "cheque",   "#2475B0"),
    ("Dolares",         "dolares",  "#E67E22"),
    ("Transf. Natasha", "nat",      "#1A3A2A"),
    ("Transf. Maira",   "mai",      "#7B68EE"),
    ("Otro",            "otro",     "#888"),
]

def _totales_caja(fecha_hoy, usuario):
    conn=conectar();c=conn.cursor()
    c.execute("SELECT medio,SUM(monto) FROM pagos WHERE fecha LIKE %s AND emitido_por=%s AND fecha NOT LIKE '%%01/01/2000%%' GROUP BY medio",
              (f"%{fecha_hoy}%", usuario))
    filas=c.fetchall(); conn.close()
    tot = {k[0]:0.0 for k in MEDIOS_CAJA}
    for medio,monto in filas:
        m=(medio or "").lower()
        if "efectivo" in m:                              tot["Efectivo"]        += monto or 0
        elif "cheque" in m:                              tot["Cheque"]          += monto or 0
        elif "dolar" in m or "lar" in m or "u$s" in m:  tot["Dolares"]         += monto or 0
        elif "natasha" in m or ("trans" in m and "nat" in m): tot["Transf. Natasha"] += monto or 0
        elif "maira" in m or ("trans" in m and "mai" in m):   tot["Transf. Maira"]   += monto or 0
        else:                                            tot["Otro"]            += monto or 0
    tot["_fisico"] = tot["Efectivo"] + tot["Cheque"] + tot["Dolares"]
    tot["_total"]  = sum(v for k,v in tot.items() if not k.startswith("_"))
    return tot

def _ci(label, valor, color, extra=""):
    vf = "${:,.0f}".format(valor).replace(",",".")
    col = color if valor > 0 else "#ccc"
    return (
        '<div style="display:flex;flex-direction:column;align-items:center;background:var(--bg);'
        'border-radius:8px;padding:8px 12px;min-width:90px;border:1.5px solid var(--border);' + extra + '">'
        '<span style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;'
        'letter-spacing:.6px;margin-bottom:4px">' + label + '</span>'
        '<span style="font-family:\'DM Serif Display\',serif;font-size:1.1rem;font-weight:600;color:' + col + '">' + vf + '</span>'
        '</div>'
    )

@app.route("/caja", methods=["GET","POST"])
@login_req
def caja():
    conn=conectar();c=conn.cursor();flash=""
    usuario=session.get("display","");rol=session.get("rol","secretaria")
    fecha_hoy=datetime.now().strftime("%d/%m/%Y")

    if request.method=="POST":
        accion=request.form.get("accion","")
        if accion=="editar_cierre" and rol=="admin":
            cid_e=request.form.get("cierre_id")
            ef2=float(request.form.get("ef",0) or 0)
            ch2=float(request.form.get("ch",0) or 0)
            dol2=float(request.form.get("dol",0) or 0)
            nat2=float(request.form.get("nat",0) or 0)
            mai2=float(request.form.get("mai",0) or 0)
            otr2=float(request.form.get("otr",0) or 0)
            fisico2=ef2+ch2+dol2
            total2=ef2+ch2+dol2+nat2+mai2+otr2
            c.execute("UPDATE cierres_caja SET efectivo=%s,cheque=%s,dolares=%s,transferencia_nat=%s,transferencia_mai=%s,otro=%s,total_fisico=%s,total_general=%s WHERE id=%s",
                      (ef2,ch2,dol2,nat2,mai2,otr2,fisico2,total2,cid_e))
            conn.commit()
            registrar_auditoria("EDITAR_CIERRE_CAJA",
                f"Cierre #{cid_e} editado: Ef:{fmt(ef2)} Ch:{fmt(ch2)} U$S:{fmt(dol2)} Nat:{fmt(nat2)} Mai:{fmt(mai2)} Total:{fmt(total2)}")
            flash=f'<div class="flash fok">✅ Cierre editado · Nat: {fmt(nat2)} · Maira: {fmt(mai2)} · Total: {fmt(total2)}</div>'

        if accion=="cerrar_caja":
            c.execute("SELECT id FROM cierres_caja WHERE fecha=%s AND usuario=%s AND cerrado=TRUE",(fecha_hoy,usuario))
            if c.fetchone():
                flash='<div class="flash ferr">Ya cerraste tu caja hoy</div>'
            else:
                tot=_totales_caja(fecha_hoy,usuario)
                c.execute("SELECT p.fecha,cl.nombre,p.monto,p.medio,p.observaciones FROM pagos p JOIN clientes cl ON cl.id=p.cliente_id WHERE p.fecha LIKE %s AND p.emitido_por=%s ORDER BY p.id",(f"%{fecha_hoy}%",usuario))
                pagos_dia=c.fetchall()
                detalle=" | ".join(f"{p[1]}:{fmt(p[2])}({p[3]})" for p in pagos_dia)
                c.execute("SELECT id FROM cierres_caja WHERE fecha=%s AND usuario=%s",(fecha_hoy,usuario))
                existe=c.fetchone()
                v=(tot["Efectivo"],tot["Cheque"],tot["Dolares"],tot["Transf. Natasha"],tot["Transf. Maira"],tot["Otro"],tot["_fisico"],tot["_total"],detalle,datetime.now().strftime("%H:%M"))
                if existe:
                    c.execute("UPDATE cierres_caja SET efectivo=%s,cheque=%s,dolares=%s,transferencia_nat=%s,transferencia_mai=%s,otro=%s,total_fisico=%s,total_general=%s,detalle_pagos=%s,cerrado=TRUE,hora_cierre=%s WHERE id=%s",v+(existe[0],))
                else:
                    c.execute("INSERT INTO cierres_caja(fecha,usuario,efectivo,cheque,dolares,transferencia_nat,transferencia_mai,otro,total_fisico,total_general,detalle_pagos,cerrado,hora_cierre) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)",(fecha_hoy,usuario)+v)
                conn.commit()
                registrar_auditoria("CIERRE_CAJA",f"Ef:{fmt(tot['Efectivo'])} Ch:{fmt(tot['Cheque'])} U$S:{fmt(tot['Dolares'])} Nat:{fmt(tot['Transf. Natasha'])} Mai:{fmt(tot['Transf. Maira'])} Total:{fmt(tot['_total'])}")
                flash=f'<div class="flash fok">Caja cerrada · Efectivo: {fmt(tot["Efectivo"])} · Cheque: {fmt(tot["Cheque"])} · U$S: {fmt(tot["Dolares"])} · Nat: {fmt(tot["Transf. Natasha"])} · Maira: {fmt(tot["Transf. Maira"])} · <b>Total: {fmt(tot["_total"])}</b></div>'

    tot_hoy=_totales_caja(fecha_hoy,usuario)
    c.execute("SELECT id FROM cierres_caja WHERE fecha=%s AND usuario=%s AND cerrado=TRUE",(fecha_hoy,usuario))
    _cierre_row=c.fetchone()
    # Also check if there are real pagos today by this user
    c.execute("SELECT COUNT(*) FROM pagos WHERE fecha LIKE %s AND emitido_por=%s AND fecha NOT LIKE '%%2000%%'",(fecha_hoy+"%",usuario))
    _pagos_hoy=c.fetchone()[0]
    # Caja is only "cerrada" if there was a closure AND it was done today
    ya_cerro=bool(_cierre_row)

    # Cobros del dia con saldo
    c.execute("""SELECT cl.nombre,p.periodo,p.monto,p.medio,p.observaciones,
                        COALESCE(cu.debe,0),COALESCE(cu.haber,0)
                 FROM pagos p
                 JOIN clientes cl ON cl.id=p.cliente_id
                 LEFT JOIN cuentas cu ON cu.cliente_id=p.cliente_id AND cu.periodo=p.periodo
                 WHERE p.fecha LIKE %s AND p.emitido_por=%s
                 AND p.fecha NOT LIKE '%%2000%%'
                 ORDER BY p.id DESC""",(fecha_hoy+"%",usuario))
    cobros_dia=c.fetchall()

    if rol=="admin":
        # Admin ve historial completo de todos
        c.execute("SELECT id,fecha,usuario,efectivo,cheque,dolares,transferencia_nat,transferencia_mai,otro,total_fisico,total_general,cerrado,hora_cierre FROM cierres_caja ORDER BY id DESC LIMIT 40")
    else:
        # Secretaria solo ve su propio historial (sin acumular dias anteriores en totales)
        c.execute("SELECT id,fecha,usuario,efectivo,cheque,dolares,transferencia_nat,transferencia_mai,otro,total_fisico,total_general,cerrado,hora_cierre FROM cierres_caja WHERE usuario=%s AND fecha=%s ORDER BY id DESC LIMIT 5",(usuario,fecha_hoy))
    cierres=c.fetchall(); conn.close()

    # Items caja hoy - todos los medios siempre visibles
    items_hoy = (
        _ci("Efectivo",        tot_hoy["Efectivo"],        "#27AE60") +
        _ci("Cheque",          tot_hoy["Cheque"],          "#2475B0") +
        _ci("Dolares U$S",     tot_hoy["Dolares"],         "#E67E22") +
        _ci("Trans. Natasha",  tot_hoy["Transf. Natasha"], "#1A3A2A") +
        _ci("Trans. Maira",    tot_hoy["Transf. Maira"],   "#7B68EE") +
        _ci("Otro",            tot_hoy["Otro"],            "#888") +
        _ci("Fisico",          tot_hoy["_fisico"],         "#1A3A2A", "border:2px solid var(--primary)") +
        _ci("TOTAL DIA",       tot_hoy["_total"],          "#C8A96E", "border:2px solid var(--accent)")
    )

    estado_badge='<span class="estado-cerrada">Cerrada</span>' if ya_cerro else '<span class="estado-abierta">Abierta</span>'

    # Tabla cobros del dia
    if cobros_dia:
        filas_c=""
        for row in cobros_dia:
            nm,per,monto,medio,obs,debe,haber=row
            saldo=max(debe-haber,0)
            cm={"Efectivo":"#27AE60","Cheque":"#2475B0","Dolares":"#E67E22"}.get(medio,"#7B68EE")
            sd=(f'<span style="color:var(--danger);font-weight:700;font-size:.78rem">Saldo: {fmt(saldo)}</span>'
                if saldo>0.5 else
                '<span style="color:var(--success);font-size:.78rem;font-weight:700">Al dia</span>')
            filas_c+=(f'<tr><td class="nm">{nm}</td><td class="mu">{per}</td>'
                     f'<td style="font-weight:700;color:var(--success)">{fmt(monto)}</td>'
                     f'<td><span style="background:#f0f4ff;color:{cm};font-size:.72rem;padding:2px 7px;border-radius:8px;font-weight:700">{medio}</span></td>'
                     f'<td class="mu">{obs or "---"}</td>'
                     f'<td>{sd}</td></tr>')
        tabla_cobros=(f'<div class="fcard" style="margin-bottom:16px"><h3>Cobros del dia ({len(cobros_dia)})</h3>'
                     f'<div class="dtable"><table><thead><tr><th>Cliente</th><th>Periodo</th>'
                     f'<th>Cobrado</th><th>Medio</th><th>Observaciones</th><th>Saldo pendiente</th></tr></thead>'
                     f'<tbody>{filas_c}</tbody></table></div></div>')
    else:
        tabla_cobros='<div class="info-box" style="margin-bottom:16px">Sin cobros registrados hoy.</div>'

    # Historial
    cierre_html=""
    es_adm=session.get("rol")=="admin"
    for ci in cierres:
        cid_c,fci,uci,ef,ch,dol,nat,mai,otr,tf,tg,cerr,hora=ci
        its=(_ci("Efectivo",ef or 0,"#27AE60","min-width:72px;padding:5px 8px")+
             _ci("Cheque",ch or 0,"#2475B0","min-width:72px;padding:5px 8px")+
             _ci("U$S",dol or 0,"#E67E22","min-width:72px;padding:5px 8px")+
             _ci("Natasha",nat or 0,"#1A3A2A","min-width:72px;padding:5px 8px")+
             _ci("Maira",mai or 0,"#7B68EE","min-width:72px;padding:5px 8px")+
             _ci("Fisico",tf or 0,"#1A3A2A","min-width:72px;padding:5px 8px;border:2px solid var(--primary)")+
             _ci("TOTAL",tg or 0,"#C8A96E","min-width:80px;padding:5px 8px;border:2px solid var(--accent)"))
        est=(f'<span class="estado-cerrada">Cerrada {hora}</span>' if cerr
             else '<span class="estado-abierta">Abierta</span>')
        btn_editar_caja=(
            f'<button data-cid="{cid_c}" data-ef="{ef or 0}" data-ch="{ch or 0}"'
            f' data-dol="{dol or 0}" data-nat="{nat or 0}" data-mai="{mai or 0}"'
            f' data-otr="{otr or 0}" data-fci="{fci}" data-uci="{uci}"'
            f' class="btn btn-xs btn-o cajEditBtn" title="Editar cierre">✏️ Editar</button>'
        ) if es_adm else ""
        cierre_html+=(f'<div class="caja-row"><div class="caja-header">'
                     f'<div><span class="caja-user">{uci}</span><span class="caja-fecha"> · {fci}</span></div>'
                     f'<div style="display:flex;align-items:center;gap:8px">{est}{btn_editar_caja}</div>'
                     f'</div>'
                     f'<div style="display:flex;gap:7px;flex-wrap:wrap;margin-top:8px">{its}</div></div>')

    if ya_cerro:
        btn_cierre='<div class="info-box" style="margin-top:12px">Caja cerrada hoy. Manana se abre automaticamente.</div>'
    else:
        btn_cierre=('<form method="post" style="margin-top:14px"><input type="hidden" name="accion" value="cerrar_caja">'
                   '<button class="btn btn-r btn-sm" onclick="return confirm(\'Cerrar caja del dia?\')">'
                   'Cerrar Caja del Dia</button></form>')

    # Modal editar cierre (solo admin)
    medios_caja_edit=MEDIOS_PAGO
    modal_editar_caja=f"""
    <div class="mo" id="mec"><div class="modal" style="max-width:560px">
      <h3>✏️ Editar Cierre de Caja</h3>
      <p class="msub" id="mec-sub"></p>
      <form method="post">
        <input type="hidden" name="accion" value="editar_cierre">
        <input type="hidden" name="cierre_id" id="mec-id">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">
          <div class="fg"><label>Efectivo $</label><input name="ef" id="mec-ef" type="number" step="0.01"></div>
          <div class="fg"><label>Cheque $</label><input name="ch" id="mec-ch" type="number" step="0.01"></div>
          <div class="fg"><label>Dólares U$S $</label><input name="dol" id="mec-dol" type="number" step="0.01"></div>
          <div class="fg"><label>Trans. Natasha $</label><input name="nat" id="mec-nat" type="number" step="0.01"></div>
          <div class="fg"><label>Trans. Maira $</label><input name="mai" id="mec-mai" type="number" step="0.01"></div>
          <div class="fg"><label>Otro $</label><input name="otr" id="mec-otr" type="number" step="0.01"></div>
        </div>
        <div class="info-box" style="margin-bottom:12px;font-size:.78rem">
          Los totales Físico y Total se recalculan automáticamente al guardar.
        </div>
        <div class="mact">
          <button type="button" class="btn btn-o" onclick="closeM('mec')">Cancelar</button>
          <button type="submit" class="btn btn-a">Guardar cambios</button>
        </div>
      </form>
    </div></div>
    <script>
    document.addEventListener('click',function(e){{
      var b=e.target.closest('.cajEditBtn');
      if(!b)return;
      document.getElementById('mec-id').value=b.dataset.cid;
      document.getElementById('mec-sub').textContent=b.dataset.uci+' · '+b.dataset.fci;
      document.getElementById('mec-ef').value=b.dataset.ef;
      document.getElementById('mec-ch').value=b.dataset.ch;
      document.getElementById('mec-dol').value=b.dataset.dol;
      document.getElementById('mec-nat').value=b.dataset.nat;
      document.getElementById('mec-mai').value=b.dataset.mai;
      document.getElementById('mec-otr').value=b.dataset.otr;
      document.getElementById('mec').classList.add('on');
    }});
    document.querySelectorAll('.mo').forEach(m=>m.addEventListener('click',e=>{{if(e.target===m)m.classList.remove('on')}}));
    function closeM(id){{document.getElementById(id).classList.remove('on')}}
    </script>
    """ if es_adm else ""


    # Admin: ver cajas en tiempo real de todas las secretarias
    cajas_live = ""
    if rol == "admin":
        conn2=conectar();c2=conn2.cursor()
        c2.execute("SELECT DISTINCT emitido_por FROM pagos WHERE fecha LIKE %s AND emitido_por IS NOT NULL",(fecha_hoy+"%",))
        usuarios_activos=[r[0] for r in c2.fetchall() if r[0] and r[0]!=usuario]
        for uact in usuarios_activos:
            t=_totales_caja(fecha_hoy,uact)
            if t["_total"]>0:
                c2.execute("SELECT id FROM cierres_caja WHERE fecha=%s AND usuario=%s AND cerrado=TRUE",(fecha_hoy,uact))
                cerr=bool(c2.fetchone())
                est='<span class="estado-cerrada">Cerrada</span>' if cerr else '<span class="estado-abierta">Abierta</span>'
                its=(_ci("Efectivo",t["Efectivo"],"#27AE60","min-width:70px;padding:5px 8px")+
                     _ci("Cheque",t["Cheque"],"#2475B0","min-width:70px;padding:5px 8px")+
                     _ci("Natasha",t["Transf. Natasha"],"#1A3A2A","min-width:70px;padding:5px 8px")+
                     _ci("Maira",t["Transf. Maira"],"#7B68EE","min-width:70px;padding:5px 8px")+
                     _ci("TOTAL",t["_total"],"#C8A96E","min-width:80px;padding:5px 8px;border:2px solid var(--accent)"))
                cajas_live+=(f'<div class="caja-row" style="border-left:4px solid var(--info)">'
                            f'<div class="caja-header"><div><span class="caja-user">{uact}</span>'
                            f'<span class="caja-fecha"> · En tiempo real</span></div>{est}</div>'
                            f'<div style="display:flex;gap:7px;flex-wrap:wrap;margin-top:8px">{its}</div></div>')
        conn2.close()
        if cajas_live:
            cajas_live=('<div class="fcard" style="margin-bottom:16px">'
                       f'<h3>Cajas en tiempo real — hoy</h3>{cajas_live}</div>')

    body=(f'<h1 class="page-title">Caja Diaria</h1>'
          f'<p class="page-sub">Cobros del dia — {usuario} · {fecha_hoy}</p>'
          f'{flash}'
          f'<div class="fcard" style="margin-bottom:16px">'
          f'<div style="display:flex;justify-content:space-between;align-items:center;'
          f'margin-bottom:14px;flex-wrap:wrap;gap:8px">'
          f'<span style="font-family:\'DM Serif Display\',serif;font-size:1.1rem;color:var(--primary)">Mi caja hoy</span>'
          f'{estado_badge}</div>'
          f'<div style="display:flex;gap:8px;flex-wrap:wrap">{items_hoy}</div>'
          f'{btn_cierre}'
          f'</div>'
          f'{cajas_live if rol=="admin" else ""}'
          f'{tabla_cobros}'
          f'<div class="fcard"><h3>Historial de cierres</h3>'
          f'{cierre_html or "<p style=\'color:var(--muted);font-size:.84rem\'>Sin cierres</p>"}'
          f'</div>'
          +(f'<div class="fcard" style="margin-top:16px">'
            f'<h3>Cobros por mes (cierres globales)</h3>'
            f'<div style="position:relative;height:180px"><canvas id="ch-caja"></canvas></div>'
            f'<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>'
            f'<script>fetch("/api/cierres_por_mes").then(r=>r.json()).then(d=>{{new Chart(document.getElementById("ch-caja"),{{type:"bar",data:{{labels:d.labels,datasets:[{{label:"Cobrado",data:d.totales,backgroundColor:"#185FA5",borderRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{y:{{ticks:{{callback:v=>"$"+(Math.abs(v)>=1000000?Math.round(v/1000000)+"M":Math.abs(v)>=1000?Math.round(v/1000)+"k":v)}}}}}}}}}})}}).catch(()=>{{}});'
            f'</script></div>' if rol=='admin' else '')
          +modal_editar_caja)
    return page("Caja",body,"Caja")


#  REPORTES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/reportes")
@admin_req
def reportes():
    conn=conectar();c=conn.cursor()
    c.execute("SELECT periodo,COALESCE(SUM(debe),0),COALESCE(SUM(haber),0),COALESCE(SUM(debe-haber),0) FROM cuentas GROUP BY periodo ORDER BY SUBSTRING(periodo,4,4) DESC,SUBSTRING(periodo,1,2) DESC LIMIT 20")
    por_mes=c.fetchall()
    c.execute("SELECT cl.nombre,SUM(cu.debe),SUM(cu.haber),SUM(cu.debe-cu.haber) d FROM cuentas cu JOIN clientes cl ON cl.id=cu.cliente_id GROUP BY cl.nombre ORDER BY SUM(cu.debe-cu.haber) DESC LIMIT 20")
    ranking=c.fetchall()
    c.execute("SELECT medio,SUM(monto),COUNT(*) FROM pagos GROUP BY medio ORDER BY SUM(monto) DESC")
    por_medio=c.fetchall()
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Natasha%'");nat=c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Maira%'");mai=c.fetchone()[0]
    c.execute("SELECT categoria,SUM(monto) FROM gastos GROUP BY categoria ORDER BY SUM(monto) DESC")
    gastos_cat=c.fetchall()
    c.execute("SELECT fecha,usuario,accion,detalle,cliente_nombre FROM auditoria ORDER BY id DESC LIMIT 100")
    auditoria=c.fetchall()
    c.execute("SELECT nombre FROM clientes WHERE abono IS NULL OR abono=0 ORDER BY nombre")
    sin_abono=c.fetchall();conn.close()
    filas_mes="".join(f'<tr><td class="nm">{r[0]}</td><td>{fmt(r[1])}</td><td style="color:var(--success);font-weight:600">{fmt(r[2])}</td><td style="color:{"var(--danger)" if (r[3] or 0)>0 else "var(--success)"};font-weight:600">{fmt(r[3] or 0)}</td></tr>' for r in por_mes)
    filas_rank="".join(f'<tr><td class="nm">{r[0]}</td><td>{fmt(r[1])}</td><td style="color:var(--success)">{fmt(r[2])}</td><td style="color:{"var(--danger)" if (r[3] or 0)>0 else "var(--success)"}"><b>{fmt(r[3] or 0)}</b></td></tr>' for r in ranking)
    filas_medio="".join(f'<tr><td class="nm"><span class="bmedio">{r[0]}</span></td><td style="font-weight:600;color:var(--success)">{fmt(r[1])}</td><td class="mu">{r[2]} pagos</td></tr>' for r in por_medio)
    filas_gastos="".join(f'<tr><td class="nm">{r[0]}</td><td style="color:var(--danger);font-weight:600">{fmt(r[1])}</td></tr>' for r in gastos_cat)
    filas_aud="".join(f'<div class="logrow"><div class="log-dot"></div><span class="log-time">{a[0]}</span><span class="log-user">{a[1]}</span><span class="log-msg"><b>{a[2]}</b>{" - "+a[4] if a[4] else ""} - {a[3]}</span></div>' for a in auditoria)
    sin_ab=f'<div class="warn-box">{len(sin_abono)} clientes sin honorarios: {", ".join(s[0] for s in sin_abono[:12])}</div>' if sin_abono else ""
    total_s=nat+mai;pct_nat=int(nat/total_s*100) if total_s>0 else 50;pct_mai=100-pct_nat
    body=f"""
    <h1 class="page-title">Reportes</h1><p class="page-sub">Resúmenes financieros y auditoría</p>{sin_ab}
    <div class="fcard"><h3>Distribución entre Socias</h3>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:12px" class="twocol">
        <div style="text-align:center;padding:14px;background:#f0f9f4;border-radius:10px">
          <div style="font-weight:600;color:var(--primary);margin-bottom:6px">Natasha Carlon</div>
          <div style="font-family:'DM Serif Display',serif;font-size:1.8rem;color:var(--primary)">{fmt(nat)}</div>
          <div style="font-size:.74rem;color:var(--muted)">{pct_nat}% del total</div>
        </div>
        <div style="text-align:center;padding:14px;background:#f0f4ff;border-radius:10px">
          <div style="font-weight:600;color:var(--info);margin-bottom:6px">Maira Carlon</div>
          <div style="font-family:'DM Serif Display',serif;font-size:1.8rem;color:var(--info)">{fmt(mai)}</div>
          <div style="font-size:.74rem;color:var(--muted)">{pct_mai}% del total</div>
        </div>
      </div>
      <div style="display:flex;height:14px;border-radius:7px;overflow:hidden">
        <div style="width:{pct_nat}%;background:var(--primary)"></div>
        <div style="width:{pct_mai}%;background:var(--info)"></div>
      </div>
    </div>
    <div class="tabs">
      <button class="tab on" onclick="showTab('t1',this)">Por Periodo</button>
      <button class="tab" onclick="showTab('t2',this)">Clientes</button>
      <button class="tab" onclick="showTab('t3',this)">Medios de Pago</button>
      <button class="tab" onclick="showTab('t4',this)">Gastos</button>
      <button class="tab" onclick="showTab('t5',this)">Auditoría</button>
    </div>
    <div id="t1" class="tabpanel on"><div class="dtable"><table><thead><tr><th>Periodo</th><th>Facturado</th><th>Cobrado</th><th>Deuda</th></tr></thead><tbody>{filas_mes or "<tr><td colspan=4 style='color:var(--muted);text-align:center;padding:20px'>Sin datos</td></tr>"}</tbody></table></div></div>
    <div id="t2" class="tabpanel"><div class="dtable"><table><thead><tr><th>Cliente</th><th>Facturado</th><th>Cobrado</th><th>Saldo</th></tr></thead><tbody>{filas_rank or "<tr><td colspan=4 style='color:var(--muted);text-align:center;padding:20px'>Sin datos</td></tr>"}</tbody></table></div></div>
    <div id="t3" class="tabpanel"><div class="dtable"><table><thead><tr><th>Medio</th><th>Total</th><th>Cant.</th></tr></thead><tbody>{filas_medio or "<tr><td colspan=3 style='color:var(--muted);text-align:center;padding:20px'>Sin datos</td></tr>"}</tbody></table></div></div>
    <div id="t4" class="tabpanel"><div class="dtable"><table><thead><tr><th>Categoría</th><th>Total</th></tr></thead><tbody>{filas_gastos or "<tr><td colspan=2 style='color:var(--muted);text-align:center;padding:20px'>Sin gastos</td></tr>"}</tbody></table></div></div>
    <div id="t5" class="tabpanel"><div class="fcard"><h3>Registro completo</h3>{filas_aud or "<p style='color:var(--muted);font-size:.84rem'>Sin actividad</p>"}</div></div>
    <script>function showTab(id,btn){{document.querySelectorAll('.tabpanel').forEach(p=>p.classList.remove('on'));document.querySelectorAll('.tab').forEach(b=>b.classList.remove('on'));document.getElementById(id).classList.add('on');btn.classList.add('on')}}</script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>
    <div class="fcard" style="margin-top:20px">
      <h3>📥 Exportar Reportes</h3>
      <p style="color:var(--muted);font-size:.83rem;margin-bottom:14px">Descargá los datos en Excel o PDF para usar en tu computadora o enviar por mail.</p>
      <div style="display:flex;gap:10px;flex-wrap:wrap">
        <a href="/exportar/excel/periodos" class="btn btn-g">📊 Excel - Por Periodo</a>
        <a href="/exportar/excel/clientes" class="btn btn-g">📊 Excel - Clientes</a>
        <a href="/exportar/excel/gastos" class="btn btn-g">📊 Excel - Gastos</a>
        <a href="/exportar/excel/deudores" class="btn btn-g">📊 Excel - Deudores</a>
        <a href="/exportar/pdf/resumen" class="btn btn-r">📄 PDF - Resumen General</a>
        <a href="/exportar/pdf/deudores" class="btn btn-r">📄 PDF - Deudores</a>
      </div>
    </div>"""
    return page("Reportes",body,"Reportes")

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTACIONES EXCEL y PDF
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/exportar/excel/<tipo>")
@admin_req
def exportar_excel(tipo):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl no instalado. Agregá openpyxl al requirements.txt", 500

    conn=conectar();c=conn.cursor()
    wb = openpyxl.Workbook()
    ws = wb.active

    # Estilos
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A3A2A")
    sub_fill = PatternFill("solid", fgColor="C8A96E")
    num_font = Font(size=10)
    thin = Border(
        left=Side(style="thin",color="E0D8CC"),
        right=Side(style="thin",color="E0D8CC"),
        top=Side(style="thin",color="E0D8CC"),
        bottom=Side(style="thin",color="E0D8CC")
    )

    def set_header(ws, row, cols):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

    def set_title(ws, title):
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value = title
        t.font = Font(bold=True, size=14, color="1A3A2A")
        t.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Estudio Contable Carlon  ·  Generado: {now_ar()}"
        ws["A2"].font = Font(size=9, color="888888")
        ws["A2"].alignment = Alignment(horizontal="center")

    if tipo == "periodos":
        ws.title = "Por Periodo"
        set_title(ws, "Reporte por Período - Estudio Carlon")
        set_header(ws, 4, ["Periodo","Facturado","Cobrado","Deuda","% Cobrado"])
        c.execute("SELECT periodo,COALESCE(SUM(debe),0),COALESCE(SUM(haber),0),COALESCE(SUM(debe-haber),0) FROM cuentas GROUP BY periodo ORDER BY SUBSTRING(periodo,4,4) DESC,SUBSTRING(periodo,1,2) DESC")
        for ri, row in enumerate(c.fetchall(), 5):
            pct = round(float(row[2])/float(row[1])*100,1) if row[1] else 0
            ws.cell(ri,1,row[0]);ws.cell(ri,2,round(float(row[1]),2));ws.cell(ri,3,round(float(row[2]),2));ws.cell(ri,4,round(float(row[3] or 0),2));ws.cell(ri,5,f"{pct}%")
            for ci in range(1,6):
                ws.cell(ri,ci).border=thin
                ws.cell(ri,ci).alignment=Alignment(horizontal="right" if ci>1 else "left")
        col_widths=[12,16,16,16,12]
        fname = "reporte_periodos.xlsx"

    elif tipo == "clientes":
        ws.title = "Clientes"
        set_title(ws, "Reporte por Cliente - Estudio Carlon")
        set_header(ws, 4, ["Cliente","CUIT","Honorario","Facturado","Cobrado","Saldo"])
        c.execute("""SELECT cl.nombre,cl.cuit,cl.abono,
                     COALESCE(SUM(cu.debe),0),COALESCE(SUM(cu.haber),0),COALESCE(SUM(cu.debe-cu.haber),0)
                     FROM clientes cl LEFT JOIN cuentas cu ON cl.id=cu.cliente_id
                     GROUP BY cl.nombre,cl.cuit,cl.abono ORDER BY cl.nombre""")
        for ri, row in enumerate(c.fetchall(), 5):
            cuit_d = dec(row[1]) if row[1] else ""
            ws.cell(ri,1,row[0]);ws.cell(ri,2,cuit_d);ws.cell(ri,3,round(float(row[2] or 0),2))
            ws.cell(ri,4,round(float(row[3]),2));ws.cell(ri,5,round(float(row[4]),2));ws.cell(ri,6,round(float(row[5] or 0),2))
            for ci2 in range(1,7):
                ws.cell(ri,ci2).border=thin
                ws.cell(ri,ci2).alignment=Alignment(horizontal="right" if ci2>2 else "left")
                if ci2==6 and float(row[5] or 0)>0:
                    ws.cell(ri,ci2).font=Font(color="C0392B",bold=True)
        col_widths=[30,16,14,14,14,14]
        fname = "reporte_clientes.xlsx"

    elif tipo == "gastos":
        ws.title = "Gastos"
        set_title(ws, "Reporte de Gastos - Estudio Carlon")
        set_header(ws, 4, ["Fecha","Categoría","Descripción","Monto","Usuario"])
        c.execute("SELECT fecha,categoria,descripcion,monto,usuario FROM gastos ORDER BY id DESC")
        for ri, row in enumerate(c.fetchall(), 5):
            ws.cell(ri,1,row[0]);ws.cell(ri,2,row[1]);ws.cell(ri,3,row[2])
            ws.cell(ri,4,round(float(row[3]),2));ws.cell(ri,5,row[4])
            for ci2 in range(1,6):
                ws.cell(ri,ci2).border=thin
        # Totales por categoría en hoja 2
        ws2 = wb.create_sheet("Por Categoría")
        set_header(ws2, 1, ["Categoría","Total","% del Total"])
        c.execute("SELECT categoria,SUM(monto) FROM gastos GROUP BY categoria ORDER BY SUM(monto) DESC")
        gastos_cat=c.fetchall();total_g=sum(float(r[1]) for r in gastos_cat) or 1
        for ri, row in enumerate(gastos_cat, 2):
            ws2.cell(ri,1,row[0]);ws2.cell(ri,2,round(float(row[1]),2));ws2.cell(ri,3,f"{round(float(row[1])/total_g*100,1)}%")
        col_widths=[14,30,30,14,16]
        fname = "reporte_gastos.xlsx"

    elif tipo == "deudores":
        ws.title = "Deudores"
        set_title(ws, "Deudores Pendientes - Estudio Carlon")
        set_header(ws, 4, ["Cliente","Teléfono","Email","Deuda Total"])
        c.execute("""SELECT cl.nombre,cl.telefono,cl.email,SUM(cu.debe-cu.haber) d
                     FROM cuentas cu JOIN clientes cl ON cl.id=cu.cliente_id
                     GROUP BY cl.nombre,cl.telefono,cl.email HAVING SUM(cu.debe-cu.haber)>0 ORDER BY d DESC""")
        for ri, row in enumerate(c.fetchall(), 5):
            tel_d=dec(row[1]) if row[1] else "";email_d=dec(row[2]) if row[2] else ""
            ws.cell(ri,1,row[0]);ws.cell(ri,2,tel_d);ws.cell(ri,3,email_d);ws.cell(ri,4,round(float(row[3]),2))
            for ci2 in range(1,5): ws.cell(ri,ci2).border=thin
            ws.cell(ri,4).font=Font(color="C0392B",bold=True)
        col_widths=[30,16,28,14]
        fname = "deudores.xlsx"
    else:
        conn.close();return "Tipo no válido",404

    # Auto-width columns
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    conn.close()
    buf = BytesIO()
    wb.save(buf);buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@app.route("/exportar/pdf/<tipo>")
@admin_req
def exportar_pdf_reporte(tipo):
    conn=conectar();c=conn.cursor()
    buf=BytesIO();cv=canvas.Canvas(buf,pagesize=A4);w,h=A4

    def header_pdf(titulo, subtitulo=""):
        cv.setFillColorRGB(0.10,0.23,0.16);cv.rect(0,h-80,w,80,fill=1,stroke=0)
        cv.setFillColorRGB(0.78,0.66,0.43);cv.setFont("Helvetica-Bold",16)
        cv.drawString(36,h-38,titulo)
        cv.setFillColorRGB(1,1,1);cv.setFont("Helvetica",9)
        cv.drawString(36,h-56,"Estudio Contable Carlon  ·  Quimilí, Santiago del Estero")
        cv.setFont("Helvetica",8);cv.drawRightString(w-36,h-38,f"Generado: {now_ar()}")
        if subtitulo:
            cv.setFont("Helvetica-Oblique",8);cv.drawRightString(w-36,h-54,subtitulo)

    def table_header(y, cols, widths, startx=36):
        cv.setFillColorRGB(0.10,0.23,0.16);cv.rect(startx,y-14,sum(widths),16,fill=1,stroke=0)
        cv.setFillColorRGB(1,1,1);cv.setFont("Helvetica-Bold",8)
        x=startx
        for col,ww in zip(cols,widths):
            cv.drawString(x+4,y-10,str(col)[:20]);x+=ww
        return y-16

    def draw_row(y, vals, widths, startx=36, alt=False):
        if alt: cv.setFillColorRGB(0.97,0.96,0.93);cv.rect(startx,y-13,sum(widths),14,fill=1,stroke=0)
        cv.setFillColorRGB(0.15,0.15,0.15);cv.setFont("Helvetica",8.5)
        x=startx
        for val,ww in zip(vals,widths):
            cv.drawString(x+4,y-10,str(val)[:28]);x+=ww
        cv.setStrokeColorRGB(0.87,0.87,0.87);cv.line(startx,y-13,startx+sum(widths),y-13)
        return y-14

    if tipo == "resumen":
        header_pdf("Resumen General Financiero", "Todos los periodos")
        y = h-110

        # Totales generales
        c.execute("SELECT COALESCE(SUM(debe),0),COALESCE(SUM(haber),0) FROM cuentas");td,th2=c.fetchone()
        c.execute("SELECT COALESCE(SUM(monto),0) FROM gastos");tg=c.fetchone()[0]
        c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Natasha%'");nat=c.fetchone()[0]
        c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE medio ILIKE '%Maira%'");mai=c.fetchone()[0]

        cv.setFillColorRGB(0.97,0.96,0.93);cv.roundRect(36,y-70,w-72,62,6,fill=1,stroke=0)
        cv.setFont("Helvetica-Bold",10);cv.setFillColorRGB(0.10,0.23,0.16)
        cv.drawString(50,y-18,"RESUMEN FINANCIERO GENERAL")
        cv.setFont("Helvetica",9);cv.setFillColorRGB(0.2,0.2,0.2)
        col_w=(w-90)/3
        datos_res=[("Total Facturado",fmt(td)),("Total Cobrado",fmt(th2)),("Deuda Pendiente",fmt(td-th2)),
                   ("Total Gastos",fmt(tg)),("Natasha Cobró",fmt(nat)),("Maira Cobró",fmt(mai))]
        for i,(lb,val) in enumerate(datos_res):
            cx=50+(i%3)*col_w;cy=y-38-(i//3)*18
            cv.setFont("Helvetica",8);cv.setFillColorRGB(0.5,0.5,0.5);cv.drawString(cx,cy,lb)
            cv.setFont("Helvetica-Bold",9);cv.setFillColorRGB(0.10,0.23,0.16);cv.drawString(cx,cy-12,val)
        y -= 88

        # Tabla por período
        cv.setFont("Helvetica-Bold",10);cv.setFillColorRGB(0.10,0.23,0.16);cv.drawString(36,y,"Por Período")
        y -= 8
        cols=["Periodo","Facturado","Cobrado","Deuda"];widths=[80,130,130,130]
        y = table_header(y, cols, widths)
        c.execute("SELECT periodo,COALESCE(SUM(debe),0),COALESCE(SUM(haber),0),COALESCE(SUM(debe-haber),0) FROM cuentas GROUP BY periodo ORDER BY SUBSTRING(periodo,4,4) DESC,SUBSTRING(periodo,1,2) DESC LIMIT 18")
        for i,row in enumerate(c.fetchall()):
            if y < 60: cv.showPage();header_pdf("Resumen (cont.)");y=h-110
            vals=[row[0],fmt(row[1]),fmt(row[2]),fmt(row[3] or 0)]
            y=draw_row(y,vals,widths,alt=i%2==0)

    elif tipo == "deudores":
        header_pdf("Listado de Deudores", f"Fecha: {now_ar()}")
        y = h-110
        cv.setFont("Helvetica-Bold",10);cv.setFillColorRGB(0.10,0.23,0.16);cv.drawString(36,y,"Clientes con Saldo Pendiente")
        y -= 8
        cols=["Cliente","Teléfono","Deuda Total"];widths=[200,120,150]
        y = table_header(y, cols, widths)
        c.execute("""SELECT cl.nombre,cl.telefono,SUM(cu.debe-cu.haber) d
                     FROM cuentas cu JOIN clientes cl ON cl.id=cu.cliente_id
                     GROUP BY cl.nombre,cl.telefono HAVING SUM(cu.debe-cu.haber)>0 ORDER BY d DESC""")
        total_d=0
        for i,row in enumerate(c.fetchall()):
            if y < 60: cv.showPage();header_pdf("Deudores (cont.)");y=h-110
            tel_d=dec(row[1]) if row[1] else "---"
            total_d+=float(row[2])
            y=draw_row(y,[row[0],tel_d,fmt(row[2])],widths,alt=i%2==0)
        y-=10
        cv.setFont("Helvetica-Bold",10);cv.setFillColorRGB(0.10,0.23,0.16)
        cv.drawString(36,y,f"Total deuda: {fmt(total_d)}")
    else:
        conn.close();return "Tipo no válido",404

    conn.close()
    cv.save();buf.seek(0)
    fname = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buf,mimetype="application/pdf",as_attachment=True,download_name=fname)


# ══════════════════════════════════════════════════════════════════════════════
#  USUARIOS (redirige a configuracion)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/usuarios")
@admin_req
def usuarios():
    return redirect("/configuracion")

# ══════════════════════════════════════════════════════════════════════════════
#  PDF RECIBOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/recibo_consolidado/<int:cliente_id>")
def ver_recibo_consolidado(cliente_id):
    periodos_raw=request.args.get("periodos","")
    total_param=request.args.get("total","0")
    if not periodos_raw:
        return "Sin periodos seleccionados",400
    periodos=[p.replace("-","/") for p in periodos_raw.split(",")]
    try: total_manual=float(total_param)
    except: total_manual=0

    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre,cuit,abono FROM clientes WHERE id=%s",(cliente_id,))
    cli=c.fetchone()
    if not cli: conn.close(); return "Cliente no encontrado",404
    cli_nombre,cuit_enc,abono_cli=cli
    cuit_cli=dec(cuit_enc) if cuit_enc else ""
    abono_cli=float(abono_cli or 0)

    # Get amounts per period - use pagos table as source of truth
    c.execute("SELECT periodo,COALESCE(SUM(monto),0) FROM pagos WHERE cliente_id=%s GROUP BY periodo",(cliente_id,))
    pagos_por_periodo={r[0]:float(r[1]) for r in c.fetchall()}

    c.execute("SELECT periodo,COALESCE(debe,0),COALESCE(haber,0) FROM cuentas WHERE cliente_id=%s",(cliente_id,))
    cuentas_dict={(r[0]):(float(r[1]),float(r[2])) for r in c.fetchall()}

    detalles=[]
    total_real=0
    for per in periodos:
        # Priority: what was actually paid (pagos table)
        monto_pagado=pagos_por_periodo.get(per,0)
        if monto_pagado>0:
            monto=monto_pagado
        else:
            # Use haber from cuentas, or abono if nothing
            cu=cuentas_dict.get(per,(0,0))
            monto=cu[1] if cu[1]>0 else (cu[0] if cu[0]>0 else abono_cli)
        detalles.append((per,monto))
        total_real+=monto

    medio_pago=request.args.get("medio","Transferencia -> Natasha Carlon")
    monto_total=total_real if total_real>0 else total_manual

    # Register payments in DB for each period
    emitido=session.get("display","sistema")
    for per,monto in detalles:
        if monto<=0: continue
        # Update or insert cuentas
        cu=cuentas_dict.get(per)
        if cu:
            if cu[1]<monto:  # haber is less than what was paid
                c.execute("UPDATE cuentas SET haber=%s WHERE cliente_id=%s AND periodo=%s",
                          (monto,cliente_id,per))
        else:
            c.execute("INSERT INTO cuentas(cliente_id,periodo,debe,haber) VALUES(%s,%s,%s,%s)",
                      (cliente_id,per,monto,monto))
        # Check if pago already exists
        c.execute("SELECT id FROM pagos WHERE cliente_id=%s AND periodo=%s",(cliente_id,per))
        existing=c.fetchone()
        if not existing:
            periodos_str=",".join(p for p,_ in detalles)
            try:
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por,concepto,periodos_incluidos) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (cliente_id,per,monto,medio_pago,"Recibo consolidado",False,now_ar(),emitido,emitido,"Honorarios mensuales",periodos_str))
            except:
                conn.rollback()
                c.execute("INSERT INTO pagos(cliente_id,periodo,monto,medio,observaciones,facturado,fecha,usuario,emitido_por) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                          (cliente_id,per,monto,medio_pago,"Recibo consolidado",False,now_ar(),emitido,emitido))
    conn.commit()
    conn.close()

    if not detalles or monto_total<=0:
        return "No hay montos para generar el recibo",400

    pdf=generar_pdf_consolidado(cliente_id,cli_nombre,cuit_cli,detalles,monto_total)
    dl=request.args.get("download")
    fname="recibo_consolidado_"+cli_nombre.replace(" ","_")[:20]+".pdf"
    return send_file(pdf,mimetype="application/pdf",
                     as_attachment=bool(dl),download_name=fname)


def generar_pdf_consolidado(cliente_id, cli_nombre, cuit_cli, detalles, monto_total):
    """Genera recibo PDF con múltiples períodos"""
    buffer=BytesIO();cv=canvas.Canvas(buffer,pagesize=A4);w,h=A4

    # Encabezado verde
    cv.setFillColorRGB(0.10,0.23,0.16);cv.rect(0,h-140,w,140,fill=1,stroke=0)
    logo_dibujado=False
    for lp in ["logo.png","static/logo.png"]:
        if os.path.exists(lp):
            try:
                cv.drawImage(ImageReader(lp),28,h-128,width=115,height=110,
                             preserveAspectRatio=True,mask="auto")
                logo_dibujado=True
            except: pass
            break
    if not logo_dibujado:
        cv.setFillColorRGB(0.78,0.66,0.43);cv.setFont("Helvetica-Bold",26)
        cv.drawString(28,h-55,"CARLON")
        cv.setFont("Helvetica",9);cv.setFillColorRGB(1,1,1)
        cv.drawString(28,h-70,"ESTUDIO CONTABLE")

    cv.setFillColorRGB(0.78,0.66,0.43);cv.setFont("Helvetica-Bold",20)
    cv.drawString(158,h-52,"RECIBO DE PAGO")
    cv.setFillColorRGB(1,1,1);cv.setFont("Helvetica",8)
    cv.drawString(158,h-66,"Estudio Contable Carlon — Servicios Contables e Impositivos")
    numero=datetime.now().strftime("%Y%m%d%H%M%S")
    cv.setFont("Helvetica-Bold",9);cv.drawRightString(w-36,h-52,f"N° {numero}")
    cv.setFont("Helvetica",8);cv.drawRightString(w-36,h-66,datetime.now().strftime("%d/%m/%Y %H:%M"))

    # Emisor
    cv.setFillColorRGB(0.15,0.15,0.15);cv.setFont("Helvetica-Bold",8);cv.drawString(36,h-160,"EMISOR")
    cv.setFont("Helvetica",8.5);cv.drawString(36,h-174,"Estudio Contable Carlon  ·  CUIT: 27-35045505-7")
    cv.drawString(36,h-186,"Absalón Rojas s/n  ·  Quimilí, Santiago del Estero  ·  CP 3740")
    cv.setStrokeColorRGB(0.87,0.87,0.87);cv.line(36,h-198,w-36,h-198)

    # Cliente
    cv.setFillColorRGB(0.53,0.53,0.53);cv.setFont("Helvetica-Bold",8);cv.drawString(36,h-216,"CLIENTE")
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",13);cv.drawString(36,h-232,cli_nombre)
    cv.setFont("Helvetica",8.5);cv.setFillColorRGB(0.3,0.3,0.3)
    cv.drawString(36,h-246,f"CUIT: {cuit_cli or '—'}")

    # Detalle de períodos
    y=h-270
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",9)
    cv.drawString(36,y,"DETALLE DE PERÍODOS")
    y-=14
    cv.setStrokeColorRGB(0.87,0.87,0.87);cv.line(36,y,w-36,y)
    y-=4

    # Header tabla
    cv.setFillColorRGB(0.10,0.23,0.16);cv.rect(36,y-14,w-72,16,fill=1,stroke=0)
    cv.setFillColorRGB(1,1,1);cv.setFont("Helvetica-Bold",8)
    cv.drawString(44,y-10,"Período");cv.drawRightString(w-44,y-10,"Importe")
    y-=18

    total_check=0
    for i,(per,monto) in enumerate(detalles):
        if i%2==0:
            cv.setFillColorRGB(0.97,0.96,0.93);cv.rect(36,y-12,w-72,14,fill=1,stroke=0)
        cv.setFillColorRGB(0.15,0.15,0.15);cv.setFont("Helvetica",9)
        cv.drawString(44,y-9,per)
        try: mf=f"$ {float(monto):,.0f}".replace(",",".")
        except: mf=f"$ {monto}"
        cv.drawRightString(w-44,y-9,mf)
        cv.setStrokeColorRGB(0.9,0.9,0.9);cv.line(36,y-12,w-36,y-12)
        total_check+=float(monto or 0)
        y-=16

    # Total
    y-=6
    cv.setFillColorRGB(0.97,0.96,0.93);cv.roundRect(36,y-22,w-72,28,4,fill=1,stroke=0)
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",11)
    cv.drawString(44,y-12,"TOTAL ABONADO")
    try: tf=f"$ {float(monto_total):,.0f}".replace(",",".")
    except: tf=f"$ {monto_total}"
    cv.drawRightString(w-44,y-12,tf)
    y-=36

    cv.setFont("Helvetica",8);cv.setFillColorRGB(0.45,0.45,0.45)
    cv.drawString(36,y,"Recibí conforme el importe indicado en concepto de honorarios profesionales.")
    y-=30

    # Firmas
    cv.setStrokeColorRGB(0.72,0.72,0.72)
    cv.line(36,y,195,y);cv.line(w-195,y,w-36,y)
    cv.setFont("Helvetica",8);cv.setFillColorRGB(0.5,0.5,0.5)
    cv.drawString(36,y-13,"Firma");cv.drawString(w-195,y-13,"Aclaración")

    # Datos bancarios + QR
    cv.setFillColorRGB(0.97,0.96,0.93);cv.roundRect(36,36,int((w-72)*0.62),118,8,fill=1,stroke=0)
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",9);cv.drawString(52,144,"DATOS PARA TRANSFERENCIA")
    cv.setFont("Helvetica",8.5);cv.setFillColorRGB(0.2,0.2,0.2)
    for ii,ln in enumerate(["Titular: Alexis Natasha Carlon","CUIL: 27-35045505-7  ·  Banco Nacion Argentina","CBU: 0110420630042013452529","Alias: ESTUDIO.CONTA.CARLON"]):
        cv.drawString(52,126-ii*16,ln)

    # QR
    try: monto_int=int(float(monto_total))
    except: monto_int=0
    qr=qrcode.QRCode(version=2,error_correction=qrcode.constants.ERROR_CORRECT_M,box_size=5,border=2)
    qr.add_data("0110420630042013452529")
    qr.make(fit=True)
    qr_img=qr.make_image(fill_color="black",back_color="white")
    qb=BytesIO();qr_img.save(qb,"PNG");qb.seek(0)
    qr_x=w-150;qr_y=30;qr_sz=116
    cv.setFillColorRGB(1,1,1);cv.roundRect(qr_x-5,qr_y-5,qr_sz+10,qr_sz+10,6,fill=1,stroke=0)
    cv.drawImage(ImageReader(qb),qr_x,qr_y,width=qr_sz,height=qr_sz)
    cv.setFont("Helvetica-Bold",7);cv.setFillColorRGB(0.10,0.23,0.16)
    cv.drawCentredString(qr_x+qr_sz//2,qr_y-10,"Escanear con cualquier home banking")
    cv.setFont("Helvetica",6.5);cv.setFillColorRGB(0.4,0.4,0.4)
    cv.drawCentredString(qr_x+qr_sz//2,qr_y-20,"Alias: ESTUDIO.CONTA.CARLON  |  $"+str(monto_int))

    cv.save();buffer.seek(0);return buffer

def generar_pdf(cliente_id, periodo, monto):
    buffer=BytesIO();cv=canvas.Canvas(buffer,pagesize=A4);w,h=A4
    conn=conectar();c=conn.cursor()
    c.execute("SELECT nombre,cuit FROM clientes WHERE id=%s",(cliente_id,))
    data=c.fetchone();conn.close()
    cli_nombre=data[0] if data else "—";cuit_cli=dec(data[1]) if data else ""

    # ── Encabezado verde ──────────────────────────────────────────────────────
    cv.setFillColorRGB(0.10,0.23,0.16);cv.rect(0,h-140,w,140,fill=1,stroke=0)

    # Logo más grande (BN se ve bien en blanco)
    logo_dibujado=False
    for lp in ["logo.png","static/logo.png"]:
        if os.path.exists(lp):
            try:
                cv.drawImage(ImageReader(lp),28,h-128,width=115,height=110,
                             preserveAspectRatio=True,mask="auto")
                logo_dibujado=True
            except: pass
            break
    # Si no hay logo, texto del estudio grande
    if not logo_dibujado:
        cv.setFillColorRGB(0.78,0.66,0.43)
        cv.setFont("Helvetica-Bold",26)
        cv.drawString(28,h-55,"CARLON")
        cv.setFont("Helvetica",9)
        cv.setFillColorRGB(1,1,1)
        cv.drawString(28,h-70,"ESTUDIO CONTABLE")

    # Titulo y datos al lado del logo
    x_txt = 158
    cv.setFillColorRGB(0.78,0.66,0.43);cv.setFont("Helvetica-Bold",22)
    cv.drawString(x_txt,h-52,"RECIBO DE PAGO")
    cv.setFillColorRGB(1,1,1);cv.setFont("Helvetica",8.5)
    cv.drawString(x_txt,h-68,"Estudio Contable Carlon — Servicios Contables e Impositivos")
    numero=datetime.now().strftime("%Y%m%d%H%M%S")
    cv.setFont("Helvetica-Bold",9);cv.drawRightString(w-36,h-52,f"N° {numero}")
    cv.setFont("Helvetica",8);cv.drawRightString(w-36,h-66,datetime.now().strftime("%d/%m/%Y %H:%M"))

    # ── Emisor ─────────────────────────────────────────────────────────────────
    cv.setFillColorRGB(0.15,0.15,0.15);cv.setFont("Helvetica-Bold",8);cv.drawString(36,h-160,"EMISOR")
    cv.setFont("Helvetica",8.5);cv.drawString(36,h-174,"Estudio Contable Carlon  ·  CUIT: 27-35045505-7")
    cv.drawString(36,h-186,"Absalón Rojas s/n  ·  Quimilí, Santiago del Estero  ·  CP 3740")
    cv.setStrokeColorRGB(0.87,0.87,0.87);cv.line(36,h-198,w-36,h-198)

    # ── Cliente ────────────────────────────────────────────────────────────────
    cv.setFillColorRGB(0.53,0.53,0.53);cv.setFont("Helvetica-Bold",8);cv.drawString(36,h-216,"CLIENTE")
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",14);cv.drawString(36,h-234,cli_nombre)
    cv.setFont("Helvetica",8.5);cv.setFillColorRGB(0.3,0.3,0.3)
    cv.drawString(36,h-250,f"CUIT: {cuit_cli or '—'}   ·   Periodo: {periodo}")

    # ── Monto ──────────────────────────────────────────────────────────────────
    cv.setFillColorRGB(0.97,0.96,0.93);cv.roundRect(36,h-320,w-72,55,8,fill=1,stroke=0)
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",10);cv.drawString(54,h-282,"TOTAL ABONADO")
    cv.setFont("Helvetica-Bold",24)
    try: mf=f"$ {float(monto):,.0f}".replace(",",".")
    except: mf=f"$ {monto}"
    cv.drawRightString(w-54,h-282,mf)
    cv.setFont("Helvetica",8);cv.setFillColorRGB(0.45,0.45,0.45)
    cv.drawString(36,h-338,"Recibí conforme el importe indicado en concepto de honorarios profesionales.")

    # ── Firmas ─────────────────────────────────────────────────────────────────
    cv.setStrokeColorRGB(0.72,0.72,0.72)
    cv.line(36,h-400,195,h-400);cv.line(w-195,h-400,w-36,h-400)
    cv.setFont("Helvetica",8);cv.setFillColorRGB(0.5,0.5,0.5)
    cv.drawString(36,h-413,"Firma");cv.drawString(w-195,h-413,"Aclaración")

    # ── Datos bancarios ────────────────────────────────────────────────────────
    cv.setFillColorRGB(0.97,0.96,0.93);cv.roundRect(36,36,int((w-72)*0.62),118,8,fill=1,stroke=0)
    cv.setFillColorRGB(0.10,0.23,0.16);cv.setFont("Helvetica-Bold",9);cv.drawString(52,144,"DATOS PARA TRANSFERENCIA")
    cv.setFont("Helvetica",8.5);cv.setFillColorRGB(0.2,0.2,0.2)
    for i,ln in enumerate([
        "Titular: Alexis Natasha Carlon",
        "CUIL: 27-35045505-7  ·  Banco Nacion Argentina",
        "Cuenta Ahorro $ 28324201345252",
        "CBU: 0110420630042013452529",
        "Alias: ESTUDIO.CONTA.CARLON"
    ]):
        cv.drawString(52,126-i*16,ln)

    # ── QR → link BNA+ para transferir ────────────────────────────────────────
    # BNA+ transfiere via alias: link universal que abre la app del banco
    try:
        monto_int=int(float(monto))
    except:
        monto_int=0
    # Link de transferencia BNA+ (alias directo)
    # QR con CBU directo - formato que lee BNA+, Mercado Pago, Personal Pay
    # Usar CBU en texto plano es el formato mas compatible
    qr_payload = "0110420630042013452529"
    # El QR del CBU lo leen todas las apps de transferencia argentinas
    # El monto se muestra abajo del QR para que el cliente lo ingrese manualmente
    qr=qrcode.QRCode(version=2,error_correction=qrcode.constants.ERROR_CORRECT_M,
                     box_size=5,border=2)
    qr.add_data(qr_payload)
    qr.make(fit=True)
    qr_img=qr.make_image(fill_color="black",back_color="white")
    qb=BytesIO();qr_img.save(qb,"PNG");qb.seek(0)
    qr_x=w-150;qr_y=30;qr_sz=116
    cv.setFillColorRGB(1,1,1);cv.roundRect(qr_x-5,qr_y-5,qr_sz+10,qr_sz+10,6,fill=1,stroke=0)
    cv.drawImage(ImageReader(qb),qr_x,qr_y,width=qr_sz,height=qr_sz)
    cv.setFont("Helvetica-Bold",7.5);cv.setFillColorRGB(0.10,0.23,0.16)
    cv.drawCentredString(qr_x+qr_sz//2,qr_y-10,"Escanear con cualquier home banking")
    cv.setFont("Helvetica",6.5);cv.setFillColorRGB(0.4,0.4,0.4)
    cv.drawCentredString(qr_x+qr_sz//2,qr_y-20,f"Alias: ESTUDIO.CONTA.CARLON  |  $"+str(monto_int))

    cv.save();buffer.seek(0);return buffer

@app.route("/recibo/<int:cliente_id>/<path:periodo>")
def ver_recibo(cliente_id,periodo):
    periodo=periodo.replace("-","/")
    conn=conectar();c=conn.cursor()
    c.execute("SELECT debe,haber FROM cuentas WHERE cliente_id=%s AND periodo=%s",(cliente_id,periodo))
    data=c.fetchone();conn.close()
    if not data: return "No hay datos para ese período",404
    monto=data[1] if data[1]>0 else data[0]
    pdf=generar_pdf(cliente_id,periodo,monto);dl=request.args.get("download")
    return send_file(pdf,mimetype="application/pdf",as_attachment=bool(dl),download_name=f"recibo_{periodo.replace('/','_')}.pdf")

# ══════════════════════════════════════════════════════════════════════════════
#  AGENDA VENCIMIENTOS
# ══════════════════════════════════════════════════════════════════════════════
MESES_ESP=["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

@app.route("/agenda")
@login_req
def agenda():
    mes=int(request.args.get("mes",datetime.now().month))
    anio=int(request.args.get("anio",datetime.now().year))
    conn=conectar();c=conn.cursor()
    c.execute("SELECT vencimiento_id,estado,nota,usuario,fecha_actualizacion FROM agenda_vencimientos WHERE mes=%s AND anio=%s",(mes,anio))
    rows={r[0]:r for r in c.fetchall()}
    c.execute("SELECT vencimiento_id,estado,nota,usuario,fecha_actualizacion FROM agenda_vencimientos WHERE mes=%s AND anio=%s ORDER BY fecha_actualizacion DESC LIMIT 8",(mes,anio))
    actividad=c.fetchall();conn.close()
    hoy=datetime.now();dia_hoy=hoy.day if hoy.month==mes and hoy.year==anio else 0
    stats={"total":len(VENCIMIENTOS_IMPOSITIVOS),"pendiente":0,"borrador":0,"presentado":0,"observado":0}
    alertas=[]
    filas=""
    for v in VENCIMIENTOS_IMPOSITIVOS:
        row=rows.get(v["id"])
        estado=row[1] if row else "pendiente"
        nota=row[2] if row else ""
        stats[estado]=stats.get(estado,0)+1
        dias_rest=v["dia"]-dia_hoy if dia_hoy>0 else 99
        if estado=="pendiente" and 0<dias_rest<=5:
            alertas.append(f"⚠️ <b>{v['nombre']}</b> vence en {dias_rest} día(s) (día {v['dia']})")
        elif estado=="pendiente" and dias_rest<=0 and dia_hoy>0:
            alertas.append(f"🔴 <b>{v['nombre']}</b> venció el día {v['dia']} y está pendiente")
        estados_opts="".join(f'<option value="{s}" {"selected" if s==estado else ""}>{l}</option>' for s,l in [("pendiente","⏳ Pendiente"),("borrador","📝 En borrador"),("presentado","✅ Presentado"),("observado","⚠ Observado")])
        tipo_badge=f'<span style="font-size:.68rem;padding:2px 7px;border-radius:8px;background:{"#f0f4ff" if v["tipo"]=="AFIP" else "#f0f9f4"};color:{"#1a3a8a" if v["tipo"]=="AFIP" else "#1a5c3a"};font-weight:600">{v["tipo"]}</span>'
        venc_badge=f'<span style="font-size:.72rem;color:var(--muted)">Día {v["dia"]}</span>'
        color_est={"pendiente":"var(--border)","borrador":"var(--warning)","presentado":"var(--success)","observado":"var(--danger)"}.get(estado,"var(--border)")
        filas+=f'''<div style="background:var(--card);border-radius:var(--r);padding:14px 18px;box-shadow:var(--shadow);margin-bottom:9px;border-left:4px solid {color_est}">
          <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px">
            <div><span style="font-weight:600;color:var(--primary)">{v["nombre"]}</span> {tipo_badge} {venc_badge}</div>
            <form method="post" action="/agenda/actualizar" style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">
              <input type="hidden" name="venc_id" value="{v["id"]}">
              <input type="hidden" name="mes" value="{mes}">
              <input type="hidden" name="anio" value="{anio}">
              <div class="fg" style="min-width:160px"><select name="estado" style="font-size:.82rem">{estados_opts}</select></div>
              <div class="fg" style="flex:1;min-width:180px"><input name="nota" value="{nota}" placeholder="Nota..." style="font-size:.82rem"></div>
              <button type="submit" class="btn btn-p btn-sm">Guardar</button>
            </form>
          </div>
        </div>'''
    act_html=""
    NOMBRES_V={v["id"]:v["nombre"] for v in VENCIMIENTOS_IMPOSITIVOS}
    for a in actividad:
        ESTADOS_LABELS2={"pendiente":"Pendiente","borrador":"En borrador","presentado":"Presentado","observado":"Observado"}
        act_html+=f'<div class="logrow"><div class="log-dot"></div><span class="log-time">{a[4]}</span><span class="log-user">{a[3]}</span><span class="log-msg"><b>{NOMBRES_V.get(a[0],a[0])}</b> → {ESTADOS_LABELS2.get(a[1],a[1])}{" · "+a[2] if a[2] else ""}</span></div>'
    alerta_html='<div class="warn-box" style="margin-bottom:16px">'+"<br>".join(alertas)+"</div>" if alertas else ""
    pct_pres=int(stats["presentado"]/stats["total"]*100) if stats["total"]>0 else 0
    mes_ant,anio_ant=(mes-1,anio) if mes>1 else (12,anio-1)
    mes_sig,anio_sig=(mes+1,anio) if mes<12 else (1,anio+1)
    body=f"""
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin-bottom:5px">
      <h1 class="page-title">Agenda de Vencimientos</h1>
      <div style="display:flex;align-items:center;gap:8px">
        <a href="/agenda?mes={mes_ant}&anio={anio_ant}" class="btn btn-o btn-sm">← Ant.</a>
        <span style="font-family:'DM Serif Display',serif;font-size:1.1rem;color:var(--primary);min-width:160px;text-align:center">{MESES_ESP[mes]} {anio}</span>
        <a href="/agenda?mes={mes_sig}&anio={anio_sig}" class="btn btn-o btn-sm">Sig. →</a>
      </div>
    </div>
    <p class="page-sub">Control de vencimientos impositivos · Santiago del Estero</p>
    {alerta_html}
    <div class="stats" style="margin-bottom:18px">
      <div class="scard"><div class="slabel">Total</div><div class="sval">{stats["total"]}</div></div>
      <div class="scard g"><div class="slabel">Presentados</div><div class="sval">{stats["presentado"]}</div></div>
      <div class="scard b"><div class="slabel">En borrador</div><div class="sval">{stats["borrador"]}</div></div>
      <div class="scard r"><div class="slabel">Pendientes</div><div class="sval">{stats["pendiente"]}</div></div>
    </div>
    <div class="fcard" style="margin-bottom:16px">
      <div style="display:flex;justify-content:space-between;margin-bottom:6px;font-size:.82rem">
        <span style="font-weight:600;color:var(--primary)">Progreso del mes</span>
        <span style="color:var(--success);font-weight:700">{pct_pres}% presentado</span>
      </div>
      <div class="progwrap"><div class="progbar" style="width:{pct_pres}%"></div></div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 320px;gap:16px;align-items:start" class="twocol">
      <div>{filas}</div>
      <div>
        <div class="fcard"><h3>🕓 Actividad reciente</h3>{act_html or "<p style='color:var(--muted);font-size:.84rem'>Sin actividad</p>"}</div>
        <div class="info-box"><b>Leyenda:</b><br>⏳ Pendiente · 📝 Borrador · ✅ Presentado · ⚠ Observado</div>
      </div>
    </div>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>"""
    return page("Agenda", body, "Agenda")

@app.route("/agenda/actualizar", methods=["POST"])
@login_req
def agenda_actualizar():
    venc_id=request.form.get("venc_id","").strip()
    mes=int(request.form.get("mes",datetime.now().month))
    anio=int(request.form.get("anio",datetime.now().year))
    estado=request.form.get("estado","pendiente")
    nota=request.form.get("nota","").strip()
    ids_validos={v["id"] for v in VENCIMIENTOS_IMPOSITIVOS}
    if venc_id not in ids_validos: return redirect(f"/agenda?mes={mes}&anio={anio}")
    conn=conectar();c=conn.cursor()
    c.execute("""INSERT INTO agenda_vencimientos(vencimiento_id,mes,anio,estado,nota,usuario,fecha_actualizacion)
                 VALUES(%s,%s,%s,%s,%s,%s,%s)
                 ON CONFLICT(vencimiento_id,mes,anio)
                 DO UPDATE SET estado=%s,nota=%s,usuario=%s,fecha_actualizacion=%s""",
              (venc_id,mes,anio,estado,nota,session.get("display",""),now_ar(),
               estado,nota,session.get("display",""),now_ar()))
    conn.commit();conn.close()
    nombre_v=next((v["nombre"] for v in VENCIMIENTOS_IMPOSITIVOS if v["id"]==venc_id),venc_id)
    registrar_auditoria("AGENDA",f"{nombre_v} → {estado}")
    return redirect(f"/agenda?mes={mes}&anio={anio}")

# ══════════════════════════════════════════════════════════════════════════════
#  IMPORTAR
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/importar", methods=["GET","POST"])
@admin_req
def importar():
    try: import pandas as pd
    except: return "pandas no instalado",500
    if request.method=="POST":
        archivo=request.files["archivo"];df=pd.read_excel(archivo)
        conn=conectar();c=conn.cursor();ok=0
        for _,row in df.iterrows():
            nombre=str(row.get("nombre y apellido","")).strip()
            if not nombre: continue
            c.execute("INSERT INTO clientes(nombre,cuit,telefono,abono) VALUES(%s,%s,%s,%s)",
                      (nombre,enc(str(row.get("cuit",""))),enc(str(row.get("telefono",""))),row.get("honorario",0)));ok+=1
        conn.commit();conn.close();registrar_auditoria("IMPORTACIÓN",f"{ok} clientes importados")
        return redirect("/clientes")
    body='<h1 class="page-title">Importar Clientes</h1><p class="page-sub">Excel con columnas: <b>nombre y apellido</b>, cuit, telefono, honorario</p><div class="fcard"><h3>📂 Archivo Excel</h3><form method="post" enctype="multipart/form-data"><div class="fg" style="margin-bottom:14px"><label>Archivo .xlsx</label><input type="file" name="archivo" accept=".xlsx,.xls"></div><button class="btn btn-p">Importar</button></form></div>'
    return page("Importar",body)



# ══════════════════════════════════════════════════════════════════════════════
#  PWA MANIFEST
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/manifest.json")
def manifest():
    from flask import Response
    data = json.dumps({
        "name": "Estudio Carlon",
        "short_name": "Carlon",
        "start_url": "/app",
        "display": "standalone",
        "background_color": "#F7F5F0",
        "theme_color": "#1A3A2A",
        "icons": [
            {"src": "https://placehold.co/192x192/1A3A2A/C8A96E?text=C", "sizes": "192x192", "type": "image/png"},
            {"src": "https://placehold.co/512x512/1A3A2A/C8A96E?text=C", "sizes": "512x512", "type": "image/png"}
        ]
    })
    return Response(data, mimetype="application/json")

# ══════════════════════════════════════════════════════════════════════════════
#  APP MOVIL SUPERVISORA
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/app", methods=["GET","POST"])
@login_req
def app_movil():
    conn=conectar();c=conn.cursor();flash=""
    hoy=datetime.now().strftime("%d/%m/%Y")
    tab=request.args.get("tab","caja")
    rol=session.get("rol","secretaria")
    usuario=session.get("display","")

    if request.method=="POST":
        accion=request.form.get("accion","")
        if accion=="registrar_arqueo":
            ef=float(request.form.get("ef",0) or 0)
            ch=float(request.form.get("ch",0) or 0)
            dol=float(request.form.get("dol",0) or 0)
            nota=request.form.get("nota","").strip()
            total=ef+ch+dol
            c.execute("INSERT INTO seguridad_eventos(tipo,detalle,ip,usuario,fecha) VALUES(%s,%s,%s,%s,%s)",
                      ("ARQUEO",
                       "Efectivo:"+fmt(ef)+" Cheque:"+fmt(ch)+" USD:"+fmt(dol)+" Total:"+fmt(total)+((" Nota:"+nota) if nota else ""),
                       "app-movil",usuario,now_ar()))
            conn.commit()
            registrar_auditoria("ARQUEO_CAJA","Ef:"+fmt(ef)+" Ch:"+fmt(ch)+" U$S:"+fmt(dol)+" Total:"+fmt(total)+((" - "+nota) if nota else ""))
            flash="ok:Arqueo guardado - Total fisico: "+fmt(total)
        elif accion=="gasto_rapido":
            cat=request.form.get("cat","Otros")
            desc=request.form.get("desc","").strip()
            monto_str=request.form.get("monto","0")
            monto=float(monto_str) if monto_str else 0
            if monto>0:
                c.execute("INSERT INTO gastos(fecha,categoria,descripcion,monto,usuario) VALUES(%s,%s,%s,%s,%s)",
                          (now_ar(),cat,desc,monto,usuario))
                conn.commit()
                registrar_auditoria("GASTO",cat+": "+fmt(monto)+" - "+desc)
                flash="ok:Gasto registrado: "+fmt(monto)
            else:
                flash="err:Ingresa un monto"

    # Datos
    c.execute("SELECT fecha,usuario,efectivo,cheque,dolares,transferencia_nat,transferencia_mai,otro,total_fisico,total_general,cerrado,hora_cierre FROM cierres_caja WHERE fecha=%s ORDER BY hora_cierre DESC",(hoy,))
    cierres=c.fetchall()
    c.execute("SELECT detalle,fecha FROM seguridad_eventos WHERE tipo='ARQUEO' AND fecha LIKE %s ORDER BY id DESC LIMIT 10",(hoy+"%",))
    arqueos=c.fetchall()
    c.execute("SELECT categoria,descripcion,monto,usuario FROM gastos WHERE fecha LIKE %s ORDER BY id DESC LIMIT 20",(hoy+"%",))
    gastos_hoy=c.fetchall()
    total_gastos_hoy=sum(g[2] for g in gastos_hoy)
    c.execute("SELECT COALESCE(SUM(monto),0) FROM pagos WHERE fecha LIKE %s",(hoy+"%",))
    total_cobrado=c.fetchone()[0]
    conn.close()

    # Flash
    flash_html=""
    if flash:
        partes=flash.split(":",1)
        ftipo=partes[0] if len(partes)>1 else "ok"
        fmsg=partes[1] if len(partes)>1 else flash
        fclass="app-fok" if ftipo=="ok" else "app-ferr"
        flash_html='<div class="app-flash '+fclass+'">'+fmsg+'</div>'

    # ── TAB CAJA
    if tab=="caja":
        cierres_html=""
        for ci in cierres:
            fci,uci,ef,ch,dol,nat,mai,otr,tf,tg,cerr,hora=ci
            est_cls="ok" if cerr else "pend"
            est_txt=("Cerrada "+hora) if cerr else "Abierta"
            cierres_html+=(
                '<div class="cierre-row '+est_cls+'">'
                '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">'
                '<div><div class="cierre-nombre">'+uci+'</div>'
                '<div class="cierre-fecha">'+est_txt+'</div></div>'
                '<div style="font-family:\'DM Serif Display\',serif;font-size:1.2rem;color:var(--primary)">'+fmt(tg or 0)+'</div>'
                '</div>'
                '<div class="medios-grid">'
                '<div class="medio-chip"><span class="mc-lbl">Efectivo</span><span class="mc-val">'+fmt(ef or 0)+'</span></div>'
                '<div class="medio-chip"><span class="mc-lbl">Cheque</span><span class="mc-val">'+fmt(ch or 0)+'</span></div>'
                '<div class="medio-chip"><span class="mc-lbl">U$S</span><span class="mc-val">'+fmt(dol or 0)+'</span></div>'
                '<div class="medio-chip"><span class="mc-lbl">Natasha</span><span class="mc-val">'+fmt(nat or 0)+'</span></div>'
                '<div class="medio-chip"><span class="mc-lbl">Maira</span><span class="mc-val">'+fmt(mai or 0)+'</span></div>'
                '<div class="medio-chip"><span class="mc-lbl">Total</span><span class="mc-val" style="color:var(--accent)">'+fmt(tg or 0)+'</span></div>'
                '</div></div>'
            )
        if not cierres_html:
            cierres_html='<p style="color:var(--muted);font-size:.85rem;text-align:center;padding:10px">Sin cierres hoy</p>'

        arqueos_html=""
        for aq in arqueos:
            det,fec=aq
            arqueos_html+=('<div style="background:var(--bg);border-radius:8px;padding:10px 12px;'
                          'margin-bottom:8px;font-size:.82rem">'
                          '<span style="color:var(--muted);font-size:.72rem">'+fec+'</span><br>'+det+'</div>')

        content_tab=(
            '<div class="app-total-box">'
            '<div class="atl">Total cobrado hoy</div>'
            '<div class="atv">'+fmt(total_cobrado)+'</div>'
            '</div>'
            '<div class="caja-card">'
            '<h2>Cierres de hoy <span style="font-size:.75rem;color:var(--muted)">'+str(len(cierres))+' sec.</span></h2>'
            +cierres_html+
            '</div>'
            '<div class="caja-card">'
            '<h2>Registrar Arqueo</h2>'
            '<p style="font-size:.8rem;color:var(--muted);margin-bottom:14px">Conta el dinero fisico del estudio.</p>'
            '<form method="post">'
            '<input type="hidden" name="accion" value="registrar_arqueo">'
            '<div class="app-input-row">'
            '<div class="app-field"><label>Efectivo $</label><input type="number" name="ef" placeholder="0" step="100"></div>'
            '<div class="app-field"><label>Cheque $</label><input type="number" name="ch" placeholder="0" step="100"></div>'
            '</div>'
            '<div class="app-input-row">'
            '<div class="app-field"><label>Dolares $</label><input type="number" name="dol" placeholder="0"></div>'
            '<div class="app-field"><label>Nota</label><input type="text" name="nota" placeholder="Opcional"></div>'
            '</div>'
            '<button type="submit" class="btn-app btn-app-g">Guardar Arqueo</button>'
            '</form></div>'
            +(('<div class="caja-card"><h2>Arqueos de hoy</h2>'+arqueos_html+'</div>') if arqueos else "")
        )

    # ── TAB GASTOS
    elif tab=="gastos":
        cats_disp=CATEGORIAS_GASTO if rol in ("admin","supervisor") else CATEGORIAS_SEC
        opts_cats="".join('<option value="'+c2+'">'+c2+'</option>' for c2 in cats_disp)
        gastos_list=""
        for g in gastos_hoy:
            gastos_list+=(
                '<div style="display:flex;justify-content:space-between;align-items:center;'
                'padding:10px 0;border-bottom:1px solid var(--border)">'
                '<div><span style="font-size:.82rem;font-weight:600">'+(g[1] or g[0])+'</span><br>'
                '<span class="bmedio" style="font-size:.65rem">'+g[0]+'</span> '
                '<span style="font-size:.7rem;color:var(--muted)">'+g[3]+'</span></div>'
                '<span style="font-family:\'DM Serif Display\',serif;color:var(--danger);font-size:1rem">'+fmt(g[2])+'</span>'
                '</div>'
            )
        if not gastos_list:
            gastos_list='<p style="color:var(--muted);font-size:.85rem;text-align:center;padding:10px">Sin gastos hoy</p>'
        content_tab=(
            '<div class="app-total-box" style="background:var(--danger)">'
            '<div class="atl">Gastos de hoy</div>'
            '<div class="atv">'+fmt(total_gastos_hoy)+'</div>'
            '</div>'
            '<div class="caja-card"><h2>Registrar Gasto</h2>'
            '<form method="post">'
            '<input type="hidden" name="accion" value="gasto_rapido">'
            '<div class="app-field" style="margin-bottom:12px"><label>Categoria</label><select name="cat">'+opts_cats+'</select></div>'
            '<div class="app-field" style="margin-bottom:12px"><label>Descripcion</label><input type="text" name="desc" placeholder="Detalle..."></div>'
            '<div class="app-field" style="margin-bottom:12px"><label>Monto $</label><input type="number" name="monto" placeholder="0" step="100" required></div>'
            '<button type="submit" class="btn-app btn-app-r">Registrar Gasto</button>'
            '</form></div>'
            '<div class="caja-card"><h2>Gastos de hoy</h2>'
            +gastos_list+
            (('<div style="text-align:right;font-weight:700;padding-top:10px;color:var(--danger)">Total: '+fmt(total_gastos_hoy)+'</div>') if gastos_hoy else "")
            +'</div>'
        )

    # ── TAB RESUMEN
    else:
        tab="resumen"
        total_ef=sum((ci[2] or 0) for ci in cierres)
        total_ch=sum((ci[3] or 0) for ci in cierres)
        total_dol=sum((ci[4] or 0) for ci in cierres)
        total_nat=sum((ci[5] or 0) for ci in cierres)
        total_mai=sum((ci[6] or 0) for ci in cierres)
        total_gen=sum((ci[9] or 0) for ci in cierres)
        content_tab=(
            '<div class="caja-card"><h2>Resumen del dia</h2>'
            '<div class="medios-grid" style="grid-template-columns:1fr 1fr;gap:10px">'
            '<div class="medio-chip"><span class="mc-lbl">Total Cobrado</span><span class="mc-val" style="font-size:1.1rem;color:var(--success)">'+fmt(total_cobrado)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">Total Gastos</span><span class="mc-val" style="font-size:1.1rem;color:var(--danger)">'+fmt(total_gastos_hoy)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">Efectivo</span><span class="mc-val">'+fmt(total_ef)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">Cheque</span><span class="mc-val">'+fmt(total_ch)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">U$S</span><span class="mc-val">'+fmt(total_dol)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">Natasha</span><span class="mc-val">'+fmt(total_nat)+'</span></div>'
            '<div class="medio-chip"><span class="mc-lbl">Maira</span><span class="mc-val">'+fmt(total_mai)+'</span></div>'
            '<div class="medio-chip" style="background:var(--primary);border-color:var(--primary)"><span class="mc-lbl" style="color:rgba(255,255,255,.7)">TOTAL</span><span class="mc-val" style="color:var(--accent);font-size:1.1rem">'+fmt(total_gen)+'</span></div>'
            '</div>'
            '<div style="margin-top:14px;font-size:.78rem;color:var(--muted);text-align:center">'+str(len(cierres))+' cierres - '+str(len(gastos_hoy))+' gastos - '+hoy+'</div>'
            '</div>'
            '<div class="caja-card"><h2>Accesos rapidos</h2>'
            '<div style="display:flex;flex-direction:column;gap:10px">'
            '<a href="/caja" class="btn-app btn-app-b" style="text-align:center;display:block;text-decoration:none;padding:13px">Ver Caja Completa</a>'
            '<a href="/clientes" class="btn-app btn-app-o" style="text-align:center;display:block;text-decoration:none;padding:13px">Ver Clientes</a>'
            '</div></div>'
        )

    act_caja="act" if tab=="caja" else ""
    act_gas="act" if tab=="gastos" else ""
    act_res="act" if tab=="resumen" else ""
    tab_bar=(
        '<div class="tab-bar">'
        '<a href="/app?tab=caja" class="tab-item '+act_caja+'"><span>&#x1F3E6;</span>Caja</a>'
        '<a href="/app?tab=gastos" class="tab-item '+act_gas+'"><span>&#x1F4DD;</span>Gastos</a>'
        '<a href="/app?tab=resumen" class="tab-item '+act_res+'"><span>&#x1F4CA;</span>Resumen</a>'
        '<a href="/logout" class="tab-item"><span>&#x1F6AA;</span>Salir</a>'
        '</div>'
    )

    return ('<!DOCTYPE html><html lang="es"><head>'
            '<meta charset="UTF-8">'
            '<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">'
            '<meta name="apple-mobile-web-app-capable" content="yes">'
            '<meta name="apple-mobile-web-app-title" content="Estudio Carlon">'
            '<meta name="theme-color" content="#1A3A2A">'
            '<link rel="manifest" href="/manifest.json">'
            '<title>Estudio Carlon</title>'
            '<style>'+CSS+'body{padding-bottom:75px}</style>'
            '</head><body>'
            '<div class="app-wrap">'
            '<div class="app-header">'
            '<div class="app-title">Estudio Carlon</div>'
            '<div class="app-date">Hola '+usuario+' - '+hoy+'</div>'
            '</div>'
            +flash_html+content_tab+
            '</div>'+tab_bar+'</body></html>')



# ══════════════════════════════════════════════════════════════════════════════
#  TAREAS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/tareas", methods=["GET","POST"])
@login_req
def tareas():
    conn=conectar();c=conn.cursor()
    flash=""; rol=session.get("rol"); usuario=session.get("display","")

    if request.method=="POST":
        accion=request.form.get("accion","")
        if accion=="nueva":
            titulo=request.form.get("titulo","").strip()
            desc=request.form.get("descripcion","").strip()
            prio=request.form.get("prioridad","normal")
            fvenc=request.form.get("fecha_vencimiento","").strip() or None
            asig=request.form.get("asignado_a",usuario).strip() or usuario
            if titulo:
                c.execute("INSERT INTO tareas(titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento) VALUES(%s,%s,%s,%s,'pendiente',%s,%s,%s,%s)",
                          (titulo,desc,usuario,asig,prio,now_ar(),now_ar(),fvenc))
                conn.commit(); flash='<div class="flash fok">✅ Tarea creada</div>'
        elif accion=="actualizar":
            tid=request.form.get("tid"); nuevo_est=request.form.get("estado")
            nuevo_titulo=request.form.get("titulo","").strip()
            nueva_prio=request.form.get("prioridad","normal")
            nueva_desc=request.form.get("descripcion","").strip()
            nueva_fvenc=request.form.get("fecha_vencimiento","").strip() or None
            if tid:
                c.execute("UPDATE tareas SET estado=%s,titulo=%s,prioridad=%s,descripcion=%s,fecha_vencimiento=%s,fecha_actualizacion=%s WHERE id=%s",
                          (nuevo_est,nuevo_titulo,nueva_prio,nueva_desc,nueva_fvenc,now_ar(),tid))
                conn.commit(); flash='<div class="flash fok">✅ Tarea actualizada</div>'
        elif accion=="borrar" and rol=="admin":
            tid=request.form.get("tid")
            if tid: c.execute("DELETE FROM tareas WHERE id=%s",(tid,)); conn.commit()
            flash='<div class="flash fok">Tarea eliminada</div>'

    # Filtros
    filtro_est=request.args.get("est","activas")
    editar_id=request.args.get("editar","")

    # Para admin: ve todas; para secretaria: las propias o asignadas
    if rol=="admin":
        if filtro_est=="completadas":
            c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas WHERE estado='completada' ORDER BY fecha_actualizacion DESC LIMIT 50")
        elif filtro_est=="todas":
            c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas ORDER BY CASE prioridad WHEN 'urgente' THEN 1 WHEN 'alta' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END,fecha_actualizacion DESC")
        else:
            c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas WHERE estado!='completada' ORDER BY CASE prioridad WHEN 'urgente' THEN 1 WHEN 'alta' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END")
    else:
        if filtro_est=="completadas":
            c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas WHERE (asignado_a=%s OR usuario=%s) AND estado='completada' ORDER BY fecha_actualizacion DESC",(usuario,usuario))
        else:
            c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas WHERE (asignado_a=%s OR usuario=%s) AND estado!='completada' ORDER BY CASE prioridad WHEN 'urgente' THEN 1 WHEN 'alta' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END",(usuario,usuario))

    lista=c.fetchall()

    # Tarea a editar
    tarea_edit=None
    if editar_id:
        c.execute("SELECT id,titulo,descripcion,usuario,asignado_a,estado,prioridad,fecha_creacion,fecha_actualizacion,fecha_vencimiento FROM tareas WHERE id=%s",(editar_id,))
        tarea_edit=c.fetchone()

    # Usuarios para asignar
    c.execute("SELECT nombre_display FROM usuarios WHERE activo IS NOT FALSE ORDER BY nombre_display")
    usuarios_lista=[r[0] for r in c.fetchall()]
    conn.close()

    PRIO_COLOR={"urgente":"var(--danger)","alta":"var(--warning)","normal":"var(--info)","baja":"var(--muted)"}
    PRIO_LABEL={"urgente":"🔴 Urgente","alta":"🟠 Alta","normal":"🔵 Normal","baja":"⚪ Baja"}
    EST_BADGE={"pendiente":'<span class="sec-badge">⏳ Pendiente</span>',"borrador":'<span class="sec-badge warn">📝 Borrador</span>',"en_progreso":'<span class="sec-badge ok">▶ En progreso</span>',"completada":'<span class="sec-badge ok">✅ Completada</span>'}

    rows=""
    for t in lista:
        tid,titulo,desc,usu,asig,est,prio,fcre,fact,fvenc=t
        col=PRIO_COLOR.get(prio,"var(--info)")
        est_b=EST_BADGE.get(est,est)
        prio_b=f'<span style="font-size:.68rem;color:{col};font-weight:700">{PRIO_LABEL.get(prio,prio)}</span>'
        venc_txt=(f'<br><span style="font-size:.72rem;color:{"var(--danger)" if fvenc and fvenc<=now_ar()[:10] else "var(--muted)"}">Vence: {fvenc}</span>' if fvenc else "")
        asig_txt=(f'<span style="font-size:.72rem;color:var(--muted)"> → {asig}</span>' if asig and asig!=usu else "")
        btn_edit=f'<a href="/tareas?editar={tid}" class="btn btn-xs btn-o">✏️ Editar</a>'
        btn_comp=('<form method="post" style="display:inline"><input type=hidden name=accion value=actualizar>'
                  +f'<input type=hidden name=tid value={tid}><input type=hidden name=titulo value="{titulo}">'
                  +'<input type=hidden name=estado value=completada><input type=hidden name=prioridad value='+prio+'>'
                  +'<button class="btn btn-xs btn-g" title="Marcar como completada">✓ Listo</button></form>') if est!="completada" else ""
        btn_del=(f'<form method="post" style="display:inline"><input type=hidden name=accion value=borrar><input type=hidden name=tid value={tid}><button class="btn btn-xs btn-r" onclick="return confirm(\'Eliminar?\')">🗑</button></form>') if rol=="admin" else ""
        rows+=(f'<div class="arow" style="border-left:3px solid {col}">'               f'<div style="flex:1"><span style="font-weight:600">{titulo}</span>{asig_txt}'
               f'<br>{est_b} {prio_b}{venc_txt}'               +(f'<br><span style="font-size:.76rem;color:var(--muted)">{desc[:80]}{"..." if len(desc or "")>80 else ""}</span>' if desc else "")               +f'<br><span style="font-size:.7rem;color:var(--muted)">Creada por {usu} · {fcre}</span></div>'               +f'<div style="display:flex;gap:5px;flex-wrap:wrap">{btn_edit}{btn_comp}{btn_del}</div></div>')

    # Form nueva/editar tarea
    if tarea_edit:
        te=tarea_edit
        opts_est="".join(f'<option value="{s}" {"selected" if s==te[5] else ""}>{l}</option>' for s,l in [("pendiente","⏳ Pendiente"),("borrador","📝 Borrador"),("en_progreso","▶ En progreso"),("completada","✅ Completada")])
        opts_prio="".join(f'<option value="{s}" {"selected" if s==te[6] else ""}>{l}</option>' for s,l in [("urgente","🔴 Urgente"),("alta","🟠 Alta"),("normal","🔵 Normal"),("baja","⚪ Baja")])
        opts_asig="".join(f'<option value="{u}" {"selected" if u==te[4] else ""}>{u}</option>' for u in usuarios_lista)
        form_html=(f'<div class="fcard" style="margin-bottom:16px;border:2px solid var(--accent)"><h3>✏️ Editar Tarea</h3>'
                   f'<form method="post"><input type="hidden" name="accion" value="actualizar"><input type="hidden" name="tid" value="{te[0]}">'
                   f'<div class="fgrid"><div class="fg"><label>Título</label><input name="titulo" value="{te[1]}" required></div>'
                   f'<div class="fg"><label>Estado</label><select name="estado">{opts_est}</select></div>'
                   f'<div class="fg"><label>Prioridad</label><select name="prioridad">{opts_prio}</select></div>'
                   f'<div class="fg"><label>Asignada a</label><select name="asignado_a">{opts_asig}</select></div>'
                   f'<div class="fg"><label>Vence</label><input name="fecha_vencimiento" type="date" value="{te[9] or ""}"></div></div>'
                   f'<div class="fg" style="margin-bottom:12px"><label>Descripción</label><textarea name="descripcion" rows="2">{te[2] or ""}</textarea></div>'
                   f'<div class="mact"><a href="/tareas" class="btn btn-o">Cancelar</a><button type="submit" class="btn btn-a">Guardar cambios</button></div>'
                   f'</form></div>')
    else:
        opts_prio_n="".join(f'<option value="{s}">{l}</option>' for s,l in [("urgente","🔴 Urgente"),("alta","🟠 Alta"),("normal","🔵 Normal"),("baja","⚪ Baja")])
        opts_prio_n=opts_prio_n.replace('<option value="normal">','<option value="normal" selected>')
        opts_asig_n="".join(f'<option value="{u}" {"selected" if u==usuario else ""}>{u}</option>' for u in usuarios_lista)
        show_form=request.args.get("nueva","")
        form_html=('<div class="fcard" style="margin-bottom:16px"><h3>➕ Nueva Tarea</h3>'
                   '<form method="post"><input type="hidden" name="accion" value="nueva">'
                   '<div class="fgrid">'
                   '<div class="fg"><label>Título *</label><input name="titulo" placeholder="Ej: Llamar a cliente X" required></div>'
                   f'<div class="fg"><label>Prioridad</label><select name="prioridad">{opts_prio_n}</select></div>'
                   f'<div class="fg"><label>Asignada a</label><select name="asignado_a">{opts_asig_n}</select></div>'
                   '<div class="fg"><label>Fecha vence</label><input name="fecha_vencimiento" type="date"></div></div>'
                   '<div class="fg" style="margin-bottom:12px"><label>Descripción</label><textarea name="descripcion" rows="2" placeholder="Detalle opcional..."></textarea></div>'
                   '<button class="btn btn-p">Crear tarea</button></form></div>') if show_form else (
                   '<div style="margin-bottom:14px"><a href="/tareas?nueva=1" class="btn btn-p btn-sm">+ Nueva tarea</a></div>')

    # Tabs filtro
    def ftab(est,lbl):
        cls="tab on" if filtro_est==est else "tab"
        return f'<a href="/tareas?est={est}" class="{cls}" style="text-decoration:none">{lbl}</a>'
    tabs=('<div class="tabs" style="margin-bottom:16px">'
          +ftab("activas","⏳ Activas")
          +ftab("completadas","✅ Completadas")
          +(ftab("todas","📋 Todas") if rol=="admin" else "")
          +'</div>')

    n_act=len([t for t in lista if t[5]!="completada"])
    body=f"""
    <h1 class="page-title">Agenda de Tareas</h1>
    <p class="page-sub">{len(lista)} tareas · {"Todas las secretarias" if rol=="admin" else "Mis tareas"}</p>
    {flash}
    {tabs}
    {form_html}
    {rows or '<div class="info-box">Sin tareas en esta vista ✨</div>'}
    """
    return page("Tareas", body, "Clientes")

@app.route("/tareas/completar/<int:tid>", methods=["POST"])
@login_req
def completar_tarea(tid):
    conn=conectar();c=conn.cursor()
    c.execute("UPDATE tareas SET estado='completada',fecha_actualizacion=%s WHERE id=%s",(now_ar(),tid))
    conn.commit();conn.close()
    return redirect(request.referrer or "/tareas")


# ══════════════════════════════════════════════════════════════════════════════
#  EMPLEADOS
# ══════════════════════════════════════════════════════════════════════════════
CONVENIOS = [
    "Empleados de Comercio (130/75)",
    "Transporte de Cargas / Camioneros (40/89)",
    "Trabajadores Agropecuarios / UATRE (1/75)",
    "Construccion (UOCRA)",
    "Gastronómicos",
    "Metalúrgicos (UOM)",
    "Municipales",
    "Docentes",
    "Personal de Casas Particulares",
    "Sin convenio / Fuera de convenio",
    "Otro",
]

@app.route("/empleados/<int:cliente_id>", methods=["GET","POST"])
@login_req
def empleados(cliente_id):
    conn=conectar();c=conn.cursor()
    flash=""
    c.execute("SELECT nombre FROM clientes WHERE id=%s",(cliente_id,))
    cli=c.fetchone()
    if not cli: conn.close(); return redirect("/clientes")
    nombre_cli=cli[0]

    if request.method=="POST":
        accion=request.form.get("accion","")
        if accion=="alta":
            nombre_emp=request.form.get("nombre","").strip()
            cuil=request.form.get("cuil","").strip()
            categoria=request.form.get("categoria","").strip()
            convenio=request.form.get("convenio","").strip()
            fecha_ing=request.form.get("fecha_ingreso","").strip()
            obs=request.form.get("observaciones","").strip()
            if nombre_emp:
                c.execute("""INSERT INTO empleados(cliente_id,nombre,cuil,categoria,convenio,
                             fecha_ingreso,activo,observaciones,fecha_alta)
                             VALUES(%s,%s,%s,%s,%s,%s,TRUE,%s,%s)""",
                          (cliente_id,nombre_emp,enc(cuil),categoria,convenio,
                           fecha_ing,obs,now_ar()))
                conn.commit()
                registrar_auditoria("ALTA EMPLEADO",f"{nombre_emp} | {convenio}",cliente_id,nombre_cli)
                flash='<div class="flash fok">✅ Empleado registrado</div>'
        elif accion=="baja":
            emp_id=request.form.get("emp_id")
            if emp_id:
                c.execute("UPDATE empleados SET activo=FALSE WHERE id=%s AND cliente_id=%s",(emp_id,cliente_id))
                conn.commit()
                flash='<div class="flash fok">Empleado dado de baja</div>'
        elif accion=="reactivar":
            emp_id=request.form.get("emp_id")
            if emp_id:
                c.execute("UPDATE empleados SET activo=TRUE WHERE id=%s AND cliente_id=%s",(emp_id,cliente_id))
                conn.commit()
        elif accion=="editar":
            emp_id=request.form.get("emp_id")
            nombre_emp=request.form.get("nombre","").strip()
            cuil=request.form.get("cuil","").strip()
            categoria=request.form.get("categoria","").strip()
            convenio=request.form.get("convenio","").strip()
            fecha_ing=request.form.get("fecha_ingreso","").strip()
            obs=request.form.get("observaciones","").strip()
            if emp_id:
                c.execute("""UPDATE empleados SET nombre=%s,cuil=%s,categoria=%s,convenio=%s,
                             fecha_ingreso=%s,observaciones=%s WHERE id=%s AND cliente_id=%s""",
                          (nombre_emp,enc(cuil),categoria,convenio,fecha_ing,obs,emp_id,cliente_id))
                conn.commit()
                flash='<div class="flash fok">✅ Empleado actualizado</div>'

    # Get employees
    c.execute("""SELECT id,nombre,cuil,categoria,convenio,fecha_ingreso,activo,observaciones
                 FROM empleados WHERE cliente_id=%s ORDER BY activo DESC,nombre""",
              (cliente_id,))
    emps=c.fetchall()
    conn.close()

    tab=request.args.get("tab","activos")
    activos=[e for e in emps if e[6]]
    bajas=[e for e in emps if not e[6]]
    lista=activos if tab=="activos" else bajas

    conv_opts="".join(f'<option value="{cv}">{cv}</option>' for cv in CONVENIOS)

    # Tabla empleados
    filas=""
    for e in lista:
        eid,enombre,ecuil_enc,ecat,econv,efing,eact,eobs=e
        ecuil=dec(ecuil_enc) if ecuil_enc else ""
        cuil_limpio=(ecuil or "").replace("-","").replace(" ","")
        # Link a ARCA constancia
        arca_btn=(f'<a href="https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do?nroCuit={cuil_limpio}" target="_blank" class="btn btn-xs btn-arca" title="Constancia ARCA">ARCA</a>' if cuil_limpio else "")
        # Badge convenio
        conv_color={"Empleados de Comercio (130/75)":"#1A5276","Transporte de Cargas / Camioneros (40/89)":"#6E2F1A","Trabajadores Agropecuarios / UATRE (1/75)":"#1A6B2F"}.get(econv or "","#555")
        conv_badge=(f'<span style="background:{conv_color};color:#fff;font-size:.65rem;padding:2px 6px;border-radius:4px;font-weight:600">{(econv or "Sin convenio").split(" (")[0][:20]}</span>' if econv else "")
        # Action buttons
        if eact:
            btns=(f'<form method="post" style="display:inline"><input type=hidden name=accion value=baja><input type=hidden name=emp_id value={eid}>'                  f'<button class="btn btn-xs btn-o" onclick="return confirm(\'Dar de baja a {enombre}?\')">Dar de baja</button></form>')
        else:
            btns=(f'<form method="post" style="display:inline"><input type=hidden name=accion value=reactivar><input type=hidden name=emp_id value={eid}>'                  f'<button class="btn btn-xs btn-g">Reactivar</button></form>')
        edit_btn=(f'<button class="btn btn-xs btn-p" onclick="editEmp({eid},\'{enombre}\',\'{ecuil}\',\'{ecat or ""}\',\'{econv or ""}\',\'{efing or ""}\',\'{(eobs or "").replace(chr(39),"")}\')" >✏️</button>')
        filas+=(f'<tr><td style="font-weight:600">{enombre}</td>'                f'<td>{ecuil or "—"}</td>'                f'<td>{ecat or "—"}</td>'                f'<td>{conv_badge or (econv or "—")}</td>'                f'<td>{efing or "—"}</td>'                f'<td style="text-align:right">{arca_btn} {edit_btn} {btns}</td></tr>')

    empty=f'<tr><td colspan="6" style="text-align:center;color:var(--muted);padding:20px">Sin empleados {"activos" if tab=="activos" else "dados de baja"}</td></tr>'

    body=f"""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;flex-wrap:wrap">
      <a href="/cuenta/{cliente_id}" class="btn btn-o btn-sm">&larr; Volver a cuenta</a>
      <h1 class="page-title" style="margin:0">Empleados — {nombre_cli}</h1>
    </div>
    {flash}

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:18px" class="twocol">
      <!-- Stats -->
      <div class="fcard" style="margin-bottom:0">
        <div class="stats" style="margin-bottom:0">
          <div class="scard g" style="margin-bottom:0"><div class="slabel">Activos</div><div class="sval">{len(activos)}</div></div>
          <div class="scard o" style="margin-bottom:0"><div class="slabel">Bajas</div><div class="sval">{len(bajas)}</div></div>
          <div class="scard b" style="margin-bottom:0"><div class="slabel">Total</div><div class="sval">{len(emps)}</div></div>
        </div>
      </div>
      <!-- Links sindicatos -->
      <div class="fcard" style="margin-bottom:0">
        <div style="font-size:.75rem;font-weight:700;color:var(--muted);text-transform:uppercase;margin-bottom:8px">Links sindicatos y escalas</div>
        <div style="display:flex;flex-wrap:wrap;gap:6px">
          <a href="https://jorgevega.com.ar/laboral/71-empleados-comercio-escala-salarial-2016-2017.html" target="_blank" class="btn btn-xs btn-o">Escala Comercio</a>
          <a href="https://www.online.faecys.org.ar/Inicio.aspx" target="_blank" class="btn btn-xs btn-o">FAECYS</a>
          <a href="https://www.sasweb.com.ar/usuarios/login" target="_blank" class="btn btn-xs btn-o">Sind. Comercio SGO</a>
          <a href="https://www.camioneros-ba.org.ar/index.php/gremiales/salarios/escalas-salariales" target="_blank" class="btn btn-xs btn-o">Escala Camioneros</a>
          <a href="https://federacion.impresiondeboletas.com.ar/login.aspx" target="_blank" class="btn btn-xs btn-o">Sind. Camioneros Nac.</a>
          <a href="http://camioneros.sirwiq.com/Login/Acceso" target="_blank" class="btn btn-xs btn-o">Camioneros SGO</a>
          <a href="https://www.ignacioonline.com.ar/paritaria-agrarios-escalas-salariales-marzo-abril-y-junio-2026-uatre/" target="_blank" class="btn btn-xs btn-o">Escala Agro 2026</a>
          <a href="https://portal.renatre.org.ar/" target="_blank" class="btn btn-xs btn-o">RENATRE</a>
          <a href="https://apps.uatre.org.ar/usupe/signin.aspx" target="_blank" class="btn btn-xs btn-o">UATRE SGO</a>
        </div>
      </div>
    </div>

    <!-- Tabs activos/bajas -->
    <div class="tabs" style="margin-bottom:14px">
      <a href="/empleados/{cliente_id}?tab=activos" class="tab {"on" if tab=="activos" else ""}" style="text-decoration:none">👷 Activos ({len(activos)})</a>
      <a href="/empleados/{cliente_id}?tab=bajas" class="tab {"on" if tab=="bajas" else ""}" style="text-decoration:none">📋 Bajas ({len(bajas)})</a>
    </div>

    <!-- Tabla -->
    <div class="fcard" style="margin-bottom:16px;overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:.84rem">
        <thead>
          <tr style="background:var(--primary);color:#fff">
            <th style="padding:8px 12px;text-align:left">Nombre</th>
            <th style="padding:8px 12px;text-align:left">CUIL</th>
            <th style="padding:8px 12px;text-align:left">Categoría</th>
            <th style="padding:8px 12px;text-align:left">Convenio</th>
            <th style="padding:8px 12px;text-align:left">Ingreso</th>
            <th style="padding:8px 12px;text-align:right">Acciones</th>
          </tr>
        </thead>
        <tbody>
          {filas or empty}
        </tbody>
      </table>
    </div>

    <!-- Formulario alta -->
    <div class="fcard" id="form-alta">
      <h3>➕ Agregar empleado</h3>
      <form method="post">
        <input type="hidden" name="accion" value="alta">
        <div class="fgrid">
          <div class="fg"><label>Nombre completo *</label><input name="nombre" required placeholder="Apellido, Nombre"></div>
          <div class="fg"><label>CUIL</label><input name="cuil" placeholder="20-12345678-9"></div>
          <div class="fg"><label>Categoría</label><input name="categoria" placeholder="Ej: Cajero A, Chofer, Peón"></div>
          <div class="fg"><label>Convenio</label><select name="convenio"><option value="">Sin especificar</option>{conv_opts}</select></div>
          <div class="fg"><label>Fecha de ingreso</label><input name="fecha_ingreso" type="date"></div>
          <div class="fg"><label>Observaciones</label><input name="observaciones" placeholder="Opcional"></div>
        </div>
        <button class="btn btn-g">Registrar empleado</button>
      </form>
    </div>

    <!-- Modal editar -->
    <div id="modal-emp" class="mo" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.5);z-index:999;align-items:center;justify-content:center">
      <div style="background:var(--card);border-radius:var(--r);padding:24px;width:90%;max-width:500px;max-height:90vh;overflow-y:auto">
        <h3>✏️ Editar empleado</h3>
        <form method="post">
          <input type="hidden" name="accion" value="editar">
          <input type="hidden" name="emp_id" id="edit-id">
          <div class="fgrid">
            <div class="fg"><label>Nombre *</label><input name="nombre" id="edit-nombre" required></div>
            <div class="fg"><label>CUIL</label><input name="cuil" id="edit-cuil"></div>
            <div class="fg"><label>Categoría</label><input name="categoria" id="edit-cat"></div>
            <div class="fg"><label>Convenio</label><select name="convenio" id="edit-conv"><option value="">Sin especificar</option>{conv_opts}</select></div>
            <div class="fg"><label>Fecha ingreso</label><input name="fecha_ingreso" id="edit-ing" type="date"></div>
            <div class="fg"><label>Observaciones</label><input name="observaciones" id="edit-obs"></div>
          </div>
          <div style="display:flex;gap:8px;margin-top:12px">
            <button type="button" onclick="document.getElementById('modal-emp').style.display='none'" class="btn btn-o">Cancelar</button>
            <button type="submit" class="btn btn-a">Guardar</button>
          </div>
        </form>
      </div>
    </div>

    <script>
    function editEmp(id,nom,cuil,cat,conv,ing,obs){{
      document.getElementById('edit-id').value=id;
      document.getElementById('edit-nombre').value=nom;
      document.getElementById('edit-cuil').value=cuil;
      document.getElementById('edit-cat').value=cat;
      var sel=document.getElementById('edit-conv');
      for(var i=0;i<sel.options.length;i++){{
        if(sel.options[i].value===conv){{sel.selectedIndex=i;break;}}
      }}
      document.getElementById('edit-ing').value=ing;
      document.getElementById('edit-obs').value=obs;
      document.getElementById('modal-emp').style.display='flex';
    }}
    document.getElementById('modal-emp').addEventListener('click',function(e){{
      if(e.target===this) this.style.display='none';
    }});
    </script>
    <style>@media(max-width:700px){{.twocol{{grid-template-columns:1fr!important}}}}</style>
    """
    return page(f"Empleados — {nombre_cli}", body, "Clientes")


# ══════════════════════════════════════════════════════════════════════════════
#  SUELDOS Y 931
# ══════════════════════════════════════════════════════════════════════════════
def vto_931(cuit_str, mes, anio):
    """Calcula vencimiento F.931 según terminación de CUIT"""
    try:
        cuit_limpio = (cuit_str or "").replace("-","").replace(" ","")
        terminacion = int(cuit_limpio[-1]) if cuit_limpio else 0
        dias = {0:13,1:13,2:14,3:14,4:15,5:15,6:16,7:16,8:17,9:17}
        dia_vto = dias.get(terminacion, 15)
        # Vence el mes siguiente
        mes_vto = mes + 1 if mes < 12 else 1
        anio_vto = anio if mes < 12 else anio + 1
        return f"{dia_vto:02d}/{mes_vto:02d}/{anio_vto}"
    except:
        return "—"

@app.route("/sueldos", methods=["GET","POST"])
@login_req
def sueldos():
    conn = conectar(); c = conn.cursor()
    flash = ""
    hoy = datetime.now()
    mes = int(request.args.get("mes", hoy.month))
    anio = int(request.args.get("anio", hoy.year))

    if request.method == "POST":
        accion = request.form.get("accion","")
        cliente_id = request.form.get("cliente_id")
        est_rec = request.form.get("estado_recibo","pendiente")
        est_931 = request.form.get("estado_931","pendiente")
        est_vep = request.form.get("estado_vep","pendiente")
        obs = request.form.get("observaciones","").strip()
        if accion == "guardar" and cliente_id:
            c.execute("""INSERT INTO sueldos_estado(cliente_id,mes,anio,estado_recibo,estado_931,
                         estado_vep,observaciones,fecha_actualizacion,usuario)
                         VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
                         ON CONFLICT(cliente_id,mes,anio) DO UPDATE SET
                         estado_recibo=EXCLUDED.estado_recibo,
                         estado_931=EXCLUDED.estado_931,
                         estado_vep=EXCLUDED.estado_vep,
                         observaciones=EXCLUDED.observaciones,
                         fecha_actualizacion=EXCLUDED.fecha_actualizacion,
                         usuario=EXCLUDED.usuario""",
                      (cliente_id,mes,anio,est_rec,est_931,est_vep,obs,now_ar(),
                       session.get("display","")))
            conn.commit()
            flash = '<div class="flash fok">✅ Estado actualizado</div>'

    # Get all clients WITH employees
    c.execute("""SELECT DISTINCT cl.id, cl.nombre, cl.cuit,
                        COUNT(e.id) as n_emp
                 FROM clientes cl
                 JOIN empleados e ON e.cliente_id=cl.id AND e.activo=TRUE
                 WHERE cl.activo IS NOT FALSE
                 GROUP BY cl.id, cl.nombre, cl.cuit
                 ORDER BY cl.nombre""")
    clientes_con_emp = c.fetchall()

    # Get estados for this month
    c.execute("""SELECT cliente_id,estado_recibo,estado_931,estado_vep,observaciones
                 FROM sueldos_estado WHERE mes=%s AND anio=%s""", (mes, anio))
    estados = {r[0]: r[1:] for r in c.fetchall()}
    conn.close()

    # Build navigation months
    meses_nav = ""
    for m in range(1,13):
        sel = "font-weight:700;color:var(--primary)" if m==mes else "color:var(--muted)"
        meses_nav += f'<a href="/sueldos?mes={m}&anio={anio}" style="text-decoration:none;font-size:.82rem;padding:4px 8px;border-radius:4px;{sel}">{["","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"][m]}</a>'

    years_nav = ""
    for y in [anio-1, anio, anio+1]:
        sel = "background:var(--primary);color:#fff" if y==anio else "background:var(--bg);color:var(--text)"
        years_nav += f'<a href="/sueldos?mes={mes}&anio={y}" style="text-decoration:none;font-size:.82rem;padding:3px 10px;border-radius:4px;{sel}">{y}</a>'

    # Stats
    total = len(clientes_con_emp)
    rec_ok = sum(1 for cid,*_ in clientes_con_emp if estados.get(cid,("","",""))[0]=="presentado")
    f931_ok = sum(1 for cid,*_ in clientes_con_emp if estados.get(cid,("","",""))[1]=="presentado")
    vep_ok = sum(1 for cid,*_ in clientes_con_emp if estados.get(cid,("","",""))[2]=="generado")

    # Badge helper
    def badge_est(val, tipo):
        if tipo == "recibo":
            opts = [("pendiente","⏳ Pendiente","#888"),("borrador","📝 Borrador","#E67E22"),
                    ("presentado","✅ Presentado","#27AE60")]
        elif tipo == "931":
            opts = [("pendiente","⏳ Pendiente","#888"),("borrador","📝 Borrador","#E67E22"),
                    ("presentado","✅ Presentado","#27AE60")]
        else:  # vep
            opts = [("pendiente","⏳ Pendiente","#888"),("generado","💳 Generado","#1A5276"),
                    ("pagado","✅ Pagado","#27AE60")]
        for k,lbl,col in opts:
            if val==k:
                return f'<span style="background:{col};color:#fff;font-size:.68rem;padding:2px 7px;border-radius:4px;font-weight:600;white-space:nowrap">{lbl}</span>'
        return ""

    # Build rows
    filas = ""
    mes_nombre = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                  "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes]

    for cid, cnombre, ccuit_enc, n_emp in clientes_con_emp:
        ccuit = dec(ccuit_enc) if ccuit_enc else ""
        cuit_limpio = (ccuit or "").replace("-","").replace(" ","")
        est = estados.get(cid, ("pendiente","pendiente","pendiente",""))
        est_rec, est_931, est_vep = est[0], est[1], est[2]
        obs_val = est[3] if len(est)>3 else ""
        vto = vto_931(ccuit, mes, anio)

        # Row color by status
        if est_rec=="presentado" and est_931=="presentado" and est_vep in ("generado","pagado"):
            row_bg = "rgba(39,174,96,.06)"
        elif est_rec=="pendiente" and est_931=="pendiente":
            row_bg = "rgba(231,76,60,.04)"
        else:
            row_bg = "transparent"

        # Select options
        def sel_rec(v): return "selected" if est_rec==v else ""
        def sel_931(v): return "selected" if est_931==v else ""
        def sel_vep(v): return "selected" if est_vep==v else ""

        filas += f'''<tr style="background:{row_bg};border-bottom:1px solid var(--border)">
          <td style="padding:8px 10px">
            <div style="font-weight:600;font-size:.85rem">{cnombre}</div>
            <div style="font-size:.7rem;color:var(--muted)">{n_emp} emp · CUIT: {ccuit or "—"}</div>
          </td>
          <td style="padding:8px 6px;text-align:center">
            <form method="post" style="display:inline" onchange="this.submit()">
              <input type=hidden name=accion value=guardar>
              <input type=hidden name=cliente_id value={cid}>
              <input type=hidden name=estado_931 value="{est_931}">
              <input type=hidden name=estado_vep value="{est_vep}">
              <input type=hidden name=observaciones value="{obs_val}">
              <select name="estado_recibo" style="font-size:.75rem;padding:3px 5px;border:1.5px solid var(--border);border-radius:6px;background:var(--bg)">
                <option value="pendiente" {sel_rec("pendiente")}>⏳ Pendiente</option>
                <option value="borrador" {sel_rec("borrador")}>📝 Borrador</option>
                <option value="presentado" {sel_rec("presentado")}>✅ Presentado</option>
              </select>
            </form>
          </td>
          <td style="padding:8px 6px;text-align:center">
            <form method="post" style="display:inline" onchange="this.submit()">
              <input type=hidden name=accion value=guardar>
              <input type=hidden name=cliente_id value={cid}>
              <input type=hidden name=estado_recibo value="{est_rec}">
              <input type=hidden name=estado_vep value="{est_vep}">
              <input type=hidden name=observaciones value="{obs_val}">
              <select name="estado_931" style="font-size:.75rem;padding:3px 5px;border:1.5px solid var(--border);border-radius:6px;background:var(--bg)">
                <option value="pendiente" {sel_931("pendiente")}>⏳ Pendiente</option>
                <option value="borrador" {sel_931("borrador")}>📝 Borrador</option>
                <option value="presentado" {sel_931("presentado")}>✅ Presentado</option>
              </select>
            </form>
            <div style="font-size:.65rem;color:var(--muted);margin-top:2px">Vto: {vto}</div>
          </td>
          <td style="padding:8px 6px;text-align:center">
            <form method="post" style="display:inline" onchange="this.submit()">
              <input type=hidden name=accion value=guardar>
              <input type=hidden name=cliente_id value={cid}>
              <input type=hidden name=estado_recibo value="{est_rec}">
              <input type=hidden name=estado_931 value="{est_931}">
              <input type=hidden name=observaciones value="{obs_val}">
              <select name="estado_vep" style="font-size:.75rem;padding:3px 5px;border:1.5px solid var(--border);border-radius:6px;background:var(--bg)">
                <option value="pendiente" {sel_vep("pendiente")}>⏳ Pendiente</option>
                <option value="generado" {sel_vep("generado")}>💳 Generado</option>
                <option value="pagado" {sel_vep("pagado")}>✅ Pagado</option>
              </select>
            </form>
          </td>
          <td style="padding:8px 6px">
            <a href="/empleados/{cid}" class="btn btn-xs btn-o" title="Ver empleados">👷 {n_emp}</a>
          </td>
        </tr>'''

    empty = '<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:24px">No hay clientes con empleados activos registrados</td></tr>'

    body = f"""
    <h1 class="page-title">Control Sueldos y F.931</h1>
    <p class="page-sub">Solo clientes con empleados registrados</p>
    {flash}

    <!-- Navegacion mes/año -->
    <div class="fcard" style="margin-bottom:14px;padding:10px 16px">
      <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap">
        <div style="display:flex;gap:4px">{years_nav}</div>
        <div style="display:flex;gap:2px;flex-wrap:wrap">{meses_nav}</div>
        <div style="margin-left:auto;font-size:.85rem;font-weight:700;color:var(--primary)">{mes_nombre} {anio}</div>
      </div>
    </div>

    <!-- Stats -->
    <div class="stats" style="margin-bottom:16px">
      <div class="scard b"><div class="sicon">👷</div><div class="slabel">Clientes con empleados</div><div class="sval">{total}</div></div>
      <div class="scard g"><div class="sicon">📄</div><div class="slabel">Recibos presentados</div><div class="sval">{rec_ok}/{total}</div></div>
      <div class="scard o"><div class="sicon">📋</div><div class="slabel">F.931 presentados</div><div class="sval">{f931_ok}/{total}</div></div>
      <div class="scard p" style="--scard-c:var(--info)"><div class="sicon">💳</div><div class="slabel">VEP generados</div><div class="sval">{vep_ok}/{total}</div></div>
    </div>

    <!-- Info -->
    <div class="info-box" style="margin-bottom:14px;font-size:.8rem">
      💡 El estado se guarda automáticamente al cambiar el selector. El vencimiento F.931 se calcula según terminación del CUIT del cliente.
    </div>

    <!-- Tabla -->
    <div class="fcard" style="overflow-x:auto;padding:0">
      <table style="width:100%;border-collapse:collapse;font-size:.84rem">
        <thead>
          <tr style="background:var(--primary);color:#fff">
            <th style="padding:10px 12px;text-align:left">Cliente</th>
            <th style="padding:10px 8px;text-align:center;min-width:130px">Recibo sueldo</th>
            <th style="padding:10px 8px;text-align:center;min-width:140px">F.931 / Sindicato</th>
            <th style="padding:10px 8px;text-align:center;min-width:120px">VEP pago</th>
            <th style="padding:10px 8px;text-align:center">Empleados</th>
          </tr>
        </thead>
        <tbody>
          {filas if clientes_con_emp else empty}
        </tbody>
      </table>
    </div>
    <style>
    .scard.p {{ border-left:4px solid var(--info) !important }}
    </style>
    """
    return page("Sueldos y F.931", body, "Clientes")


# ══════════════════════════════════════════════════════════════════════════════
#  NOVEDADES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/novedades")
@login_req
def novedades():
    rol = session.get("rol","secretaria")
    script = (
        "<script>"
        "fetch('https://dolarapi.com/v1/dolares/oficial').then(r=>r.json()).then(d=>{"
        "var f=function(n){return '$'+n.toLocaleString('es-AR',{minimumFractionDigits:2})};"
        "document.getElementById('nc').textContent=f(d.compra);"
        "document.getElementById('nv').textContent=f(d.venta);"
        "var od=document.getElementById('idx-oficial');if(od)od.textContent=f(d.venta);"
        "}).catch(()=>{});"
        "fetch('https://dolarapi.com/v1/dolares/tarjeta').then(r=>r.json()).then(d=>{"
        "var f=function(n){return '$'+n.toLocaleString('es-AR',{minimumFractionDigits:2})};"
        "var el=document.getElementById('nd');if(el)el.textContent=f(d.venta);"
        "}).catch(()=>{});"
        "fetch('https://dolarapi.com/v1/dolares/bolsa').then(r=>r.json()).then(d=>{"
        "var el=document.getElementById('idx-mep');if(el)el.textContent='$'+d.venta.toLocaleString('es-AR',{maximumFractionDigits:0});"
        "}).catch(()=>{});"
        "fetch('https://dolarapi.com/v1/dolares/blue').then(r=>r.json()).then(d=>{"
        "var el=document.getElementById('idx-blue');if(el)el.textContent='$'+d.venta.toLocaleString('es-AR',{maximumFractionDigits:0});"
        "}).catch(()=>{});"
        "fetch('https://apis.datos.gob.ar/series/api/series/?ids=143.3_SALM_DICI_0_36_6&limit=1&sort=desc&format=json')"
        ".then(r=>r.json()).then(d=>{"
        "if(d.data&&d.data.length){var v=parseFloat(d.data[0][1]),f=function(n){return '$'+Math.round(n).toLocaleString('es-AR')};"
        "var el=document.getElementById('smvym-val');if(el)el.textContent=f(v);"
        "var el2=document.getElementById('idx-smvym');if(el2)el2.textContent=f(v);}"
        "}).catch(()=>{});"
        "function tickN(){var n=new Date(),pad=function(x){return x.toString().padStart(2,'0')};"
        "var dias=['Dom','Lun','Mar','Mie','Jue','Vie','Sab'];"
        "var rj=document.getElementById('nreloj');var fh=document.getElementById('nfecha');"
        "if(rj)rj.textContent=pad(n.getHours())+':'+pad(n.getMinutes())+':'+pad(n.getSeconds());"
        "if(fh)fh.textContent=dias[n.getDay()]+' '+pad(n.getDate())+'/'+(pad(n.getMonth()+1))+'/'+n.getFullYear();}"
        "tickN();setInterval(tickN,1000);"
        "function showTab(id,btn){"
        "document.querySelectorAll('.tabpanel').forEach(function(p){p.classList.remove('on')});"
        "document.querySelectorAll('.tab').forEach(function(b){b.classList.remove('on')});"
        "document.getElementById(id).classList.add('on');btn.classList.add('on');}"
        "</script>"
        "<style>@media(max-width:700px){.twocol{grid-template-columns:1fr!important}}</style>"
    )

    bna_bar = (
        '<div style="background:var(--card);border-radius:var(--r);padding:14px 20px;'
        'box-shadow:var(--shadow);margin-bottom:18px;display:flex;align-items:center;'
        'justify-content:space-between;flex-wrap:wrap;gap:10px">'
        '<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:center">'
        '<div><div style="font-size:.62rem;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:3px">Dolar BNA Oficial</div>'
        '<div style="display:flex;gap:14px;align-items:baseline">'
        '<div><span style="font-size:.65rem;color:var(--muted)">Compra</span> '
        '<span id="nc" style="font-weight:700;font-size:1.1rem;color:var(--success)">---</span></div>'
        '<div><span style="font-size:.65rem;color:var(--muted)">Venta</span> '
        '<span id="nv" style="font-weight:700;font-size:1.1rem;color:var(--danger)">---</span></div>'
        '<div><span style="font-size:.65rem;color:var(--muted)">Divisa/Tarjeta</span> '
        '<span id="nd" style="font-weight:700;font-size:1rem;color:var(--warning)">---</span></div>'
        '</div></div></div>'
        '<div style="text-align:right">'
        '<div id="nreloj" style="font-family:serif;font-size:1.2rem;color:var(--primary);font-weight:600"></div>'
        '<div id="nfecha" style="font-size:.74rem;color:var(--muted)"></div>'
        '</div></div>'
    )

    tabs = (
        '<div class="tabs">'
        '<button class="tab on" onclick="showTab(\'t-imp\',this)">Impositivo</button>'
        '<button class="tab" onclick="showTab(\'t-sal\',this)">Escalas Salariales</button>'
        '<button class="tab" onclick="showTab(\'t-srt\',this)">SRT / ART</button>'
        '<button class="tab" onclick="showTab(\'t-yt\',this)">Holistor</button>'
        '</div>'
    )

    t_imp = (
        '<div id="t-imp" class="tabpanel on">'

        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px" class="twocol">'

        '<div class="fcard" style="margin-bottom:0"><h3>ARCA / AFIP — Novedades</h3>'
        '<div style="display:flex;flex-direction:column;gap:8px">'
        '<a href="https://servicioscf.afip.gob.ar/publico/sitio/contenido/novedad/listado.aspx" target="_blank" class="btn btn-arca btn-sm">Novedades AFIP/ARCA (oficial)</a>'
        '<a href="https://www.arca.gob.ar/vencimientos/" target="_blank" class="btn btn-o btn-sm">Calendario Vencimientos ARCA</a>'
        '<a href="https://www.argentina.gob.ar/trabajo/bo" target="_blank" class="btn btn-o btn-sm">Boletin Oficial</a>'
        '<a href="https://contadoresenred.com/art-suma-fija-marzo-2026-nuevo-importe-respecto-del-regimen-general/" target="_blank" class="btn btn-o btn-sm">Contadores en Red — Novedades</a>'
        '<a href="https://www.instagram.com/p/DVGNz5tDB6o/" target="_blank" class="btn btn-o btn-sm">Instagram novedades impositivas</a>'
        '<a href="https://documento.errepar.com/actualidad/fondo-fiduciario-de-enfermedades-profesionales-nueva-suma-fija-con-vencimiento-en-mayo-2026-20260428091549513" target="_blank" class="btn btn-o btn-sm">Errepar — Fondo Enf. Profesionales</a>'
        '<a href="https://siap.blogdelcontador.com.ar/novedades/nuevo-valor-de-la-suma-fija-del-fondo-fiduciario-de-enfermedades-profesionales-desde-marzo-2026/" target="_blank" class="btn btn-o btn-sm">Blog Contador — Suma fija FFEP</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>Links impositivos utiles</h3>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.arca.gob.ar/landing/default.asp" target="_blank" class="btn btn-arca btn-sm">ARCA Login</a>'
        '<a href="https://seti.afip.gob.ar/padron-puc-constancia-internet/ConsultaConstanciaAction.do" target="_blank" class="btn btn-o btn-sm">Constancia Inscripcion</a>'
        '<a href="https://www.arca.gob.ar/monotributo/categorias.asp" target="_blank" class="btn btn-o btn-sm">Categorias Monotributo</a>'
        '<a href="http://dgronline.dgrsantiago.gob.ar" target="_blank" class="btn btn-o btn-sm">Rentas Santiago del Estero</a>'
        '<a href="http://www.dgrsantiago.gov.ar/?page_id=992" target="_blank" class="btn btn-o btn-sm">Biblioteca IIBB Rentas SGO</a>'
        '<a href="https://servicioscorp.anses.gob.ar/clavelogon/logon.aspx?system=miansesv2" target="_blank" class="btn btn-o btn-sm">ANSES</a>'
        '<a href="https://www.bcra.gob.ar" target="_blank" class="btn btn-b btn-sm">BCRA</a>'
        '<a href="https://www.ilovepdf.com/es/jpg_a_pdf" target="_blank" class="btn btn-o btn-sm">iLovePDF</a>'
        '</div></div>'

        '</div>'

        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px" class="twocol">'

        '<div class="fcard" style="margin-bottom:0"><h3>CPCESE — Honorarios y Cursos</h3>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://cpcese.org.ar/matriculados/honorarios-minimos-eticos" target="_blank" class="btn btn-a btn-sm">Honorarios Minimos Eticos</a>'
        '<a href="https://cpcese.org.ar/documentos/contadores%20afiche%2001-12-25.pdf" target="_blank" class="btn btn-o btn-sm">Ultima actualizacion (Dic 2025)</a>'
        '<a href="https://autogestion.cpcese.org.ar/materiales" target="_blank" class="btn btn-o btn-sm">Cursos y materiales CPCESE</a>'
        '<a href="https://www.ambito.com/edicion-impresa/hacienda-carnes-y-la-factura-electronica-n3969539" target="_blank" class="btn btn-o btn-sm">Liquidaciones Hacienda</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>Mercados e Inversiones</h3>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.rava.com/" target="_blank" class="btn btn-o btn-sm">RAVA — Bolsa y Mercados</a>'
        '<a href="https://rofex.primary.ventures/fyo/futurosfinancieros" target="_blank" class="btn btn-o btn-sm">ROFEX — Futuros Financieros</a>'
        '</div></div>'

        '</div>'

        '<div class="fcard"><h3>Indices del dia</h3>'
        '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:10px">'
        '<div class="scard" style="margin-bottom:0"><div class="slabel">Dolar Oficial</div><div class="sval" id="idx-oficial">---</div></div>'
        '<div class="scard b" style="margin-bottom:0"><div class="slabel">Dolar MEP</div><div class="sval" id="idx-mep">---</div></div>'
        '<div class="scard o" style="margin-bottom:0"><div class="slabel">Dolar Blue</div><div class="sval" id="idx-blue">---</div></div>'
        '<div class="scard g" style="margin-bottom:0"><div class="slabel">SMVYM</div><div class="sval" id="idx-smvym">---</div></div>'
        '</div></div>'

        '</div>'
    )

    t_sal = (
        '<div id="t-sal" class="tabpanel">'
        '<div class="warn-box" style="margin-bottom:14px">Las escalas se actualizan por paritaria. Verificar siempre en la fuente oficial antes de liquidar.</div>'
        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px" class="twocol">'

        '<div class="fcard" style="margin-bottom:0"><h3>Transporte de Cargas (Camioneros/FADEEAC)</h3>'
        '<div style="font-size:.8rem;color:var(--muted);margin-bottom:10px">Convenio 40/89</div>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.camioneros-ba.org.ar/index.php/gremiales/salarios/escalas-salariales" target="_blank" class="btn btn-o btn-sm">Escalas Camioneros/Transporte</a>'
        '<a href="https://federacion.impresiondeboletas.com.ar/login.aspx" target="_blank" class="btn btn-o btn-sm">Sindicato Camioneros Nacional</a>'
        '<a href="http://camioneros.sirwiq.com/Login/Acceso" target="_blank" class="btn btn-o btn-sm">Camioneros Sgo. del Estero</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>Empleados de Comercio (FAECYS)</h3>'
        '<div style="font-size:.8rem;color:var(--muted);margin-bottom:10px">Convenio 130/75</div>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://jorgevega.com.ar/laboral/71-empleados-comercio-escala-salarial-2016-2017.html" target="_blank" class="btn btn-o btn-sm">Escalas Empleados Comercio</a>'
        '<a href="https://www.online.faecys.org.ar/Inicio.aspx" target="_blank" class="btn btn-o btn-sm">FAECYS Nacional</a>'
        '<a href="https://www.sasweb.com.ar/usuarios/login" target="_blank" class="btn btn-o btn-sm">Sindicato Comercio Sgo. Estero</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>Trabajadores Agropecuarios (UATRE)</h3>'
        '<div style="font-size:.8rem;color:var(--muted);margin-bottom:10px">Convenio 1/75 — Personal rural</div>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.ignacioonline.com.ar/paritaria-agrarios-escalas-salariales-marzo-abril-y-junio-2026-uatre/" target="_blank" class="btn btn-o btn-sm">Escalas Agropecuarios 2026 (UATRE)</a>'
        '<a href="https://portal.renatre.org.ar/" target="_blank" class="btn btn-o btn-sm">RENATRE Nacional</a>'
        '<a href="https://apps.uatre.org.ar/usupe/signin.aspx" target="_blank" class="btn btn-o btn-sm">UATRE Santiago del Estero</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>SMVyM — Salario Minimo</h3>'
        '<div style="font-size:.8rem;color:var(--muted);margin-bottom:10px">Fijado por el CNEPSMVYM</div>'
        '<div id="smvym-box" style="background:var(--bg);border-radius:8px;padding:10px;margin-bottom:10px">'
        '<div style="font-size:.72rem;font-weight:700;color:var(--muted);text-transform:uppercase;margin-bottom:4px">Valor actual</div>'
        '<div id="smvym-val" style="font-size:1.3rem;color:var(--success)">cargando...</div>'
        '</div>'
        '<a href="https://www.argentina.gob.ar/trabajo/smvm" target="_blank" class="btn btn-g btn-sm">Ver SMVYM oficial</a>'
        '</div>'

        '</div></div>'
    )

    t_srt = (
        '<div id="t-srt" class="tabpanel">'
        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px" class="twocol">'

        '<div class="fcard" style="margin-bottom:0"><h3>SRT — Suma fija y actualizaciones</h3>'
        '<div style="font-size:.8rem;color:var(--muted);margin-bottom:12px">Actualizacion trimestral por resolucion SRT. Verificar monto vigente antes de liquidar.</div>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.srt.gob.ar" target="_blank" class="btn btn-o btn-sm">SRT Oficial</a>'
        '<a href="https://www.srt.gob.ar/index.shtml/montosfijos" target="_blank" class="btn btn-o btn-sm">Montos fijos vigentes SRT</a>'
        '<a href="https://siap.blogdelcontador.com.ar/novedades/nuevo-valor-de-la-suma-fija-del-fondo-fiduciario-de-enfermedades-profesionales-desde-marzo-2026/" target="_blank" class="btn btn-o btn-sm">Nueva suma fija FFEP (Mar 2026)</a>'
        '<a href="https://documento.errepar.com/actualidad/fondo-fiduciario-de-enfermedades-profesionales-nueva-suma-fija-con-vencimiento-en-mayo-2026-20260428091549513" target="_blank" class="btn btn-o btn-sm">FFEP — Vencimiento Mayo 2026</a>'
        '<a href="https://servicios.srt.gob.ar/srt/emision/index.xhtml" target="_blank" class="btn btn-o btn-sm">Sistema SRT Emision</a>'
        '<a href="https://www.srt.gob.ar/index.shtml/siniestros" target="_blank" class="btn btn-o btn-sm">Denuncia Siniestros</a>'
        '</div></div>'

        '<div class="fcard" style="margin-bottom:0"><h3>Higiene y Seguridad Laboral</h3>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.argentina.gob.ar/trabajo/seguridadysalud" target="_blank" class="btn btn-b btn-sm">Seguridad e Higiene MTSS</a>'
        '<a href="https://www.srt.gob.ar/estadisticas/" target="_blank" class="btn btn-o btn-sm">Estadisticas SRT</a>'
        '<a href="https://www.argentina.gob.ar/trabajo" target="_blank" class="btn btn-o btn-sm">Ministerio de Trabajo</a>'
        '<a href="https://contadoresenred.com/art-suma-fija-marzo-2026-nuevo-importe-respecto-del-regimen-general/" target="_blank" class="btn btn-o btn-sm">ART — Suma fija vigente</a>'
        '</div></div>'

        '</div></div>'
    )

    t_yt = (
        '<div id="t-yt" class="tabpanel">'
        '<div class="fcard">'
        '<h3>Canal Holistor — Tutoriales y novedades del sistema</h3>'
        '<p style="color:var(--muted);font-size:.83rem;margin-bottom:16px">Videos oficiales de capacitacion, liquidacion de sueldos y novedades impositivas.</p>'
        '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px;margin-bottom:14px">'

        '<div style="background:var(--bg);border-radius:10px;padding:16px">'
        '<div style="font-weight:700;color:var(--primary);margin-bottom:8px;font-size:1rem">Canal oficial Holistor</div>'
        '<p style="font-size:.83rem;color:var(--muted);margin-bottom:14px;line-height:1.6">'
        'Tutoriales de liquidacion de sueldos, manejo del sistema Holistor, '
        'novedades impositivas y actualizaciones para estudios contables.</p>'
        '<a href="https://www.youtube.com/@HolistorSA/videos" target="_blank" '
        'class="btn btn-sm" style="background:#FF0000;color:#fff;width:100%;justify-content:center;display:flex;margin-bottom:8px">'
        'Ver canal completo en YouTube</a>'
        '<a href="https://www.youtube.com/@HolistorSA/videos" target="_blank" class="btn btn-o btn-sm" style="width:100%;justify-content:center;display:flex">'
        'Ultimos videos subidos</a>'
        '</div>'

        '<div style="background:var(--bg);border-radius:10px;padding:16px">'
        '<div style="font-weight:700;color:var(--primary);margin-bottom:8px">Buscar videos por tema</div>'
        '<div class="info-box" style="margin-bottom:10px">Busca en YouTube escribiendo el tema + HolistorSA</div>'
        '<div style="display:flex;flex-direction:column;gap:7px">'
        '<a href="https://www.youtube.com/results?search_query=HolistorSA+liquidacion+sueldos" target="_blank" class="btn btn-o btn-sm">Liquidacion de sueldos</a>'
        '<a href="https://www.youtube.com/results?search_query=HolistorSA+novedades+impositivas" target="_blank" class="btn btn-o btn-sm">Novedades impositivas</a>'
        '<a href="https://www.youtube.com/results?search_query=HolistorSA+tutorial" target="_blank" class="btn btn-o btn-sm">Tutoriales del sistema</a>'
        '<a href="https://www.youtube.com/results?search_query=HolistorSA+ART+SRT" target="_blank" class="btn btn-o btn-sm">ART / SRT</a>'
        '</div></div>'

        '</div></div></div>'
    )

    body = (
        '<h1 class="page-title">Novedades</h1>'
        '<p class="page-sub">Actualizaciones impositivas, escalas salariales y capacitaciones</p>'
        + bna_bar + tabs + t_imp + t_sal + t_srt + t_yt + script
    )
    return page("Novedades", body, "Novedades")


@app.route("/api/cierres_por_mes")
@admin_req
def api_cierres_por_mes():
    conn=conectar();c=conn.cursor()
    c.execute("""SELECT SUBSTRING(fecha,7,4)||'/'||SUBSTRING(fecha,4,2) as mes_anio,
                        SUM(total_general)
                 FROM cierres_caja WHERE cerrado=TRUE
                 GROUP BY mes_anio ORDER BY SUBSTRING(fecha,7,4),SUBSTRING(fecha,4,2) DESC LIMIT 12""")
    rows=list(reversed(c.fetchall())); conn.close()
    from flask import jsonify
    return jsonify({"labels":[r[0] for r in rows],"totales":[float(r[1] or 0) for r in rows]})

if __name__=="__main__":
    app.run(debug=True)
